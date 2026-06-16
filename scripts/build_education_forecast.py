# -*- coding: utf-8 -*-
"""
build_education_forecast.py
===========================
Extracts Yotam Tzabari's per-חומש education forecast from his Gonenim tracking
workbook and emits a clean JSON the app joins against live plan status.

Source : Downloads/טבלת מעקב גוננים (יותם צברי).xlsx  ->  sheet "היצע שירותי חינוך "
         (header row 6, data from row 7; ~193 rows = קיים + עתידי).
Output : oranim-app/data/education_forecast_gonenim.json

The sheet is self-contained: every education facility carries, on one row, its
type, class counts, the תב"ע it belongs to (col 21 מספר תכנית), the חומש class
distribution (cols 35-37 כיתות למימוש חומש 1/2/3) and ITM coords (cols 43/44).

taba is normalised to match plans.geojson / all_permits.json keys, which store the
bare number with no "101-" prefix and no leading zeros (e.g. "101-0110510" -> "110510").

Run:  py scripts/build_education_forecast.py
"""
import json
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
APP_ROOT = os.path.dirname(HERE)
SRC_XLSX = os.path.join(
    os.path.expanduser("~"), "Downloads", "טבלת מעקב גוננים (יותם צברי).xlsx"
)
SHEET = "היצע שירותי חינוך "  # note trailing space — exact sheet name
OUT = os.path.join(APP_ROOT, "data", "education_forecast_gonenim.json")

# 1-based column -> 0-based tuple index
C = lambda n: n - 1

# Manual taba links for facilities the projector left unlinked (col 21 blank) but a
# statutory plan now exists. Keyed by (name, neighborhood). Yotam's table is ~1y old,
# so plans entered since (e.g. בבדיקה תכנונית) are missing from it.
MANUAL_TABA = {
    ("עתידי בית הנוער", "רסקו"): "1158146",  # "בית הנוער העברי" (בבדיקה תכנונית)
}


def norm_tabas(v):
    """Yotam 'מספר תכנית' -> list of plan keys matching plans.geojson `taba`.
    '101-0565317'      -> ['565317']
    '101-0565317 / 101-0252379' -> ['565317','252379']
    'המלצת הצוות' / '' -> []   (recommendation, not tied to a plan)
    """
    s = str(v or "")
    out, seen = [], set()
    for tok in re.split(r"[,;/\\\s]+", s):
        tok = tok.strip()
        if "-" in tok:
            tok = tok.split("-")[-1]
        tok = tok.lstrip("0")
        if tok.isdigit() and len(tok) >= 4 and tok not in seen:
            seen.add(tok)
            out.append(tok)
    return out


def norm_one_taba(v):
    """Single bare plan key from any taba-ish cell ('101-0515635' / '4938ב' -> '515635'/'4938')."""
    s = str(v or "")
    if "-" in s:
        s = s.split("-")[-1]
    m = re.search(r"\d+", s)
    if not m:
        return None
    t = m.group(0).lstrip("0") or "0"
    return t if len(t) >= 4 else None


# ── Demand (צורך) per domain per חומש — from "מאזנים בסיס הנתונים" ──────────
# Rollup service-type labels per domain (regular + special-ed); breakdown rows
# (ממלכתי/ממ"ד/חרדי/ערבי) are excluded to avoid double-counting.
DEMAND_SHEET = "מאזנים בסיס הנתונים"
DEMAND_NEIGHBORHOODS = {"רסקו", "גוננים א-ו", "פת", "קטמונים ח-ט"}
# Regular education stream only — matches Yotam's per-domain balance rows.
# שירותי חינוך מיוחד is a SEPARATE service in his table (not folded into the domain),
# so including "X חינוך מיוחד" here would inflate demand and break the מאזן match.
DOMAIN_ROLLUPS = {
    "maon": {"מעון יום"},
    "gan": {"גן ילדים"},
    "yesodi": {"יסודי"},
    "al_yesodi": {"על יסודי"},
}


# ── Area context (population + units per חומש per neighborhood) ────────────
# From "דמוגרפיה-קיים ועתידי " — row 5 = cluster total (כלל_עירוני), rows 7-10 = neighborhoods.
# cols: 3 יחד קיימות · 4 תושבים · 5 משק בית · 6/7/8 תוספת אוכ' חומש1/2/3 · 11/12/13 אוכ' מצטבר חומש1/2/3
CONTEXT_SHEET = "דמוגרפיה-קיים ועתידי "


def extract_context(wb):
    if CONTEXT_SHEET not in wb.sheetnames:
        print("  WARNING: demographics sheet not found, skipping context")
        return {}
    ws = wb[CONTEXT_SHEET]
    rows = list(ws.iter_rows(min_row=1, max_row=12, values_only=True))

    def mk(r):
        units_ex = num(r[C(3)]) or 0.0
        residents = num(r[C(4)]) or 0.0
        hh = num(r[C(5)]) or 0.0
        out = {"units_existing": round(units_ex), "residents_now": round(residents)}
        for n in (1, 2, 3):
            added_pop = num(r[C(5 + n)]) or 0.0
            cumul = num(r[C(10 + n)]) or 0.0
            out["c%d" % n] = {
                "pop_total": round(cumul),
                "pop_added": round(added_pop),
                "units_added": round(added_pop / hh) if hh else None,
            }
        return out

    ctx = {}
    for r in rows:
        nm = txt(r[C(1)])
        if nm in DEMAND_NEIGHBORHOODS:
            ctx[nm] = mk(r)
        elif nm == "כלל_עירוני":
            ctx["__all__"] = mk(r)
    return ctx


def extract_demand(wb):
    """Per domain per scope: existing supply/demand + per-חומש ADDED supply & demand.
    {scope: {domain: {supply_existing, demand_existing,
                      add_supply:[c1,c2,c3], add_demand:[c1,c2,c3]}}}
    היצע קיים=col7 · ביקוש קיים=col12 · תוספת היצע ח1/2/3=col8/9/10 · תוספת ביקוש ח1/2/3=col13/14/15."""
    ws = wb[DEMAND_SHEET]

    def blank():
        return {d: {"supply_existing": 0.0, "demand_existing": 0.0,
                    "add_supply": [0.0, 0.0, 0.0], "add_demand": [0.0, 0.0, 0.0]} for d in DOMAIN_ROLLUPS}

    scopes = {"__all__": blank()}  # __all__ = cluster total; plus one per neighborhood
    type_to_domain = {}
    for dom, labels in DOMAIN_ROLLUPS.items():
        for lb in labels:
            type_to_domain[lb] = dom
    for r in ws.iter_rows(min_row=6, values_only=True):
        nb = txt(r[C(2)])
        if nb not in DEMAND_NEIGHBORHOODS:
            continue
        dom = type_to_domain.get(txt(r[C(6)]))
        if not dom:
            continue
        if nb not in scopes:
            scopes[nb] = blank()
        sup = num(r[C(7)]) or 0.0    # היצע קיים (existing supply)
        dem = num(r[C(12)]) or 0.0   # ביקוש קיים (existing demand)
        as_ = [num(r[C(c)]) or 0.0 for c in (8, 9, 10)]    # תוספת היצע ח1/2/3
        ad = [num(r[C(c)]) or 0.0 for c in (13, 14, 15)]   # תוספת ביקוש ח1/2/3
        for tgt in (scopes["__all__"], scopes[nb]):
            tgt[dom]["supply_existing"] += sup
            tgt[dom]["demand_existing"] += dem
            for i in range(3):
                tgt[dom]["add_supply"][i] += as_[i]
                tgt[dom]["add_demand"][i] += ad[i]
    for sc in scopes.values():
        for d in sc:
            sc[d]["supply_existing"] = round(sc[d]["supply_existing"], 1)
            sc[d]["demand_existing"] = round(sc[d]["demand_existing"], 1)
            sc[d]["add_supply"] = [round(x, 1) for x in sc[d]["add_supply"]]
            sc[d]["add_demand"] = [round(x, 1) for x in sc[d]["add_demand"]]
    return scopes


# ── מנח"י realization year/status — from "טבלה ראשית " ─────────────────────
MANHI_XLSX = os.path.join(os.path.expanduser("~"), "Downloads",
                          "טבלת מעקב תכניות 24-11 מעודכן למנחי - 24-12-25.xlsx")
MANHI_SHEET = "טבלה ראשית "


def parse_manhi_year(v):
    """col 26 שנת אכלוס -> (year:int|None, text:str|None)."""
    if v is None:
        return None, None
    if hasattr(v, "year"):  # datetime
        return v.year, None
    if isinstance(v, (int, float)):
        y = int(v)
        return (y, None) if 1990 <= y <= 2100 else (None, str(v))
    s = str(v).strip()
    m = re.search(r"\b(20\d{2})\b", s)
    if m:
        return int(m.group(1)), None
    return None, (s or None)


def load_manhi():
    """{ taba -> {year:int|None, year_text:str|None, status:str|None, function:str|None} }
    Prefers the row carrying a numeric year when a taba appears more than once."""
    if not os.path.exists(MANHI_XLSX):
        print("  WARNING: מנח\"י file not found, skipping realization join")
        return {}
    wb = openpyxl.load_workbook(MANHI_XLSX, data_only=True, read_only=True)
    ws = wb[MANHI_SHEET]
    out = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        taba = norm_one_taba(r[C(14)])  # col 14 תב"ע/מגרש
        if not taba:
            continue
        year, year_text = parse_manhi_year(r[C(26)])  # col 26 שנת אכלוס
        rec = {"year": year, "year_text": year_text,
               "status": txt(r[C(23)]), "function": txt(r[C(7)])}
        if taba not in out or (rec["year"] and not out[taba]["year"]):
            out[taba] = rec
    return out


def num(v):
    """Coerce a cell to float, else None. Treats blanks / text as None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    s = str(v).strip() if v is not None else ""
    return s if s and s != "#N/A" else None


def main():
    if not os.path.exists(SRC_XLSX):
        sys.exit("Source workbook not found: %s" % SRC_XLSX)
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True, read_only=True)
    ws = wb[SHEET]
    demand = extract_demand(wb)
    context = extract_context(wb)
    manhi = load_manhi()

    facilities = []
    counts = {"קיים": 0, "עתידי": 0, "other": 0}
    matched_taba = 0
    manhi_hits = 0
    for r in ws.iter_rows(min_row=7, values_only=True):
        status = txt(r[C(4)])
        name = txt(r[C(6)])
        if not status and not name:
            continue  # blank spacer row
        tabas = norm_tabas(r[C(21)])
        if not tabas:
            ov = MANUAL_TABA.get((name, txt(r[C(2)])))
            if ov:
                tabas = [ov]
        if tabas:
            matched_taba += 1
        cls_c1 = num(r[C(35)]) or 0.0
        cls_c2 = num(r[C(36)]) or 0.0
        cls_c3 = num(r[C(37)]) or 0.0
        # Primary חומש: explicit col 30 if it names one, else the earliest חומש
        # that carries class-realisation weight.
        chumash_raw = txt(r[C(30)])
        primary = None
        m = re.search(r"חומש\s*([123])", chumash_raw or "")
        if m:
            primary = int(m.group(1))
        elif cls_c1 > 0:
            primary = 1
        elif cls_c2 > 0:
            primary = 2
        elif cls_c3 > 0:
            primary = 3

        fac = {
            "id": txt(r[C(1)]),
            "name": name,
            "neighborhood": txt(r[C(2)]),          # שכונה / מתחם
            "cluster": txt(r[C(3)]),               # רובע / אשכול
            "status": status,                      # קיים / עתידי
            "area_type": txt(r[C(5)]),             # סוג השטח
            "semel_iri": txt(r[C(7)]),
            "semel_moe": txt(r[C(8)]),             # סמל משרד החינוך
            "address": txt(r[C(9)]),
            "edu_type": txt(r[C(10)]),             # סוג חינוך
            "edu_stage": txt(r[C(11)]),            # שלב חינוך (גן/יסודי/על-יסודי/מעון)
            "pikuach": txt(r[C(12)]),
            "service": txt(r[C(13)]),
            "classes_existing": num(r[C(18)]),     # כיתות לפרוגרמה מצב קיים
            "classes_delta": num(r[C(19)]),        # תוספת/גריעת כיתות
            "classes_potential": num(r[C(20)]),    # סה"כ כיתות פוטנציאליות
            "taba": tabas[0] if tabas else None,
            "tabas": tabas,
            "plan_name_yotam": txt(r[C(22)]),
            "est_year": txt(r[C(29)]),             # שנת אכלוס מוערכת
            "chumash": chumash_raw,
            "chumash_primary": primary,
            "mimush_c1": num(r[C(31)]),            # אחוז מימוש חומש 1 (הנחת יותם)
            "mimush_c2": num(r[C(32)]),
            "mimush_c3": num(r[C(33)]),
            "classes_c1": cls_c1,                  # כיתות למימוש חומש 1
            "classes_c2": cls_c2,
            "classes_c3": cls_c3,
            "classes_mimush_total": num(r[C(38)]),
            "classes_remaining_full": num(r[C(39)]),
            "migrash": txt(r[C(40)]),
            "gush": txt(r[C(41)]),
            "helka": txt(r[C(42)]),
            "x_itm": num(r[C(43)]),
            "y_itm": num(r[C(44)]),
            "notes": txt(r[C(45)]),
        }
        # מנח"י realization year/status (actual municipal build tracking), by taba
        mrec = None
        for t in tabas:
            if t in manhi:
                mrec = manhi[t]
                break
        if mrec:
            manhi_hits += 1
            fac["manhi_year"] = mrec["year"]
            fac["manhi_year_text"] = mrec["year_text"]
            fac["manhi_status"] = mrec["status"]
            fac["manhi_function"] = mrec["function"]
        else:
            fac["manhi_year"] = None
            fac["manhi_year_text"] = None
            fac["manhi_status"] = None
            fac["manhi_function"] = None
        facilities.append(fac)
        counts[status if status in counts else "other"] += 1

    payload = {
        "source": "טבלת מעקב גוננים (יותם צברי).xlsx / היצע שירותי חינוך",
        "sheet": SHEET.strip(),
        "count": len(facilities),
        "demand": demand,
        "context": context,
        "facilities": facilities,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    print("Wrote %s" % OUT)
    print("  facilities: %d  (קיים=%d  עתידי=%d  other=%d)"
          % (len(facilities), counts["קיים"], counts["עתידי"], counts["other"]))
    print("  rows with a taba link: %d" % matched_taba)
    fut_no_taba = sum(1 for f in facilities if f["status"] == "עתידי" and not f["tabas"])
    print("  future facilities without a plan (ללא תכנית): %d" % fut_no_taba)
    print("  facilities matched to מנח\"י: %d" % manhi_hits)
    print("  demand block: %s (scopes: %d)" % ("OK" if demand and demand.get("__all__") else "MISSING", len(demand)))


if __name__ == "__main__":
    main()
