# -*- coding: utf-8 -*-
"""Ad-hoc: does every xlsx projector recommendation appear spatially?

Replicates build_projector_gonenim's row-reading + pid logic, then matches each
xlsx row against the emitted projector_gonenim.geojson and reports the
geometry_source quality (real geometry vs. centroid-only vs. missing).
"""
import io, json, sys
from collections import Counter, defaultdict
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import build_projector_gonenim as B
import openpyxl

# --- load xlsx rows exactly like the builder ---
wb = openpyxl.load_workbook(B.XLSX_PATH, data_only=True)
ws = wb[B.SHEET]
header = [c.value for c in ws[2]]
h = {name: idx for idx, name in enumerate(header)}

rows = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if all(c is None for c in row):
        continue
    if row[h["אזור"]] not in B.GONENIM_AREAS:
        continue
    rows.append(row)

print(f"xlsx rows in 4 gonenim sub-neighborhoods: {len(rows)}")

# --- load emitted geojson, index by base project_id ---
gj = json.load(open(B.DATA_DIRS[0] / "projector_gonenim.geojson", encoding="utf-8"))
feats = gj["features"]
# Also load the tzatal (official transport shapefile) layer — transport &
# passage recommendations are emitted there, not into projector_gonenim.
tz = json.load(open(B.DATA_DIRS[0] / "projector_gonenim_tzatal.geojson", encoding="utf-8"))
tz_feats = tz["features"]
# A feature pid may be "<base>" or "<base>-1","<base>-2" for multi-part geoms.
feat_by_base = defaultdict(list)
for f in feats + tz_feats:
    pid = (f["properties"].get("project_id") or "")
    base = pid.split("-")[0] if pid else ""
    feat_by_base[base].append(f)
print(f"tzatal (transport) features: {len(tz_feats)}")

CENTROID_SOURCES = {"sub_neighborhood_centroid", "minahak_centroid"}

cat = Counter()
real_rows, centroid_rows, skipped_rows, missing_rows, dup_rows = [], [], [], [], []
seen = set()

for row in rows:
    area = row[h["אזור"]]
    proj_num = row[h["מס פרויקט"]]
    proj_name = row[h["שם פרויקט"]]
    orig_name = str(proj_name or "").strip()
    name_ovr = B.MANUAL_NAME_OVERRIDES.get((orig_name, area, str(proj_num or "").strip()))
    if name_ovr:
        proj_name = name_ovr
    service_he = (row[h["שירות"]] or "").strip()
    domain_he = row[h["תחום"]]
    name = str(proj_name or "").strip()

    pid = B.project_id(area, proj_num, service_he, name)
    rec = {"area": area, "num": proj_num, "name": orig_name,
           "service": service_he, "domain": domain_he}

    if pid in seen:
        cat["duplicate_merged"] += 1
        dup_rows.append(rec)
        continue
    seen.add(pid)

    if ((name, area) in B.SKIP_FEATURES
            or (name, area, service_he) in B.SKIP_FEATURES):
        cat["skipped_intentionally"] += 1
        skipped_rows.append(rec)
        continue

    matches = feat_by_base.get(pid, [])
    if not matches:
        cat["MISSING"] += 1
        missing_rows.append(rec)
        continue

    srcs = {m["properties"].get("geometry_source") for m in matches}
    if srcs & CENTROID_SOURCES and not (srcs - CENTROID_SOURCES):
        cat["centroid_only"] += 1
        rec["src"] = ",".join(sorted(s for s in srcs if s))
        centroid_rows.append(rec)
    else:
        cat["real_geometry"] += 1
        real_rows.append(rec)

print("\n=== coverage breakdown ===")
for k, v in cat.most_common():
    print(f"  {k}: {v}")

def dump(title, recs):
    print(f"\n=== {title} ({len(recs)}) ===")
    for r in recs:
        extra = f"  [{r.get('src')}]" if r.get("src") else ""
        print(f"  [{r['area']}] #{r['num']} {r['name']} — {r['service']} ({r['domain']}){extra}")

dump("MISSING — no feature at all", missing_rows)
dump("CENTROID-ONLY — appears but not at a real spot", centroid_rows)
dump("skipped intentionally (SKIP_FEATURES)", skipped_rows)
dump("duplicate-merged (same area+num+service+name)", dup_rows)

# geometry_source distribution across all emitted features
print("\n=== geometry_source across all emitted features ===")
gs = Counter(f["properties"].get("geometry_source") for f in feats)
for k, v in gs.most_common():
    print(f"  {k}: {v}")
