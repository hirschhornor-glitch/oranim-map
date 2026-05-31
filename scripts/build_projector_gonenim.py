"""
Build projector_gonenim.geojson + summary.json + overlap.json from the
"אורנים — תכנית עבודה מתכללת (מהדורה 4)" Excel.

Reads 131 גוננים rows, geocodes each via plan_number / migrash / centroid,
detects overlap with planned תב"עות, extracts the relevant PDF pages
(via PyMuPDF) so popups can show the consultants' description verbatim,
and writes outputs to both oranim-map/data/ and oranim-app/data/.

Run manually each time the xlsx is updated.
"""
from __future__ import annotations
import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

try:
    import fitz  # PyMuPDF — used for PDF text extraction
except ImportError:
    fitz = None

try:
    from shapely.geometry import shape, mapping, Point as ShPoint
    from shapely.ops import unary_union, nearest_points as _shp_nearest_points
except ImportError:
    shape = mapping = unary_union = ShPoint = _shp_nearest_points = None

try:
    from pyproj import Transformer as _PyProjTransformer
    _ITM_TRANSFORMER = _PyProjTransformer.from_crs(4326, 2039, always_xy=True)
except ImportError:
    _ITM_TRANSFORMER = None

try:
    from scipy.interpolate import RBFInterpolator as _RBFInterpolator
except ImportError:
    _RBFInterpolator = None

try:
    import numpy as _np
except ImportError:
    _np = None

# --- Paths ----------------------------------------------------------------
XLSX_PATH = Path(
    r"C:\Users\orlev\OneDrive - Municipality of Jerusalem"
    r"\פרוייקטור שכונתי תוצרים"
    r"\אורנים - תכנית עבודה מתכללת לכלל השכונות  (מהדורה 4).xlsx"
)
SHEET = "תכנית עבודה"
# מינה"ק גוננים includes 4 sub-neighborhoods: רסקו, גוננים א-ו, קטמונים ח-ט, פת
GONENIM_AREAS = {"רסקו", "גוננים א-ו", "קטמונים ח-ט", "פת"}

REPO_ROOT = Path(r"C:\ORANIM")
# The repo holds a single app (oranim-app). Older versions wrote to a sibling
# oranim-map/data — that directory no longer exists.
DATA_DIRS = [REPO_ROOT / "oranim-app" / "data"]

LANDUSE_PATH = DATA_DIRS[0] / "landuse_xplan.geojson"
PLANS_PATH = DATA_DIRS[0] / "plans.geojson"
FUTURE_SHAVAZ_PATH = DATA_DIRS[0] / "future_shavaz.geojson"
MINAHAK_PATH = DATA_DIRS[0] / "minahak_gonen.geojson"
SUB_NEIGH_PATH = DATA_DIRS[0] / "sub_neighborhoods.geojson"
MINAHAK_UNION_PATH = DATA_DIRS[0] / "minahak_gonenim_union.geojson"
MIGRASH_PANUI_PATH = DATA_DIRS[0] / "migrash_panui.geojson"
BUILDINGS_PATH = DATA_DIRS[0] / "buildings.geojson"
ROADS_PATH = DATA_DIRS[0] / "roads.geojson"
GEOCODE_CACHE_PATH = DATA_DIRS[0] / "projector_geocode_cache.json"
PARCELS_PATH = DATA_DIRS[0] / "parcels_gonenim.geojson"
YIUD_KAYAM_PATH = DATA_DIRS[0] / "yiud_karka_kayam.geojson"
# Official projector shapefiles (transport recommendations) — exact LineString/Point
# geometries provided by the projector team. Cover all 4 sub-neighborhoods.
PROJECTOR_OFFICIAL_PATHS = [
    REPO_ROOT / "oranim-app" / "data" / "projector_official_layer3.geojson",  # 53 polylines (richer attrs — checked first)
    REPO_ROOT / "oranim-app" / "data" / "projector_official_layer1.geojson",  # 53 polylines (basic attrs)
    REPO_ROOT / "oranim-app" / "data" / "projector_official_layer2.geojson",  # 18 bus-stop points
]
# Sub-neighborhood name mapping: xlsx 'אזור' → shp 'שכונה'
AREA_TO_SHP_SUB = {
    "גוננים א-ו": "א-ו",
    "קטמונים ח-ט": "ח-ט",
    "פת": "פת",
    "רסקו": "רסקו",
}

# Map xlsx area name → sub_neighborhoods.geojson `schn_nama` (for centroid fallback)
AREA_TO_SUBNEIGH = {
    "רסקו": "רסקו - גבעת הורדים",
    "גוננים א-ו": "גוננים",
    "קטמונים ח-ט": "קטמונים",
    "פת": "פת",
}

# Features to skip entirely (don't emit to the geojson). Keyed by (project_name, area).
# Use for recommendations the planner decided to drop from the map.
# Features that qualify as is_eivuy_brown or is_shatzap_polygon but the user
# wants rendered as POINT markers instead of polygons (e.g. multiple services
# share a single lot, so a polygon would be ambiguous). Keyed by (name, area).
FORCE_POINT_FEATURES = {
    # User-flagged 2026-05-26: should be point, not 2.76-dunam polygon.
    ("טשרניחובסקי 52-54-55", "רסקו"),
    # User-flagged 2026-05-26: should be point, not 2.85-dunam polygon.
    ("התקופה 2-שמעוני 48-הרצוג", "רסקו"),
    # User-flagged 2026-05-26: should be point — מעון יום inside the larger
    # מתחם המשתלה compound (shared lot with יונתן זקס school).
    ("מעון יום במתחם המשתלה", "גוננים א-ו"),
    # User-flagged 2026-05-26: programa services (gan, me'on, beit knesset)
    # within plan 101-0987016 should appear as הפרשה (point), not the full plan
    # polygon.
    ("סן מרטין 1", "קטמונים ח-ט"),
}

# Per-compound categorization of inner שצ"פים. Keyed by (project_name, area) →
# dict of {(centroid_lon, centroid_lat) → {category, label}}. The matcher picks
# the NEAREST override centroid within ~60m. Categories:
#   "מוצלח"          — successful, no upgrade needed
#   "לשדרוג"          — upgrade only
#   "לשדרוג+תצפית"   — upgrade + add a viewpoint
INNER_SHATZAP_CATEGORIES = {
    # מתחם בן זכאי-אנטיגונוס-אליעזר הגדול — page 18 PDF:
    # 5 שצ"פים labeled with categories + sizes. The 6th (450 m²) is unlabeled
    # in PDF — defaults to "מוצלח". Centroids match yiud_karka_kayam polygons.
    ("מתחם בן זכאי-אנטיגונוס-אליעזר הגדול", "גוננים א-ו"): {
        (35.20676, 31.75682): {"category": "לשדרוג+תצפית", "label": "1,350 מ\"ר"},
        (35.20641, 31.75645): {"category": "לשדרוג",        "label": "1,100 מ\"ר"},
        (35.20602, 31.75609): {"category": "מוצלח",         "label": None},
        (35.20557, 31.75580): {"category": "מוצלח",         "label": None},
        (35.20531, 31.75536): {"category": "לשדרוג+תצפית", "label": "1,200 מ\"ר"},
    },
    # מתחם הפסגה (בית שוויץ) — page 11 PDF (מצגת קטמונים מתחם הפסגה 05.2024):
    # גן הנוטרים (5.8 דונם) → "הגדלת שטח הפארק כדי שיוכל לשאת את נפח התושבים
    # ההולך וגדל ולהוסיף לו פרוגרמות" + שילוב שב"צ ב-הפרשה מבונה במגורים החדשים.
    ("מתחם הפסגה (בית שוויץ)", "קטמונים ח-ט"): {
        (35.19627, 31.75513): {"category": "לשדרוג+הרחבה", "label": "5.8 דונם → להרחבה"},
    },
}

# Extra features not in xlsx — crosswalks from PDF tables that the xlsx
# skipped (e.g. feasible-but-not-prioritized rows). Each entry creates a
# new feature in the geojson at build time. Coords are user-provided.
# Keyed by synthetic project_id; value contains all needed fields.
EXTRA_FEATURES = {
    # === גוננים א-ו crosswalks missing from xlsx (5 feasible rows) ===
    # Per "גוננים מרחב ציבורי 05.2024.pdf" עמ' 12 — these 5 of 14 feasible
    # crosswalks didn't make it to the xlsx. Coords user-provided.
    "extra_gonenim_cw_4": {
        "geometry": {"type": "Point", "coordinates": [35.198876, 31.757731]},
        "properties": {
            "project_name": "חננאל",
            "project_num": "ממעבר חציה מס'4 (לא ב-xlsx)",
            "service_he": "מעבר חציה",
            "sub_neighborhood": "גוננים א-ו",
            "service_key": "crosswalk",
            "description": "מעבר חציה מומלץ עפ\"י יועצת התנועה (לא נכלל בתכנית העבודה)",
            "plan_status": "המלצת יועצת",
            "is_xlsx_gap": True,
            "crosswalk_table_row": {
                "main_street": "חננאל",
                "secondary_street": "מעבר לה\"ר",
                "distance": "אין מעברים קרובים",
                "justification": "מעבר המשכי לה\"ר",
                "type": "מעבר חציה מוגבה",
                "implications": "ביטול 2 חניות",
                "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 — שורה 4 (יועצת תנועה: יורוברידג'). לא ב-xlsx.",
            },
        },
    },
    "extra_gonenim_cw_7": {
        "geometry": {"type": "Point", "coordinates": [35.199747, 31.755081]},
        "properties": {
            "project_name": "ניתאי הארבלי",
            "project_num": "ממעבר חציה מס'7 (לא ב-xlsx)",
            "service_he": "מעבר חציה",
            "sub_neighborhood": "גוננים א-ו",
            "service_key": "crosswalk",
            "description": "מעבר חציה מומלץ עפ\"י יועצת התנועה (לא נכלל בתכנית העבודה)",
            "plan_status": "המלצת יועצת",
            "is_xlsx_gap": True,
            "crosswalk_table_row": {
                "main_street": "ניתאי הארבלי",
                "secondary_street": "מעבר לה\"ר מונגש (רח' ללא מוצא)",
                "distance": "אין מעברים קרובים",
                "justification": "מעבר המשכי מומלץ לה\"ר",
                "type": "מומלץ גם להסדיר את המדרכה עם הרחוב ללא מוצא כך שכ\"ר יכנסו דרך אבן עליה לרכב.",
                "implications": "ביטול 2 מקומות חנייה.",
                "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 — שורה 7 (יועצת תנועה: יורוברידג'). לא ב-xlsx.",
            },
        },
    },
    "extra_gonenim_cw_8": {
        "geometry": {"type": "Point", "coordinates": [35.201968, 31.752337]},
        "properties": {
            "project_name": "קנאי הגליל",
            "project_num": "ממעבר חציה מס'8 (לא ב-xlsx)",
            "service_he": "מעבר חציה",
            "sub_neighborhood": "גוננים א-ו",
            "service_key": "crosswalk",
            "description": "מעבר חציה מומלץ עפ\"י יועצת התנועה (לא נכלל בתכנית העבודה)",
            "plan_status": "המלצת יועצת",
            "is_xlsx_gap": True,
            "crosswalk_table_row": {
                "main_street": "קנאי הגליל",
                "secondary_street": "רחוב ללא שם",
                "distance": "100 / 190",
                "justification": "מעבר המשכי לה\"ר, וצומת.",
                "type": "מעבר חציה מערבית לצומת, כולל הסדרת צומת.",
                "implications": "ללא",
                "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 — שורה 8 (יועצת תנועה: יורוברידג'). לא ב-xlsx.",
            },
        },
    },
    "extra_gonenim_cw_11": {
        "geometry": {"type": "Point", "coordinates": [35.208194, 31.756187]},
        "properties": {
            "project_name": "יוסי בן יועזר",
            "project_num": "ממעבר חציה מס'11 (לא ב-xlsx)",
            "service_he": "מעבר חציה",
            "sub_neighborhood": "גוננים א-ו",
            "service_key": "crosswalk",
            "description": "מעבר חציה מומלץ עפ\"י יועצת התנועה (לא נכלל בתכנית העבודה)",
            "plan_status": "המלצת יועצת",
            "is_xlsx_gap": True,
            "crosswalk_table_row": {
                "main_street": "יוסי בן יועזר",
                "secondary_street": "מעבר לה\"ר",
                "distance": "אין מעברים קרובים",
                "justification": "מעבר המשכי לה\"ר",
                "type": "מעבר חציה מתומרר שלא בצומת",
                "implications": "ביטול כ-5 חניות",
                "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 — שורה 11 (יועצת תנועה: יורוברידג'). לא ב-xlsx.",
            },
        },
    },
    "extra_gonenim_cw_15": {
        "geometry": {"type": "Point", "coordinates": [35.210431, 31.758637]},
        "properties": {
            "project_name": "יוסי בן יועזר",
            "project_num": "ממעבר חציה מס'15 (לא ב-xlsx)",
            "service_he": "מעבר חציה",
            "sub_neighborhood": "גוננים א-ו",
            "service_key": "crosswalk",
            "description": "מעבר חציה מומלץ עפ\"י יועצת התנועה (לא נכלל בתכנית העבודה)",
            "plan_status": "המלצת יועצת",
            "is_xlsx_gap": True,
            "crosswalk_table_row": {
                "main_street": "יוסי בן יועזר",
                "secondary_street": "מעבר לה\"ר",
                "distance": "50 / 50",
                "justification": "מעבר לה\"ר",
                "type": "במידה ויתוכננו מעברים לה\"ר בין שני צידי הכביש, יש לתכנון גם מעבר חציה, דיוק מיקום המעבר צריך לקחת בחשבון את חציית הכביש, יש לתאם בין אדריכל הנוף ליועץ התנועה.",
                "implications": "ביטול 4-6 חניות, מתוכן 1 חניות נכים.",
                "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 — שורה 15 (יועצת תנועה: יורוברידג'). לא ב-xlsx.",
            },
        },
    },
}

# Per-compound design principles (extracted from PDF deep-dive booklets like
# "מצגת קטמונים מתחם הפסגה"). Rendered in popup as a structured list.
# Keyed by (project_name, area) → list of {label, items[], source}.
# Per-crosswalk table data from "רסקו מרחב ציבורי 17.4.24.pdf" עמ' 8 — table of
# 7 recommended crosswalks analyzed by יורוברידג'. Each row maps to a feature
# via (area, project_num). Row 3 (דוד שמעוני - רח' ללא מוצא) is "לא מומלץ" and
# has no feature.
CROSSWALK_TABLE_DATA = {
    ("רסקו", "מעבר חציה מס' 1"): {
        "main_street": "הרב חיים ברלין",
        "secondary_street": "הרב ברודי",
        "distance": "אין מעברי חציה קרובים",
        "justification": "מעבר המשכי לה\"ר, צומת לא מוסדר",
        "type": "הסדרת צומת כלל מעברי חציה",
        "implications": "ביטול חניות בקרבת הצומת",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    ("רסקו", "מעבר חציה מס' 2"): {
        "main_street": "הטייסים",
        "secondary_street": "מעבר לה\"ר",
        "distance": "100 / 160",
        "justification": "",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ביטול 4-5 חניות",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    ("רסקו", "מעבר חציה מס' 4"): {
        "main_street": "טשרניחובסקי",
        "secondary_street": "מעבר לה\"ר",
        "distance": "120 / 200",
        "justification": "",
        "type": "מעבר חציה מוגבה לה\"ר",
        "implications": "ביטול חניה או שתיים",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    ("רסקו", "מעבר חציה מס' 5"): {
        "main_street": "דוד שמעוני",
        "secondary_street": "מעבר לה\"ר",
        "distance": "115 / 200",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מוגבה",
        "implications": "מתוכנן ברח' שמעוני בעת הפיכתו לחד סטרי (קראוס חן)",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    ("רסקו", "מעבר חציה מס' 6"): {
        "main_street": "ש\"י עגנון",
        "secondary_street": "המעפילים",
        "distance": "140 / 170",
        "justification": "מעבר מבוקש לה\"ר",
        "type": "",
        "implications": "מתוכנן ע\"י רולי פלד",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    ("רסקו", "מעבר חציה מס' 7"): {
        "main_street": "טשרניחובסקי — חד-סטרי מצפון לדרום",
        "secondary_street": "מעבר לה\"ר",
        "distance": "110",
        "justification": "מעבר המשכי לה\"ר מומלץ",
        "type": "אזור החצייה, ללא סימון 811",
        "implications": "ביטול 4 מקומות חניה",
        "source": "רסקו מרחב ציבורי 17.4.24, עמ' 8 (יועצת תנועה: יורוברידג')",
    },
    # === גוננים א-ו — מ-"גוננים מרחב ציבורי 05.2024.pdf" עמ' 12 (16 רשומות, 14 ישימים) ===
    # פיצ'רים #72-#80 (9 מתוך 16). הקידומת בxlsx היא "ממעבר חציה מס'".
    # NOTE: simple rule applies — implication is in the השלכות column of the
    # row itself, same as Rasko. (Earlier shift-rule was a parsing mistake.)
    # User-confirmed via PDF p12 image 2026-05-28.
    ("גוננים א-ו", "ממעבר חציה מס' 1"): {
        "main_street": "ניקנור",
        "secondary_street": "מעבר לה\"ר",
        "distance": "50 / 150",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ביטול כ-5 חניות",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס' 2"): {
        "main_street": "ניקנור",
        "secondary_street": "מעבר לה\"ר",
        "distance": "80 / 160",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ביטול 4-6 חניות, מתוכן 5 חניות נכים, סיכוי נמוך עד אפסי לביצוע.",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'5"): {  # NOTE: xlsx has no space after מס'
        "main_street": "ניקנור",
        "secondary_street": "מעבר לה\"ר",
        "distance": "60",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "ביטול 2 חניות, אין צורך לבטל את חניית רכב הנכה",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'9"): {
        "main_street": "אלעזר בן יאיר",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין מעברים קרובים",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מוגבה",
        "implications": "ביטול כ-5 חניות",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'10"): {
        "main_street": "רשב\"ג",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין מעברים קרובים",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "יש לתכנן כך שלא יבוטלו חניות.",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'12"): {
        "main_street": "אנטיגונוס",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין מעברים קרובים",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ביטול כ-5 חניות",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'13"): {
        "main_street": "מעגלי יבנה",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין מעברים קרובים",
        "justification": "מעבר המשכי לה\"ר",
        "type": "",  # PDF p12 image: empty type cell for row 13
        "implications": "מתוכננת במסגרת מעגלי יבנה",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'14"): {
        "main_street": "רשב\"ג",
        "secondary_street": "רשב\"ג / יוסי בן יועזר",
        "distance": "צמוד לצומת",
        "justification": "מעבר המשכי לה\"ר",
        "type": "צמוד לצומת נוסף, אין אפשרות לתכנן מעבר בשלב זה. כאשר יתוכנן מתחמית (בהתאם להמלצת הפרויקטור) אזי יש לשלב בתכנון את המעבר, באזור החיבור לאמוראים.",
        "implications": "",  # PDF p12: empty implications cell for row 14
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    ("גוננים א-ו", "ממעבר חציה מס'16"): {
        "main_street": "אנטיגונוס",
        "secondary_street": "רחובות ללא מוצא",
        "distance": "50",
        "justification": "מעבר לה\"ר",
        "type": "לתכנן כמעבר חציה בצומת, ובמידה ויתוכנן המתחם שמבטל את כניסת הרכבים לרח' ללא מוצא אנטיגונוס (בהתאם להמלצת הפרויקטור) אזי יש לשלב בתכנון את המעבר.",
        "implications": "ביטול 2 מקומות חנייה.",
        "source": "גוננים מרחב ציבורי 05.2024, עמ' 12 (יועצת תנועה: יורוברידג')",
    },
    # === קטמונים ח-ט — מ-"קטמונים ופת 05.2024.pdf" עמ' 16 (12 רשומות, 4 ישימים) ===
    # פיצ'רים #168-#171 (4 מתוך 12). הקידומת בxlsx היא "מעבר חציה מס' " (כמו רסקו).
    # שאר 8 השורות (3,4,5,7,8,9,10,12) אין להן פיצ'ר ב-xlsx.
    # NOTE: unlike the יורוברידג' tables (גוננים), this PDF's implications column
    # is correctly aligned to each row — NO 1-row shift here. Source: ברידג'.
    ("קטמונים ח-ט", "מעבר חציה מס' 1"): {
        "main_street": "דב הוז",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "ביטול 5-6 חניות",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 2"): {
        "main_street": "אלמליח",
        "secondary_street": "מעבר לה\"ר",
        "distance": "60 / ללא",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ללא",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 3"): {
        "main_street": "הנוטרים",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "ביטול 2 חניות",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 4"): {
        "main_street": "הנוטרים",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה, יש להרחיק מהסיבוב ככל שניתן",
        "implications": "ביטול 2 חניות",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 5"): {
        "main_street": "השומר",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "ביטול 5-6 חניות",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 6"): {
        "main_street": "יצחק שדה",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "ביטול 5-6 חניות",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 7"): {
        "main_street": "בר יוחאי",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חצייה מוגבה",
        "implications": "בצמוד לאזור החצייה מתוכננים פרויקטים שונים, אשר יתקדמו בלוחות זמנים נפרדים. יש לוודא כי המעבר יתוכנן בכל התכניות, גם אם מיקומו ישתנה מעט בהתאם.",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 8"): {
        # PDF row 8: main_street=בר יוחאי (xlsx project_name=סן מרטין — kept as the key)
        "main_street": "בר יוחאי",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "משולב בתכנית התחדשות עירונית בר יוחאי 5-15",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 9"): {
        "main_street": "בר יוחאי / סן מרטין",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "",
        "implications": "כיום לא קיים מעבר חציה בצומת, בהסדרת הצומת בעתיד יש לוודא תכנון מעבר חציה כחלק מהסדר הצומת. ללא.",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 10"): {
        "main_street": "חיבור לדוד איילון / בר יוחאי",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "",
        "implications": "כיום לא קיים מעבר חציה בצומת, בהסדרת הצומת בעתיד יש לוודא תכנון מעבר חציה כחלק מהסדר הצומת. ללא.",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    ("קטמונים ח-ט", "מעבר חציה מס' 11"): {
        "main_street": "אלמליח",
        "secondary_street": "מעבר לה\"ר / חניון",
        "distance": "100 / ללא",
        "justification": "מעבר המשכי לה\"ר 144-145",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "תוכנן צומת עם חניון, במסגרת דוד אילון, יש להוסיף מעבר חציה כחלק מההסדר בצומת, בשלב זה או בשלב תכניות התחדשות עירונית בר יוחאי 17-21/אלמליח. ללא.",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
    # === פת — מ-"קטמונים ופת 05.2024.pdf" עמ' 16 (row 12, רק רשומה אחת בxlsx) ===
    ("פת", "מעבר חציה מס' 12"): {
        "main_street": "ברל לוקר",
        "secondary_street": "מעבר לה\"ר",
        "distance": "אין בקרבתו",
        "justification": "מעבר המשכי לה\"ר",
        "type": "מעבר חציה מתומרר שלא בצומת",
        "implications": "ביטול כ-6 חניות. מעבר החציה ימוקם בהתאם לתכנית התחדשות עירונית הקרובה, ומיקום תחנת האוטובוס במתכונתה הסופית.",
        "source": "קטמונים ופת 05.2024, עמ' 16 (יועצת תנועה: ברידג')",
    },
}

COMPOUND_PRINCIPLES = {
    ("מתחם הפסגה (בית שוויץ)", "קטמונים ח-ט"): {
        "source": "מצגת מתחם הפסגה — עמ' 11 (סיכום המלצות לאחר הצגה לעירייה)",
        "headline": "פינוי בינוי של כל המתחם",
        "items": [
            "כל המתחם יתפס כמרחב תכנוני אחד",
            "מענה לשב\"צים במסגרת הפרשות מבונות במגורים החדשים",
            "הגדלת שטח הפארק (גן הנוטרים) להוספת פרוגרמות לנפח התושבים הגדל",
            "שימור עצים ערכיים במגרש בית שוויץ (חורשת האורנים)",
            "הקצאת שטח לכיכר עירונית ולרחבה מסחרית (ניתן לאחד לרחבה אחת)",
            "הגדלת רוחב רצועת הדרך ושינוי חתך הרחוב (השומר/נוטרים)",
        ],
    },
}

# Per-compound extra trail features (e.g. the yellow path inside the בן זכאי
# compound shown on PDF page 18). Keyed by (project_name, area) → list of trails.
INNER_TRAILS_BY_COMPOUND = {
    ("מתחם בן זכאי-אנטיגונוס-אליעזר הגדול", "גוננים א-ו"): [
        # Main statutory trail spine — N-S ~310m (yellow line on PDF page 18)
        {
            "label": "הסדרה סטטוטורית של השביל",
            "kind": "main_trail",
            "geometry": {
                "type": "LineString",
                "coordinates": [
                    [35.204691, 31.755225],
                    [35.205957, 31.756260],
                    [35.206965, 31.757232],
                ],
            },
        },
        # 3 accessibility passage upgrades (orange arrows on PDF p18). One of
        # these two will be accessibility-upgraded as a connector to the
        # purple light-rail line (per PDF: "אחד משני המעברים יונגש במסגרת
        # מצרנים לקו הסגול").
        {
            "label": "הסדרת שביל דרומי",
            "kind": "accessibility_trail",
            "geometry": {
                "type": "LineString",
                "coordinates": [[35.205088, 31.755539], [35.204514, 31.755649]],
            },
        },
        {
            "label": "הסדרת שביל אמצעי",
            "kind": "accessibility_trail",
            "geometry": {
                "type": "LineString",
                "coordinates": [[35.205485, 31.755877], [35.205002, 31.756105]],
            },
        },
        {
            "label": "הסדרת שביל שלישי",
            "kind": "accessibility_trail",
            "geometry": {
                "type": "LineString",
                "coordinates": [[35.206343, 31.756662], [35.206000, 31.756994]],
            },
        },
    ],
}

SKIP_FEATURES = {
    # User-removed 2026-05-20: too generic to map (whole-neighborhood transit route).
    ("מסלול תח\"צ שכונת רסקו", "רסקו"),
    # User-merged 2026-05-26: actually describes the same passage as
    # "מעבר ה\"ר מס' 7" (מתחם לוריא); attached as a note to that feature.
    ("שימור מעבר קיים טשרניחובסקי 48", "רסקו"),
    # User-merged 2026-05-26: same physical road as the "מתחם מכוור/הורקניה"
    # road_segment (just a different recommendation #); attached as a note.
    ("הארכת דרך רח' מכוור שכונת גוננים", "גוננים א-ו"),
    # User-merged 2026-05-26: same passage as "שדרוג והנגשה מעבר ה\"ר בין דב הוז
    # לגולומב" (project #2, service "הנגשה"); attached as a note.
    ("דב הוז לגולומב", "קטמונים ח-ט"),
    # User-merged 2026-05-26: transport aspect of בית הנוער (the 8th feature) —
    # merged as a note onto the other 7 services. 3-tuple key, service-specific.
    ("בית הנוער", "רסקו", "שיפור רמת שירות"),
}

# Extra note text to APPEND to a feature's popup. Keyed by (project_name, area).
# Used when an xlsx row is redundant/merged with another feature but its note
# text should still be surfaced on the kept feature.
# Bidirectional cross-feature links. Keyed by (name, area) → list of related
# (name, area) tuples. Resolved at build time to {project_id, name, display_point}
# entries on the feature's `related_features` property; the frontend renders them
# as clickable "ראי גם" badges that pan the map to the linked feature.
RELATED_FEATURES = {
    # The southern boundary of חוות הנוער's 3 proposed passages is the same passage
    # as מבני בתירא לחיים יחיל (project #15 in גוננים). Link both directions.
    ("מבני בתירא לחיים יחיל", "גוננים א-ו"): [("חוות הנוער", "רסקו")],
    ("חוות הנוער", "רסקו"): [("מבני בתירא לחיים יחיל", "גוננים א-ו")],
}

MANUAL_NOTES = {
    ("מתחם לוריא", "רסקו"): (
        "הערה (משימור מעבר קיים טשרניחובסקי 48): "
        "הצגת מעבר ה\"ר בשטח מגרש טשרניחובסקי 48 במסגרת תב\"ע 15791, "
        "בהתאם למעבר הקיים הפופולרי. הוספת המעבר למסמכי התב\"ע."
    ),
    ("מבני בתירא לחיים יחיל", "גוננים א-ו"): (
        "הערה (מחוות הנוער): זהו הגבול הדרומי של 3 ההצעות למעברי הולכי רגל "
        "בשטח חוות הנוער הציוני (פרויקט מעברי ה\"ר מס' 11,12,13 ברסקו, "
        "גוש/חלקה 30122/124 ו-30006/132,133)."
    ),
    ("חוות הנוער", "רסקו"): (
        "הערה: הפיצ'ר מציג את הקטע הצפוני (מעבר 11) ואת הקטע המרכזי "
        "(מעבר 12, שני קווים נפרדים). הקטע הדרומי (מעבר 13) מוצג בנפרד "
        "כ-\"מבני בתירא לחיים יחיל\" (מעבר ה\"ר מס' 15 בגוננים)."
    ),
    ("בית הנוער", "רסקו"): (
        "היבט תנועתי (תושי\"ה, חומש 1, 2030, רסקו כבישים ותנועה שקופית 12): "
        "המתחם ממוקם בסמוך לתחנת רכבת קלה; עירוב שימושים טוב מאפשר הקטנת חניה (~100 מ\"ח). "
        "כניסה ויציאת רכב מרחוב הרצוג לא מעמיסה על השכונה. "
        "מפרצי העלאה/הורדה: דרך השירות של הרצוג + רח' התקופה. "
        "מומלץ לשקול גם ברח' ניקנור מול מעבר ה\"ר המוביל למתחם (~1000 תלמידי תיכון)."
    ),
    ("מתחם מכוור/הורקניה", "גוננים א-ו", "רחוב חדש"): (
        "הערה (מהארכת דרך רח' מכוור שכונת גוננים — המלצה 8): "
        "חיבור רח' מכור ללא מוצא לבני בתירא. שיפור רמת שירות, חומש 1 (2030)."
    ),
    ("שדרוג והנגשה מעבר ה\"ר בין דב הוז לגולומב", "קטמונים ח-ט"): (
        "הערה (מ-\"דב הוז לגולומב\" מעבר ה\"ר מס' 2): "
        "חיבור דב הוז לתחנת האוטובוס. בתב\"ע 0987016 מתאפשר מעבר בין הבניינים. "
        "יש לוודא רציפות המעבר לכשתגיע יוזמת התחדשות בחלקה מס' 11 הסמוכה. "
        "גוש/חלקה 30169/11. הוספת מעבר ה\"ר, חומש 3 (2040)."
    ),
}

# Features the planner couldn't locate precisely YET — kept at fallback (centroid)
# until ITM/CAD coords arrive. Format: (project_name, area[, service]) → "note".
# Each build prints how many pending items remain.
PENDING_REVIEW = {
}

# Hard overrides for features whose xlsx `קישורים` (PDF reference) is wrong or
# points to the wrong page. Keyed by (project_name, area) → str. Applied AFTER
# reading the xlsx, so the popup shows the corrected reference.
MANUAL_PDF_OVERRIDES = {
    # User-corrected 2026-05-26: xlsx said pages 9,26; actual recommendation is page 25.
    ("מאלעזר בין יאיר לקנאי הגליל", "גוננים א-ו"): "עמ' 25 חוברת מרחב ציבורי גוננים",
    # User-corrected 2026-05-26: xlsx said page 22; actual page is 21.
    ("חיבור בין השכונה לפארק המסילה", "פת"): "עמ' 21 חוברת מרחב ציבורי קטמונים ופת",
}

# Fix mislabeled project_name in xlsx where one row was copied with the wrong
# name. Keyed by (xlsx_name, area, project_num). Applied very early, before any
# other name-based lookup runs (overrides, notes, force-point, skip).
MANUAL_NAME_OVERRIDES = {
    # User-flagged 2026-05-26: pn=10 had project_name "3 גני ילדים ברחוב גוננים
    # (משרה מלכה, מנח"י)" but its description / plan / location all describe a
    # different project — "מתחם קנאי הגליל (מחסן תרופות)".
    ("3 גני ילדים ברחוב גוננים (משרה מלכה, מנח\"י)", "גוננים א-ו", "10"): "מתחם קנאי הגליל (מחסן תרופות)",
}

# Hard overrides for features whose xlsx `מזהה מיקום` points to a wrong/stale
# parcel. Keyed by (project_name, area) → [lon, lat] for a Point, or
# {"type": "LineString", "coordinates": [[lon, lat], ...]} for a passage line.
# These take precedence over every other geocoding step. Use sparingly — file an
# xlsx fix instead when possible.
MANUAL_LOCATION_OVERRIDES = {
    # תו חניה לשכונת קטמונים ח-ט — area-wide parking permit recommendation.
    # No specific lot; user chose the whole sub-neighborhood polygon 2026-05-28.
    ("תו חניה לשכונת קטמונים ח-ט", "קטמונים ח-ט"): {"area_polygon": True},
    # חשיבה תנועתית מתחמית (פת) — area-wide "add internal streets" concept.
    # No specific location; user chose the whole sub-neighborhood polygon 2026-05-28.
    ("חשיבה תנועתית מתחמית", "פת"): {"area_polygon": True},
    # צומת ניות (רסקו) — נקודת רמזור. User accepted nominatim geocode 2026-05-28;
    # frozen as Point override so it won't drift on cache rebuild.
    ("צומת ניות", "רסקו"): [35.2046894, 31.7670463],
    # מגינת ניקנור לרח' יהודה הנשיא — passage from garden to יהודה הנשיא.
    # User confirmed the street-named match is in the right place 2026-05-28;
    # frozen as LineString override.
    ("מגינת ניקנור לרח' יהודה הנשיא", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.1996924, 31.7569512],
            [35.1997439, 31.7566296],
            [35.1997723, 31.7563742],
        ],
    },
    # מתחם אנטיגונוס — רחוב חדש (טוביה בן חפץ extension). User confirmed the
    # street-named match traces the right segment 2026-05-28; frozen as LineString.
    ("מתחם אנטיגונוס", "גוננים א-ו", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.207667, 31.757085],
            [35.208368, 31.757694],
            [35.208475, 31.757752],
            [35.208574, 31.757798],
            [35.208698, 31.757833],
            [35.208831, 31.757847],
        ],
    },
    # תחנות רח' אליעזר הגדול — 2 bus stops, user-provided ITM 2026-05-28.
    # Rendered as separate small_circle markers (not a line).
    ("תחנות רח' אליעזר הגדול", "גוננים א-ו"): {
        "type": "MultiPoint",
        "coordinates": [
            [35.204546, 31.755184],
            [35.205780, 31.754372],
        ],
    },
    # מתחם ברניקי פת — road/שצ"פ swap proposal. User chose a point at the
    # מתחם center 2026-05-28 (midpoint of the original LineString match).
    ("מתחם ברניקי פת", "גוננים א-ו"): [35.200499, 31.757483],
    # מפרצי העלאה והורדת למתחם מבני ציבור (פת) — pickup/dropoff bays on
    # ברל לוקר. User-provided ITM 2026-05-28.
    ("מפרצי העלאה והורדת למתחם מבני ציבור", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.200622, 31.751329],
            [35.200871, 31.750980],
        ],
    },
    # התחדשות עירונית הנושקת לרח' יעקב פת — חניונים. User-provided 5-point
    # ITM line along the relevant segment of יעקב פת 2026-05-28.
    ("התחדשות עירונית הנושקת לרח' יעקב פת - חניונים", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.199691, 31.753396],
            [35.200431, 31.752611],
            [35.201676, 31.751607],
            [35.202459, 31.751133],
            [35.202974, 31.750951],
        ],
    },
    # מתחם מכוור/הורקניה — outline polygon from "תכנון מתחמי 09.2023" PDF.
    # User-provided 15-vertex ring 2026-05-28 (rendered in pink).
    ("מתחם מכוור/הורקניה", "גוננים א-ו", "הצעה להכנת תוכנית אב מתחמית להתחדשות"): {
        "type": "Polygon",
        "coordinates": [[
            [35.202287, 31.756452],
            [35.203178, 31.756707],
            [35.203339, 31.756424],
            [35.203607, 31.756251],
            [35.204047, 31.756050],
            [35.204186, 31.755996],
            [35.204283, 31.755850],
            [35.204304, 31.755685],
            [35.204165, 31.755512],
            [35.203789, 31.755348],
            [35.203339, 31.755257],
            [35.202963, 31.755229],
            [35.202835, 31.755293],
            [35.202534, 31.755859],
            [35.202341, 31.756205],
            [35.202287, 31.756452],
        ]],
    },
    # מתחם ברניקי, פת — outline polygon from "תכנון מתחמי 09.2023" PDF.
    # User-provided 25-vertex ring 2026-05-28 (rendered in pink).
    ("מתחם ברניקי, פת", "גוננים א-ו", "הצעה להכנת תוכנית אב מתחמית להתחדשות"): {
        "type": "Polygon",
        "coordinates": [[
            [35.198618, 31.756069],
            [35.198361, 31.756379],
            [35.198221, 31.756616],
            [35.198232, 31.757017],
            [35.198597, 31.757072],
            [35.198693, 31.757282],
            [35.199037, 31.757236],
            [35.199101, 31.757519],
            [35.199316, 31.757911],
            [35.199766, 31.758185],
            [35.200067, 31.758304],
            [35.200678, 31.758468],
            [35.200925, 31.758358],
            [35.200657, 31.758021],
            [35.200496, 31.757592],
            [35.200485, 31.757291],
            [35.200560, 31.756671],
            [35.200485, 31.756871],
            [35.200582, 31.756616],
            [35.200560, 31.756415],
            [35.200496, 31.756251],
            [35.200206, 31.756333],
            [35.199788, 31.756370],
            [35.199552, 31.756278],
            [35.198961, 31.756123],
            [35.198618, 31.756069],
        ]],
    },
    # מתחם רשב"ג — outline polygon from "תכנון מתחמי 09.2023" PDF.
    # User-provided 16-vertex ring 2026-05-28 (rendered in pink).
    ("מתחם רשב\"ג", "גוננים א-ו", "הצעה להכנת תוכנית אב מתחמית להתחדשות"): {
        "type": "Polygon",
        "coordinates": [[
            [35.210463, 31.759152],
            [35.210752, 31.759170],
            [35.211600, 31.759252],
            [35.212083, 31.759261],
            [35.212104, 31.758878],
            [35.211933, 31.758386],
            [35.211718, 31.758030],
            [35.211353, 31.757473],
            [35.211128, 31.757017],
            [35.210592, 31.757218],
            [35.210592, 31.757546],
            [35.210559, 31.757765],
            [35.210431, 31.757875],
            [35.210302, 31.757957],
            [35.210398, 31.758477],
            [35.210441, 31.758887],
            [35.210463, 31.759152],
        ]],
    },
    # מתחם אנטיגונוס — outline polygon from "תכנון מתחמי 09.2023" PDF.
    # User-provided 21-vertex ring 2026-05-28 (rendered in pink).
    ("מתחם אנטיגונוס", "גוננים א-ו", "הצעה להכנת תוכנית אב מתחמית להתחדשות"): {
        "type": "Polygon",
        "coordinates": [[
            [35.207384, 31.758285],
            [35.207684, 31.758550],
            [35.208102, 31.758741],
            [35.208671, 31.758924],
            [35.209926, 31.759097],
            [35.210356, 31.759125],
            [35.210463, 31.759125],
            [35.210474, 31.758751],
            [35.210377, 31.758340],
            [35.210313, 31.757939],
            [35.210205, 31.757875],
            [35.209959, 31.757619],
            [35.209647, 31.757674],
            [35.209122, 31.757802],
            [35.208746, 31.757829],
            [35.208575, 31.757802],
            [35.208499, 31.758103],
            [35.208510, 31.758267],
            [35.208178, 31.758221],
            [35.207845, 31.758057],
            [35.207727, 31.757975],
            [35.207384, 31.758285],
        ]],
    },
    # שצ"פ זהרה (רסקו, פרויקט 8) — 1.4 דונם, gush/helka 30121/136. User wants
    # the whole lot polygon (2026-05-26). Parcel exists in parcels file.
    ("שצ\"פ זהרה", "רסקו"): {"gush_helka": "30121/136"},
    # Passage extension from המשוריינים to the existing מעבר לגרלף. No xlsx parcel,
    # and the chip's adjacent green polyline points NW instead of the actual E-W
    # direction — coords traced manually from the תשריט on 2026-05-17, then
    # refined to 4 ITM points provided by user (avoids existing buildings).
    ("הארכת נתיבי לגרלף", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.206125, 31.764960],
            [35.206845, 31.764822],
            [35.207306, 31.764748],
            [35.207669, 31.764509],
        ],
    },
    # Connection from קצנלסון to מדרגות עבדי — 53m east segment, user-provided 2026-05-17.
    # (Previously misassigned to #10 — corrected per user 2026-05-17.)
    ("חיבור למדרגות עבדי", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.20180734053431, 31.762854776698624],  # west
            [35.20236792218434, 31.76282969113318],   # east
        ],
    },
    # Western connection between שמעוני and הרצוג. User-provided 2 endpoints 2026-05-17.
    ("המשך מעבר עבדי", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2004352612508, 31.763052694532448],   # west (~דוד שמעוני 40)
            [35.20091537663527, 31.762975157474063],  # east (~60m)
        ],
    },
    # פת יעקב 3 (בית כנסת) — was nominatim on the junction; user-corrected to the
    # actual plot 2026-05-20.
    ("פת יעקב 3", "גוננים א-ו"): [35.19831472381227, 31.756940954015516],
    # חיבור בין השכונה לפארק המסילה — was nominatim at the lake; user-corrected
    # to the actual connection point near פאני קפלן plot 2026-05-20.
    # חיבור בין השכונה לפארק המסילה (פת, פרויקט 1) — LineString from neighborhood
    # to railway park. User-provided 2 ITM endpoints 2026-05-26 (~93m W→E).
    ("חיבור בין השכונה לפארק המסילה", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.2007844, 31.7498401],
            [35.1998964, 31.7495442],
        ],
    },
    # מתחם קנגורו (אולם ספורט) — was nominatim on a random street; user said
    # "kangaroo = plan 1261510" 2026-05-20, so use the polygon of plan 101-1261510.
    ("מתחם קנגורו", "קטמונים ח-ט"): {"plan_name": "101-1261510"},
    # מתחם קנגורו — road segment (כניסה מזרחית לשכונה מדב יוסף דרך המתחם).
    # Was inheriting the plan polygon; user provided 3 ITM points 2026-05-26
    # (~205m winding S→E).
    ("מתחם קנגורו", "קטמונים ח-ט", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.1975225, 31.7535408],
            [35.1980979, 31.7519708],
            [35.1985838, 31.7519136],
        ],
    },
    # התחדשות עירונית בן זכאי 31 — לפי "גוננים א-ו כבישים ותנועה.pdf" עמ' 4
    # המלצה #7: "טוביה בן חפץ - הפיכת שני רחובות ללא מוצא לרחוב אחד המשכי".
    # ה-xlsx פירש את התב"ע (101-1025295) כפוליגון 10×10 מ', אבל ההמלצה היא
    # יצירת קו (כביש) חדש. 3 נקודות ITM מהמשתמשת (2026-05-28), אורך 165 מ':
    #   ITM (220008.63, 629480.18) — קצה מזרחי
    #   ITM (219941.95, 629509.28) — מקטע אמצעי
    #   ITM (219850.14, 629516.16) — קצה מערבי
    ("התחדשות עירונית בן זכאי 31", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.21027, 31.757965],   # מזרח
            [35.209567, 31.758228],  # אמצע
            [35.208597, 31.75829],   # מערב
        ],
    },
    # שביל גוננים — הציר הראשי של 2.3 ק"מ החוצה את כל השכונה (פרויקט #4 בפרק ג'
    # של חוברת "גוננים מרחב ציבורי 05.2024.pdf", עמ' 20-24). שתי נקודות מוצא:
    # רח' בן זכאי במזרח ↔ רח' הרצוג במערב. עובר ליד אילנות, דנמרק, בית הנוער,
    # פלא, יהודה הנשיא, אליעזר הגדול. 35 waypoints מ-Google Earth KMZ שסומן
    # ידנית ע"י המשתמשת (2026-05-27); אורך מחושב 2,362 מ' (102.7% מ-2.30 ק"מ
    # המוצהרים בחוברת). זהו ה-PARENT trail — בהמשך יסומנו 9 מקטעי תכנון
    # נפרדים כפיצ'רים-בנים שיקושרו עם parent_project_id, כדי לתעד מקטעים
    # קיימים/בתכנון/תלויים-בתב"ע בנפרד.
    ("שביל גוננים", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.200043, 31.759621],
            [35.201251, 31.759883],
            [35.201695, 31.759132],
            [35.201362, 31.759023],
            [35.201448, 31.758708],
            [35.201645, 31.758328],
            [35.201437, 31.758018],
            [35.201155, 31.757418],
            [35.201164, 31.756659],
            [35.201391, 31.756117],
            [35.201415, 31.755288],
            [35.201535, 31.754574],
            [35.201909, 31.753832],
            [35.202486, 31.753315],
            [35.202752, 31.753039],
            [35.203124, 31.752815],
            [35.203489, 31.752956],
            [35.203758, 31.752716],
            [35.203767, 31.752477],
            [35.204834, 31.752452],
            [35.205247, 31.752061],
            [35.20586, 31.7526],
            [35.2062, 31.752956],
            [35.206537, 31.753645],
            [35.207181, 31.755217],
            [35.2076, 31.754943],
            [35.208148, 31.755637],
            [35.208586, 31.755899],
            [35.209408, 31.756429],
            [35.209749, 31.756706],
            [35.210271, 31.757285],
            [35.210481, 31.757399],
            [35.210635, 31.757455],
            [35.21074, 31.75775],
            [35.211911, 31.759211],
        ],
    },
    # מתחם אגד (קטמונים, פרויקט 10) — new road per master-plan, gush 30169/1,68,69
    # and 30170/3. Auto was 100m E-W parcel-edge; user provided 2 ITM endpoints
    # 2026-05-26 (~190m NW→SE).
    ("מתחם אגד", "קטמונים ח-ט"): {
        "type": "LineString",
        "coordinates": [
            [35.1968074, 31.7569172],
            [35.1978547, 31.7555166],
        ],
    },
    # שיפור תח"צ בקטמונים ח-ט — widening of רח' סן מרטין for bi-directional
    # transit service. User provided 7 ITM points 2026-05-26 (~290m winding N).
    ("שיפור תח\"צ בקטמונים ח-ט", "קטמונים ח-ט"): {
        "type": "LineString",
        "coordinates": [
            [35.1927751, 31.7541012],
            [35.1926019, 31.7546882],
            [35.1927470, 31.7551225],
            [35.1932273, 31.7555377],
            [35.1939589, 31.7558146],
            [35.1943667, 31.7560628],
            [35.1945230, 31.7565448],
        ],
    },
    # חידוש רחוב אברהם אלמליח — full street upgrade. Auto was 78m parcel-edge;
    # user requested 2026-05-26 to pull from roads.geojson and CUT at the given
    # point (keeping the longer portion — the actual project area).
    ("חידוש רחוב אברהם אלמליח", "קטמונים ח-ט"): {
        "road_name": "אברהם אלמליח",
        "cut_at": [35.1924346, 31.7530370],
    },
    # סן מרטין 1 (מעון יום) — xlsx has plan_raw=101-987016 but auto-match failed
    # due to leading-zero mismatch (plans.geojson has 101-0987016).
    ("סן מרטין 1", "קטמונים ח-ט"): {"plan_name": "101-0987016"},
    # סן מרטין 1 — transport variant (parking entrance, שיפור רמת שירות).
    # Keep as point at the specific intersection (סן מרטין-גולומב), not the plan polygon.
    # User-provided ITM 2026-05-26.
    ("סן מרטין 1", "קטמונים ח-ט", "שיפור רמת שירות"): [35.1954628, 31.7561411],
    # בית שוויץ (מתחם הפסגה) — user-corrected from "בית שלום" (wrong building) 2026-05-20.
    ("בית שוויץ (מתחם הפסגה)", "קטמונים ח-ט"): [35.19706010789749, 31.755262137288366],
    # בית שוויץ — road segment (traffic study for the מתחם הפסגה plan).
    # User-flagged 2026-05-28: change geometry to be along רחוב השומר (cross-
    # section already linked to PDF page 33 of דוח מסכם). Use roads.geojson
    # street-name match — 14 segments will be unioned.
    ("בית שוויץ (מתחם הפסגה)", "קטמונים ח-ט", "שדרוג רחובות"): {"road_name": "השומר"},
    # מתחם לוריא — has 6 services in the xlsx (school/kindergarten/daycare/synagogue +
    # this new street + a pedestrian crossing). Each needs its own location, so the
    # key includes service_he. Coords below are for the access-road service only.
    ("מתחם לוריא", "רסקו", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.20458076209259, 31.762520035568322],
            [35.2049166, 31.7629963],
        ],
    },
    # מעבר ה"ר מס' 12 — passage in גוננים, מאליעזר הגדול למרכז מסחרי.
    # User-corrected from ITM (219672.37,629224.18 → 219598.28,629153.81), 102m.
    ("מרח' אליעזר הגדול למרכז מסחרי", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.206721, 31.755657],
            [35.205939, 31.755022],
        ],
    },
    # שדרוג מעבר בן בבא-רח' בן זכאי — auto-derived from parcels was a 33m N-S
    # segment on a single lot boundary; user provided 4 ITM points 2026-05-26
    # forming an L-shaped passage from בן בבא south then east to בן זכאי.
    ("שדרוג מעבר בן בבא-רח' בן זכאי", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2065896, 31.7575217],
            [35.2070070, 31.7572080],
            [35.2071731, 31.7572080],
            [35.2071718, 31.7570791],
        ],
    },
    # מעבר ה"ר מס' 5 — מאלעזר בין יאיר לקנאי הגליל. Auto-derived was 25m E-W
    # inside one parcel; user provided 2 ITM endpoints 2026-05-26 (66m diagonal).
    ("מאלעזר בין יאיר לקנאי הגליל", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2024578, 31.7528497],
            [35.2019887, 31.7523772],
        ],
    },
    # מתחם דנמרק (פרויקט 8) — הנגשת קמפוס גוננים מדר'-מז'. Auto was 30m N-S
    # within one parcel; user provided 6 ITM points 2026-05-26 (~140m total,
    # winding NE→W through the campus). The passage and the road segment (חיבור
    # רח' טוביה לרח' אלכסנדריון, service="רחוב חדש") share the project name but
    # take different physical routes — split by service via 3-tuple override.
    ("מתחם דנמרק", "גוננים א-ו", "הוספת מעבר ה\"ר"): {
        "type": "LineString",
        "coordinates": [
            [35.2059429, 31.7528211],
            [35.2056860, 31.7529213],
            [35.2054793, 31.7531647],
            [35.2049879, 31.7532649],
            [35.2049823, 31.7531790],
            [35.2045522, 31.7531694],
        ],
    },
    # מתחם דנמרק — road segment (separate physical route from the passage above).
    # חיבור רח' טוביה לרח' אלכסנדריון. User provided 3 ITM points 2026-05-26.
    ("מתחם דנמרק", "גוננים א-ו", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.2031029, 31.7531933],
            [35.2032537, 31.7535751],
            [35.2031979, 31.7539760],
        ],
    },
    # מעבר ה"ר מס' 15 — מבני בתירא לחיים יחיל. Auto was 75m E-W on one parcel;
    # user provided 2 ITM endpoints 2026-05-26 (~114m east-west).
    ("מבני בתירא לחיים יחיל", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2038206, 31.7585431],
            [35.2050214, 31.7584452],
        ],
    },
    # מעבר ה"ר מס' 6 (פת) — מדב יוסף למבוא רוזנשטיין. Exit toward bus/rakl
    # stations on דב יוסף. Auto was 27m NE diagonal; user provided 2 ITM
    # endpoints 2026-05-26 (~38m WSW-ENE).
    ("מדב יוסף למבוא רוזנשטיין", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.1998440, 31.7525561],
            [35.1994782, 31.7527041],
        ],
    },
    # מעבר ה"ר מס' 7 (פת) — מהכניסה לבי"ס הדו לשוני עד לדב יוסף.
    # Auto was a 37m L over 3 parcels; user provided 2 ITM endpoints 2026-05-26
    # (~58m diagonal SE→NW).
    ("מהכניסה לבי\"ס הדו לשוני עד לדב יוסף", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.1994978, 31.7514251],
            [35.1991124, 31.7518355],
        ],
    },
    # מעבר ה"ר מס' 8 (פת) — מגן גולדשטיין למרכז פאני קפלן. Continuous walking
    # axis along the public boulevard. Auto was 80m N-S over 2 parcels; user
    # provided 2 ITM endpoints 2026-05-26 (~71m NNW→SSE).
    ("מגן גולדשטיין למרכז פאני קפלן", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.2001626, 31.7503585],
            [35.2005340, 31.7498026],
        ],
    },
    # מעבר ה"ר מס' 10 (פת) — מגן אולריך למעבר שמואל פינלס. Continuous link
    # from the neighborhood park to SE Pat. Auto was 35m L; user provided 2 ITM
    # endpoints 2026-05-26 (~55m NNW→SSE).
    ("מגן אולריך למעבר שמואל פינלס", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.2020530, 31.7501128],
            [35.2022960, 31.7496809],
        ],
    },
    # מתחם גבעת גונן (קטמונים, פרויקט 7) — passage (gush/helka 30199/16).
    # 156m E→W. User-provided 2026-05-26.
    ("מתחם גבעת גונן", "קטמונים ח-ט", "הוספת מעבר ה\"ר"): {
        "type": "LineString",
        "coordinates": [
            [35.1952223, 31.7546032],
            [35.1936919, 31.7540829],
        ],
    },
    # מתחם גבעת גונן — one-way road (כביש חד-סטרי) from רח' ר' צדוק to צומת
    # הנוטרים/אליעזר מרגולין. gush/helka 30199/15,53. User provided 6 ITM points
    # 2026-05-26 (~140m winding W→E).
    ("מתחם גבעת גונן", "קטמונים ח-ט", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.1950817, 31.7549962],
            [35.1945986, 31.7548411],
            [35.1944170, 31.7550272],
            [35.1941964, 31.7550845],
            [35.1940568, 31.7550630],
            [35.1936463, 31.7550176],
        ],
    },
    # שדרוג והנגשה מעבר ה"ר בין דב הוז לגולומב (קטמונים, פרויקט 2). Service is
    # "הנגשה" but the project name says "מעבר ה\"ר". Auto was a Polygon (the lot);
    # user provided 2 ITM endpoints 2026-05-26 (~69m N-S, the existing passage).
    ("שדרוג והנגשה מעבר ה\"ר בין דב הוז לגולומב", "קטמונים ח-ט"): {
        "type": "LineString",
        "coordinates": [
            [35.1953618, 31.7560397],
            [35.1953728, 31.7566601],
        ],
    },
    # מעלות שנהר (רסקו, פרויקט 1) — upgrade existing stairs + add seating.
    # Auto was 32m N-S inside parcel boundary; user provided 2 ITM endpoints
    # 2026-05-26 (~122m NW-SE, full length of the stairs).
    ("מעלות שנהר", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2027062, 31.7637519],
            [35.2018794, 31.7645919],
        ],
    },
    # מעלות סלמה לגרלף (רסקו, פרויקט 2) — upgrade existing stairs + seating.
    # Auto was 30m E-W within one parcel; user provided 2 ITM endpoints
    # 2026-05-26 (~105m W→E, full length of the stairs).
    ("מעלות סלמה לגרלף", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2060354, 31.7650786],
            [35.2049852, 31.7653793],
        ],
    },
    # נתיבי זהרה (רסקו, פרויקט 5) — accessibility upgrade for the existing
    # passage. Auto was 30m E-W on one parcel; user provided 2 ITM endpoints
    # 2026-05-26 (~91m W→E, full length of the passage).
    ("נתיבי זהרה", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2061303, 31.7655940],
            [35.2070408, 31.7655749],
        ],
    },
    # ציר הרצוג (רסקו, פרויקט 10) — upgrade of full הרב הרצוג street section.
    # Was PENDING_REVIEW; now pulled from roads.geojson by street-name match.
    ("ציר הרצוג", "רסקו"): {"road_name": "הרב הרצוג"},
    # בית הנוער (רסקו) — 24.08 dunam compound at הרצוג 105. 8 services all
    # share the same lot. User-provided plan 101-1158146 (2026-05-26).
    ("בית הנוער", "רסקו"): {"plan_name": "101-1158146"},
    # ציר סן מרטין (קטמונים) — full street + 2 missing connecting segments +
    # eastern part of בר יוחאי. User-confirmed by visual annotation 2026-05-26.
    ("ציר סן מרטין", "קטמונים ח-ט"): {
        "road_name": "חוזה סן מרטין",
        "extra_lines": [
            # Small connector: from סן מרטין south end to missing seg 1 start
            [
                [35.192753, 31.754111],
                [35.1930111, 31.7536498],
            ],
            # Missing seg 1 — west connector
            [
                [35.1930111, 31.7536498],
                [35.1940220, 31.7533779],
                [35.1951446, 31.7534018],
            ],
            # Missing seg 2 — east connector
            [
                [35.1961108, 31.7536166],
                [35.1966246, 31.7535403],
                [35.1968481, 31.7533637],
            ],
            # Cyan — eastern part of בר יוחאי connecting to יצחק שדה
            [
                [35.1968481, 31.7533637],
                [35.1993166, 31.7543756],
            ],
        ],
    },
    # מעגלי יבנה (גוננים, פרויקט 16) — section of future כביש 34 (רחוב המסילה).
    # The xlsx lists addresses on 3 streets (מעגלי יבנה / רבי חנינא / רבן יוחנן
    # בן זכאי). Auto-derived was a 60m parcel-edge fragment; user provided 3 ITM
    # points 2026-05-26 (~380m diagonal across the project's footprint).
    ("מעגלי יבנה", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2138771, 31.7585618],
            [35.2115869, 31.7562235],
            [35.2105815, 31.7554314],
        ],
    },
    # פינוי בינוי מתחם נהוראי 5-11, 18-28 (גוננים, פרויקט 17) — continuation of
    # כביש 34 along נהוראי. Auto was 46m E-W; user provided 2 ITM endpoints
    # 2026-05-26 (~95m NE→SW along the future road).
    ("פינוי בינוי מתחם נהוראי 5-11, 18-28", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2094644, 31.7542574],
            [35.2086824, 31.7536848],
        ],
    },
    # תב"ע לכביש 34 עתידי (גוננים) — section of the future כביש 34 on נהוראי 6-30
    # (even numbers). Was PENDING_REVIEW; user provided 2 ITM endpoints 2026-05-26
    # (~225m NE→SW).
    ("תב\"ע לכביש 34 עתידי", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.2105592, 31.7554600],
            [35.2086378, 31.7537325],
        ],
    },
    # חוות הנוער (רסקו, מעברי ה"ר 11,12,13) — 3 הצעות למעברים בשטח חוות הנוער
    # הציוני. הגבול הדרומי (מעבר 13) הוא אותו מעבר כמו "מבני בתירא לחיים יחיל"
    # (מעבר 15 בגוננים) — מסומן שם בנפרד. כאן מוצגים הקטע הצפוני (מעבר 11)
    # והקטע המרכזי (מעבר 12, שני קווים). User-provided ITM points 2026-05-26.
    ("חוות הנוער", "רסקו"): {
        "type": "MultiLineString",
        "coordinates": [
            # Northern segment (מעבר 11) — 52m W→E
            [
                [35.2023711, 31.7621055],
                [35.2018824, 31.7619623],
            ],
            # Central — line A (4 points, NE→S, winding through campus)
            [
                [35.2044937, 31.7611439],
                [35.2034435, 31.7606524],
                [35.2032871, 31.7603278],
                [35.2034157, 31.7601608],
            ],
            # Central — line B (3 points, E→W)
            [
                [35.2040635, 31.7602897],
                [35.2033709, 31.7601513],
                [35.2028795, 31.7599413],
            ],
        ],
    },
    # שב"צ קנאי הגליל — at plan 101-1056019 (per user 2026-05-20).
    ("שב\"צ קנאי הגליל", "גוננים א-ו"): {"plan_name": "101-1056019"},
    # מתחם דנמרק (פרויקט 9) — קמפוס הציבור של גוננים, 55.86 דונם, שטח לבניין
    # ציבורי. 9 שירותי פרוגרמה (gan/yesodi/al_yesodi/me'on/matnas/noar/knesset/
    # welfare) משתפים את אותו מגרש. גוש 30143 / חלקות
    # 136,146,165,166,167,170,186,189,201,218,219 (כל החלקות של המתחם).
    # User-requested 2026-05-27: להציג כפוליגון (חום) במקום 9 מלבני
    # "הפרשה מבונה" קטנים. ה-3-tuple overrides למעבר ה"ר ולרחוב חדש נשארים
    # יותר ספציפיים ולכן מנצחים את ה-2-tuple הזה לאותם שירותים.
    ("מתחם דנמרק", "גוננים א-ו"): {"gush_helka": "30143/136,146,165,166,167,170,186,189,201,218,219"},
    # מתחם ברניקי, פת (מתחם 3, בין יהודה הנשיא/ברניקי/פת/ניקנור) — 40 דונם
    # תיאוריים, 20.6 דונם פיזיים מ-30 החלקות שלהלן בגוש 30171.
    # xlsx location_id = "45-57,150,151,156,194-200,110,,121,214,217-220/30171"
    # אבל ה-regex extract_migrash_pairs נשבר על ה-,, ותפס רק 121,214,217-220.
    # User-confirmed 2026-05-28: להוסיף את כל 30 החלקות במפורש.
    ("מתחם ברניקי, פת", "גוננים א-ו"): {
        # User-confirmed 2026-05-28: extended with 6 sub-lot helkas from PDF
        # (בית כנסת 71/157/173, גנ"י 98/134/158, גן משחקים 56/151) — total ~22 דונם.
        "gush_helka": "30171/45,46,47,48,49,50,51,52,53,54,55,56,57,71,98,110,121,134,150,151,156,157,158,173,194,195,196,197,198,199,200,214,217,218,219,220"
    },
    # סן מרטין 15-5 (קדמת גונן) — 4 services (גני ילדים, מעון יום, בית כנסת,
    # שירותי תרבות). xlsx address is WRONG: plan 101-0511923 "קדמת גונן" is
    # actually at בר יוחאי 5-15, NOT סן מרטין 5-15 (~400m west of where xlsx
    # implies). The auto-match landed on landuse_xplan num=25 mc=400 (the
    # plan's standalone SHAVAZ allocation at the eastern edge [35.198, 31.7535])
    # which is geographically detached from the actual residential lots.
    # Override to use the full plan polygon so all 4 services display centered
    # on the actual 52-dunam plan area in בר יוחאי. User-flagged 2026-05-28.
    ("סן מרטין  15-5 (קדמת גונן)", "קטמונים ח-ט"): {"plan_name": "101-0511923"},
    # Same project under alternative xlsx label ("קדמת גונן, בר יוחאי 5,7,11,13,15"
    # with svc="טרם נקבע" — pn=9 row 5). Stripped — lookup uses .strip().
    ("קדמת גונן, בר יוחאי 5,7,11,13,15", "קטמונים ח-ט"): {"plan_name": "101-0511923"},
    # צומת הכניסה הצפונית (קטמונים, פרויקט 3) — תב"ע 1069764 בסן מרטין 4-6
    # מרחיבה את הצומת על חשבון השצ"פ. xlsx location_id="95/30199" — חלקה 95
    # היא 12,726 m² של גן ציבורי גדול שעוטף את האזור — ענק מדי. הפלאן עצמה
    # קטנה יותר (7,275 m², centered at [35.193486, 31.755944]) ומכוונת
    # למיקום המדויק של הצומת. User-flagged 2026-05-28.
    ("צומת הכניסה הצפונית", "קטמונים ח-ט"): {"plan_name": "101-1069764"},
    # מכוור/רבי טרפון (בית כנסת) — at parcel גוש 30123 / חלקה 16 (per PDF excerpt
    # + user 2026-05-20).
    ("מכוור/רבי טרפון", "גוננים א-ו", "בית כנסת"): {"gush_helka": "30123/16"},
    # מתחם מכוור/הורקניה — same parcels (גוש 30143 / 9,149,150) for both gan ילדים
    # and מועדון נוער (same physical complex, different services).
    ("מתחם מכוור/הורקניה", "גוננים א-ו", "גן ילדים"): {"gush_helka": "30143/9,149,150"},
    ("מתחם מכוור/הורקניה", "גוננים א-ו", "מועדון נוער"): {"gush_helka": "30143/9,149,150"},
    # גן אנטיגנוס (גן ילדים) — union of גוש 30133 / 35, 36 (per PDF excerpt).
    ("גוננים, רח' אנטיגנוס 2, גן אנטיגנוס", "גוננים א-ו", "גן ילדים"): {"gush_helka": "30133/35,36"},
    # אלעזר הגדול 4 (תחנה ט' - גוננים, תחנת בריאות) — user-provided specific point
    # within the מתחם דנמרק complex.
    ("אלעזר הגדול 4", "גוננים א-ו"): [35.204501490529324, 31.75434455925096],
    # פינוי בינוי מתחם נהוראי 5-11, 18-29 — gan + me'on yom (440 m² combined, 2 כיתות גן).
    # Per תב"ע 101-0778035, the הפרשה מבונה is on מגרש 1 (residential lot מגורים ד'
    # at center [35.2084544, 31.7544346] per landuse_xplan). User-confirmed 2026-05-26.
    ("פינוי בינוי מתחם נהוראי 5-11, 18-29", "גוננים א-ו", "גן ילדים"): [35.2084544, 31.7544346],
    ("פינוי בינוי מתחם נהוראי 5-11, 18-29", "גוננים א-ו", "מעון יום"): [35.2084544, 31.7544346],
    # מתחם 06 אלמליח בר יוחאי — same plan (101-0855585) for all 3 services
    # (מעון יום + בית כנסת + טרם נקבע).
    ("מתחם 06 אלמליח בר יוחאי", "קטמונים ח-ט", "מעון יום"): {"plan_name": "101-0855585"},
    ("מתחם 06 אלמליח בר יוחאי", "קטמונים ח-ט", "בית כנסת"): {"plan_name": "101-0855585"},
    ("מתחם 06 אלמליח בר יוחאי", "קטמונים ח-ט", "טרם נקבע"): {"plan_name": "101-0855585"},
    # מוריס שיפר 20 — same parcels (30501/82,83) for both social services and
    # additional welfare services (same complex).
    ("מוריס שיפר 20", "פת", "משרדי שירותים חברתיים"): {"gush_helka": "30501/82,83"},
    ("מוריס שיפר 20", "פת", "שירותי רווחה נוספים"): {"gush_helka": "30501/82,83"},
    # מבוא בנין 14 — building #14 doesn't exist in OSM (gap between #12 and #16);
    # midpoint of #12 and #16 placed manually.
    ("מבוא בנין 14", "רסקו"): [35.203730, 31.763574],
    # יצחק קצנלסון 3 — user-corrected location 2026-05-20.
    ("יצחק קצנלסון 3", "רסקו"): [35.20263157249951, 31.76267720490414],
    # השומר 8 (בית כנסת, קטמונים) — user-corrected from ITM 218675.73,629141.48.
    ("השומר 8", "קטמונים ח-ט"): [35.196202, 31.754911],
    # מרגולין 3 (מועדון נוער, קטמונים) — user-corrected 2026-05-20.
    ("מרגולין 3", "קטמונים ח-ט"): [35.19543649592314, 31.7547198685205],
    # השומר 7-5 (מעון יום, קטמונים) — user-corrected 2026-05-20.
    ("השומר 7-5", "קטמונים ח-ט"): [35.19733013537251, 31.754925131840395],
    # אברהם ארסט 10 (פת, מעון יום + גני ילדים) — user-corrected from ITM
    # 219233.48,628671.45. xlsx spells "ארסט" missing נ — proper name is "ארנסט".
    ("אברהם ארסט 10", "פת"): [35.202089, 31.750672],
    # עלמא אמונה (בי"ס על יסודי, גוננים) — parcel גוש 30133 / חלקה 129 (per PDF excerpt #42).
    ("עלמא אמונה", "גוננים א-ו"): {"gush_helka": "30133/129"},
    # בי"ס פלא (מעבר ה"ר ליד בית הספר) — at parcel גוש 30143 / חלקה 211 (per xlsx
    # location_id 211/30143 which auto-match parsed as helka/gush wrong).
    ("בי\"ס פלא", "גוננים א-ו"): {"gush_helka": "30143/211"},
    # מעבר הפילבוקס (הנגשה, רסקו) — at parcel גוש 30024 / חלקה 34 (per xlsx
    # location_id 34/30024 — same helka/gush format issue).
    # מעבר הפילבוקס (רסקו, פרויקט 3) — הנגשת מעבר 25×2 מ' בצומת חיים הזז.
    # gush 30024 isn't in parcels file; user provided 2 ITM endpoints 2026-05-26.
    ("מעבר הפילבוקס", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2094275, 31.7695996],
            [35.2096565, 31.7690604],
        ],
    },
    # גינה קהילתית ברודי (הנגשה, רסקו) — at parcel גוש 30024 / חלקה 140.
    # גינה קהילתית ברודי (רסקו, פרויקט 4) — accessibility ramp + steps at the
    # entrance to the community garden (15×6m path). gush 30024 isn't in the
    # parcels file; user provided 3 ITM points 2026-05-26 (~28m winding S).
    ("גינה קהילתית ברודי", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2100405, 31.7680903],
            [35.2100670, 31.7679830],
            [35.2099553, 31.7678434],
        ],
    },
    # מעבר חציה בחיים ברלין (רסקו, מעבר חציה מס' 1) — המשכי למעבר ה"ר ברודי.
    # User-provided point 2026-05-28 to replace previous gush_helka 30024/24 polygon.
    ("מעבר חציה בחיים ברלין", "רסקו"): [35.209894, 31.768392],
    # מעבר חציה בטייסים (רסקו, מעבר חציה מס' 2) — מתומרר שלא בצומת, ביטול 4-5 חניות.
    ("מעבר חציה בטייסים", "רסקו"): [35.207067, 31.765565],
    # מעבר חציה בטשרניחובסקי (רסקו, מעבר חציה מס' 4) — מוגבה לה"ר.
    # 4-tuple form: (name, area, service_he, project_num) to distinguish from
    # the other טשרניחובסקי crosswalk (#7).
    ("מעבר חציה בטשרניחובסקי", "רסקו", "הוספת מעבר חציה", "מעבר חציה מס' 4"): [35.203934, 31.764338],
    # מעבר חציה בדוד שמעוני (רסקו, מעבר חציה מס' 5) — מוגבה, המשכי לה"ר.
    # מתוכנן בעת הפיכתו לחד סטרי (קראוס חן).
    ("מעבר חציה בדוד שמעוני", "רסקו"): [35.202298, 31.764151],
    # מעבר חציה בשי עגנון (רסקו, מעבר חציה מס' 6) — מבוקש לה"ר, מתוכנן ע"י רולי פלד.
    ("מעבר חציה בשי עגנון", "רסקו"): [35.203859, 31.762691],
    # מעבר חציה בטשרניחובסקי (רסקו, מעבר חציה מס' 7) — חד-סטרי מצפון לדרום,
    # אזור חצייה ללא סימון 811. ביטול 4 מקומות חניה.
    ("מעבר חציה בטשרניחובסקי", "רסקו", "הוספת מעבר חציה", "מעבר חציה מס' 7"): [35.201775, 31.762846],
    # === גוננים א-ו crosswalks (#72-#80) — multiple ניקנור/רשב"ג/אנטיגונוס need 4-tuple ===
    # ניקנור #1 (גוננים, ממעבר חציה מס' 1) — מתומרר שלא בצומת, 50/150.
    ("ניקנור", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס' 1"): [35.201778, 31.758883],
    # ניקנור #2 (גוננים, ממעבר חציה מס' 2) — מתומרר שלא בצומת, 80/160, ביטול ~5 חניות.
    ("ניקנור", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס' 2"): [35.199932, 31.758272],
    # ניקנור #5 (גוננים, ממעבר חציה מס'5) — מוגבה, 60מ', ביטול 2 חניות.
    ("ניקנור", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'5"): [35.199026, 31.756826],
    # אלעזר בן יאיר #9 (גוננים, ממעבר חציה מס'9) — מוגבה, ביטול כ-5 חניות.
    ("אלעזר בן יאיר", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'9"): [35.201944, 31.753277],
    # רשב"ג #10 (גוננים, ממעבר חציה מס'10) — מתומרר שלא בצומת, ביטול כ-5 חניות.
    ("רשב\"ג", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'10"): [35.208907, 31.755430],
    # אנטיגונוס #12 (גוננים, ממעבר חציה מס'12) — מתומרר שלא בצומת, ביטול כ-5 חניות.
    ("אנטיגונוס", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'12"): [35.207427, 31.756853],
    # מעגלי יבנה #13 (גוננים, ממעבר חציה מס'13) — מתומרר שלא בצומת, מתוכננת במסגרת מעגלי יבנה.
    ("מעגלי יבנה", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'13"): [35.211729, 31.758071],
    # רשב"ג #14 (גוננים, ממעבר חציה מס'14) — צמוד לצומת, מותנה ביוזמה מתחמית עתידית.
    ("רשב\"ג", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'14"): [35.210581, 31.757615],
    # אנטיגונוס #16 (גוננים, ממעבר חציה מס'16) — 50מ', מעבר לה"ר, מותנה במתחם.
    ("אנטיגונוס", "גוננים א-ו", "מעבר חציה", "ממעבר חציה מס'16"): [35.209352, 31.757797],
    # === קטמונים ח-ט crosswalks (#168-#171) ===
    # דב הוז #1 (קטמונים, מעבר חציה מס' 1) — מוגבה, אין בקרבתו, מעבר לה"ר.
    ("דב הוז", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 1"): [35.195862, 31.756168],
    # אלמליח #2 (קטמונים, מעבר חציה מס' 2) — מתומרר שלא בצומת, 60/ללא, ביטול 5-6 חניות.
    # 4-tuple — אלמליח חוזרת ב-#171 (מעבר חציה מס' 11).
    ("אלמליח", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 2"): [35.193498, 31.756228],
    # הנוטרים #3 (קטמונים, מעבר חציה מס' 3) — מוגבה, אין בקרבתו, ביטול 2 חניות.
    # 4-tuple — הנוטרים חוזרת ב-#173 (מעבר חציה מס' 4).
    ("הנוטרים", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 3"): [35.196204, 31.755549],
    # הנוטרים #4 (קטמונים, מעבר חציה מס' 4) — מוגבה, להרחיק מהסיבוב, ביטול 2 חניות.
    ("הנוטרים", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 4"): [35.195732, 31.755316],
    # השומר #5 (קטמונים, מעבר חציה מס' 5) — מוגבה, אין בקרבתו, ביטול 5-6 חניות.
    ("השומר", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 5"): [35.196826, 31.754778],
    # בר יוחאי #7 (קטמונים, מעבר חציה מס' 7) — מוגבה, מותנה בתכניות מקבילות.
    # 4-tuple — בר יוחאי חוזרת ב-#177 (מס' 9) ו-#178 (מס' 10).
    ("בר יוחאי", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 7"): [35.197594, 31.753578],
    # סן מרטין #8 (קטמונים, מעבר חציה מס' 8) — מתומרר שלא בצומת,
    # משולב בתכנית התחדשות עירונית בר יוחאי 5-15. xlsx name=סן מרטין;
    # PDF main_street=בר יוחאי (same physical crosswalk on the corner).
    ("סן מרטין", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 8"): [35.195566, 31.752848],
    # בר יוחאי #9 (קטמונים, מעבר חציה מס' 9) — צומת בר יוחאי/סן מרטין,
    # מותנה בהסדרת הצומת בעתיד.
    ("בר יוחאי", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 9"): [35.193602, 31.753473],
    # בר יוחאי #10 (קטמונים, מעבר חציה מס' 10) — צומת בר יוחאי/חיבור לדוד איילון,
    # מותנה בהסדרת הצומת בעתיד.
    ("בר יוחאי", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 10"): [35.193720, 31.752880],
    # יצחק שדה #6 (קטמונים, מעבר חציה מס' 6) — מוגבה, אין בקרבתו, ביטול 5-6 חניות.
    ("יצחק שדה", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 6"): [35.197223, 31.754162],
    # אלמליח #11 (קטמונים, מעבר חציה מס' 11) — מתומרר שלא בצומת, 100/ללא,
    # מותנה בצומת חניון/דוד אילון או בתכניות התחדשות בר יוחאי 17-21/אלמליח.
    # 4-tuple — אלמליח הראשונה ב-#169 (מעבר חציה מס' 2).
    ("אלמליח", "קטמונים ח-ט", "מעבר חציה", "מעבר חציה מס' 11"): [35.192961, 31.752565],
    # === פת crosswalk (#179) ===
    # ברל לוקר #12 (פת, מעבר חציה מס' 12) — מתומרר שלא בצומת, ביטול ~6 חניות,
    # מיקום סופי ימוקם בהתאם לתכנית התחדשות עירונית הקרובה ולתחנת אוטובוס.
    ("ברל לוקר", "פת", "מעבר חציה", "מעבר חציה מס' 12"): [35.202459, 31.749409],
    # מתחם לוריא — has 4 public-services (school/kindergartens/daycare/synagogue)
    # all on the same complex at גוש 30185 / חלקות 178,203 (per PDF #10).
    # The 5th service (רחוב חדש) and 6th (מעבר ה"ר) keep their own dedicated coords.
    ("מתחם לוריא", "רסקו", "בי\"ס יסודי"): {"gush_helka": "30185/178,203"},
    ("מתחם לוריא", "רסקו", "גני ילדים"): {"gush_helka": "30185/178,203"},
    ("מתחם לוריא", "רסקו", "מעון יום"): {"gush_helka": "30185/178,203"},
    ("מתחם לוריא", "רסקו", "בית כנסת"): {"gush_helka": "30185/178,203"},
    # 3 גני ילדים ברחוב גוננים (משרה מלכה, מנח"י) — at parcels גוש 30171 / חלקות
    # 98,134,158 (per PDF #31, address ברניקי).
    ("3 גני ילדים ברחוב גוננים (משרה מלכה, מנח\"י)", "גוננים א-ו"): {"gush_helka": "30171/98,134,158"},
    # גשר בין דב הוז לגולומב (קטמונים) — proposed bridge. User refined to 2 ITM
    # endpoints 2026-05-26 (~64m N→S across the highway/park to bus stop).
    ("גשר בין דב הוז לגולומב", "קטמונים ח-ט"): {
        "type": "LineString",
        "coordinates": [
            [35.1958579, 31.7570364],
            [35.1961148, 31.7564781],
        ],
    },
    # גשר בין רח' חננאל לעמק הצבאים (גוננים, פרויקט 7) — bridge approved within
    # תב"ע 101-1182146. User clarified 2026-05-26 — keep only the segment ABOVE
    # הרצוג (the actual bridge crossing), not the access spine going south.
    ("גשר בין רח' חננאל לעמק הצבאים", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.1985603, 31.7579751],
            [35.1984030, 31.7581114],
            [35.1983039, 31.7582447],
        ],
    },
    # דוד איילון (שיפור רמת שירות, קטמונים) — entry/exit reorganization. User
    # refined to 3 ITM points 2026-05-26 (~78m L-shape south then west).
    ("דוד איילון", "קטמונים ח-ט"): {
        "type": "LineString",
        "coordinates": [
            [35.1937750, 31.7528319],
            [35.1934902, 31.7521590],
            [35.1931719, 31.7521542],
        ],
    },
    # רח' אברהם ארנסט (שיפור רמת שירות, פת) — merged from roads.geojson, 124m.
    ("רח' אברהם ארנסט", "פת"): {
        "type": "LineString",
        "coordinates": [
            [35.20062, 31.751351],
            [35.201022, 31.751537],
            [35.201517, 31.750826],
        ],
    },
    # מאנטיגונוס לגן אנטיגונוס (מעבר ה"ר מס' 13, גוננים) — 61m vertical connection.
    ("מאנטיגונוס לגן אנטיגונוס", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.208503, 31.757060],
            [35.207944, 31.757337],
        ],
    },
    # מעבר ה"ר מגן חביבה — name shared by 3 different passages in רסקו; need
    # project_num to disambiguate. User-provided 2026-05-20.
    ("מעבר ה\"ר מגן חביבה", "רסקו", "הוספת מעבר ה\"ר", "מעבר ה\"ר מס' 1"): {
        "type": "LineString",
        "coordinates": [
            [35.20838734746782, 31.76699409290235],
            [35.20850018986884, 31.76755207720798],
            [35.20844893755101, 31.76787466753757],
            [35.20848998690371, 31.76831059507781],
            [35.20832591472203, 31.76842394520724],
            [35.20831566826115, 31.76855472478972],
            [35.2085310257691, 31.76858958816536],
        ],
    },
    ("מעבר ה\"ר מגן חביבה", "רסקו", "הוספת מעבר ה\"ר", "מעבר ה\"ר מס' 2"): {
        "type": "LineString",
        "coordinates": [
            [35.207413, 31.768325],
            [35.207570, 31.767736],
        ],
    },
    ("מעבר ה\"ר מגן חביבה", "רסקו", "הוספת מעבר ה\"ר", "מעבר ה\"ר מס' 4"): {
        "type": "LineString",
        "coordinates": [
            [35.207710, 31.766959],
            [35.207380, 31.766968],
            [35.207280, 31.767054],
            [35.206458, 31.767050],
            [35.205978, 31.766973],
            [35.205056, 31.766821],
        ],
    },
    # מעבר 10 גוננים (מדרך האמוראים לרח' אליעזר הגדול) — 159m from user ITM 2026-05-20.
    ("מדרך האמוראים לרח' אליעזר הגדול", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.209778, 31.758647],
            [35.211459, 31.758690],
        ],
    },
    # מעבר 7 גוננים (מרח' קנאי הגליל לת. רק"ל בפת) — 77m from user ITM 2026-05-20.
    ("מרח' קנאי הגליל לת. רק\"ל בפת", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.203656, 31.750992],
            [35.203628, 31.751684],
        ],
    },
    # מעבר 11 גוננים (רח' אליעזר הגדול לפארק גוננים) — 193m from user ITM 2026-05-20.
    ("רח' אליעזר הגדול לפארק גוננים", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.207595, 31.753174],
            [35.208863, 31.754539],
        ],
    },
    # מעבר 3 רסקו (מעבר ה"ר מפארק נילי) — 86m from user ITM (3 pts) 2026-05-20.
    ("מעבר ה\"ר מפארק נילי", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.207199, 31.767644],
            [35.206741, 31.767830],
            [35.206512, 31.768116],
        ],
    },
    # מעבר 4 גוננים (מגן אלכסנדריון לרח' קנאי הגליל) — 158m / 5 pts from user ITM 2026-05-20.
    ("מגן אלכסנדריון לרח' קנאי הגליל", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.200203, 31.755147],
            [35.200387, 31.755210],
            [35.200728, 31.755229],
            [35.200728, 31.755310],
            [35.201750, 31.755186],
        ],
    },
    # מעבר 14 גוננים (מבן זכאי לבני בתירא) — 331m / 3 pts from user ITM 2026-05-20.
    ("מבן זכאי לבני בתירא", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.202158, 31.757634],
            [35.204129, 31.757453],
            [35.205229, 31.756570],
        ],
    },
    # מעבר 5 רסקו (המשך נתיבי זהרה) — 93m / 4 pts from user ITM 2026-05-20.
    ("המשך נתיבי זהרה", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.206037, 31.765601],
            [35.205573, 31.765611],
            [35.205562, 31.765682],
            [35.205135, 31.765723],
        ],
    },
    # מעבר 5 גוננים (מציר יהודה הנשיא לרח' אלכסנדריון) — 461m / 9 pts from user ITM 2026-05-20.
    ("מציר יהודה הנשיא לרח' אלכסנדריון", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.201390, 31.755680],
            [35.201406, 31.755174],
            [35.201445, 31.754687],
            [35.201769, 31.754014],
            [35.201937, 31.753790],
            [35.202233, 31.753432],
            [35.202691, 31.753122],
            [35.203479, 31.752477],
            [35.203819, 31.752444],
        ],
    },
    # מעבר 2 גוננים (מגינת ניקנור לרח' ניקנור) — 65m from user ITM (3 pts) 2026-05-20.
    ("מגינת ניקנור לרח' ניקנור", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.199918, 31.757596],
            [35.199583, 31.757672],
            [35.199315, 31.757853],
        ],
    },
    # מעבר 1 גוננים (מהורקניה לניקנור) — T-shaped passage, 2 segments / 145m total.
    # User-provided 2 pairs of ITM 2026-05-20.
    ("מהורקניה לניקנור", "גוננים א-ו"): {
        "type": "MultiLineString",
        "coordinates": [
            [[35.202479, 31.758295], [35.201739, 31.758844]],  # main arm (93m)
            [[35.202063, 31.758562], [35.201644, 31.758252]],  # cross arm (53m)
        ],
    },
    # מעבר 2 גוננים (מגינת ניקנור לרח' ניקנור) — chip-derived line was slightly too
    # long on west end (crossing onto a building). Trimmed left endpoint per user
    # 2026-05-20.
    ("מגינת ניקנור לרח' ניקנור", "גוננים א-ו"): {
        "type": "LineString",
        "coordinates": [
            [35.199335, 31.757865],  # trimmed west (was 35.1992459, 31.7579061)
            [35.199543, 31.757768],  # mid
            [35.1998814, 31.757642], # east
        ],
    },
    # ציר מבני ציבור (מעבר ה"ר מס' 8, רסקו) — 207m from user KMZ 2026-05-20.
    ("ציר מבני ציבור", "רסקו"): {
        "type": "LineString",
        "coordinates": [
            [35.2042124995081, 31.76388419657148],
            [35.20401645267174, 31.76438030965316],
            [35.2038297290477, 31.76432077928888],
            [35.20369436405558, 31.76485261248754],
            [35.2035216441796, 31.76483277047479],
            [35.20322755719656, 31.76523363419333],
        ],
    },
    # מתחם לוריא / passage (מעבר ה"ר מס' 7) — 176m from user KMZ 2026-05-20.
    ("מתחם לוריא", "רסקו", "הוספת מעבר ה\"ר"): {
        "type": "LineString",
        "coordinates": [
            [35.20512676749412, 31.76368704855995],
            [35.20506142294555, 31.76392121413493],
            [35.20505209464176, 31.76415537830304],
            [35.20503810007597, 31.76445304467871],
            [35.20497275127446, 31.76459592604018],
            [35.2047346786098, 31.76459593157189],
            [35.20465532454198, 31.76473087557864],
            [35.2046413251979, 31.76492535162129],
            [35.20473468945329, 31.76500075841066],
        ],
    },
    # מתחם מכוור/הורקניה — new street. User refined to 2 ITM endpoints 2026-05-26
    # (~30m E-W between Horkania and Yehuda Hanasi).
    ("מתחם מכוור/הורקניה", "גוננים א-ו", "רחוב חדש"): {
        "type": "LineString",
        "coordinates": [
            [35.2038520, 31.7561224],
            [35.2041634, 31.7560388],
        ],
    },
    # מתחם בן זכאי-אנטיגונוס-אליעזר הגדול — 30 dunam preservation complex.
    # User-drawn polygon from KMZ (Google Earth) 2026-05-20, ~35 dunam.
    ("מתחם בן זכאי-אנטיגונוס-אליעזר הגדול", "גוננים א-ו"): {
        "type": "Polygon",
        "coordinates": [[
            [35.20450431065631, 31.75538571606955],
            [35.20516818833218, 31.75477628474467],
            [35.20744666578539, 31.75683539889364],
            [35.20658664363136, 31.75752823934303],
            [35.20457976782419, 31.75578344112736],
            [35.20450431065631, 31.75538571606955],
        ]],
    },
}

# Source PDFs (in OneDrive — same folder as the xlsx)
PDF_DIR = XLSX_PATH.parent
# Booklet definitions: (label, [substring triggers that must ALL be in the link], pdf_path).
# Order = priority (most-specific first). The first definition where ALL triggers match wins.
BOOKLET_DEFS = [
    # Single-page traffic note (no page reference; 2 pages)
    ("בית שוויץ",            ["שוויץ"],                                     PDF_DIR / "התייחסות תנועתית בית שוויץ.pdf"),
    ("מצגת מתחם הפסגה",      ["הפסגה"],                                     PDF_DIR / "מצגת קטמונים מתחם הפסגה 05.2024.pdf"),
    # Traffic booklets — one per sub-neighborhood
    ("רסקו כבישים ותנועה",   ["רסקו", "כבישים"],                            PDF_DIR / "רסקו כבישים ותנועה.pdf"),
    ("קטמונים כבישים ותנועה",["קטמונים", "כבישים"],                         PDF_DIR / "קטמונים ח-ט כבישים ותנועה.pdf"),
    ("פת כבישים ותנועה",     ["פת", "כבישים"],                              PDF_DIR / "פת כבישים ותנועה.pdf"),
    ("גוננים כבישים ותנועה", ["גוננים", "כבישים"],                          PDF_DIR / "גוננים א-ו כבישים ותנועה.pdf"),
    # Brown-land cards — Katmonim has its own booklet, others share
    ("כרטסת קטמונים",        ["קטמונים", "כרטסת"],                          PDF_DIR / "קטמון ח-ט - כרטסת שטחים חומים 06.062024.pdf"),
    ("כרטסת רסקו-גוננים-פת", ["כרטסת"],                                     PDF_DIR / "כרטסת שטחים חומים מרחב רסקו-גוננים א-ו ופת 01052024.pdf"),
    # Public-space booklets (one per area cluster)
    ("מרחב ציבורי גוננים",   ["מרחב ציבורי", "גוננים"],                     PDF_DIR / "גוננים מרחב ציבורי_05.2024.pdf"),
    ("מרחב ציבורי רסקו",     ["מרחב ציבורי", "רסקו"],                       PDF_DIR / "רסקו מרחב ציבורי_17.4.24.pdf"),
    ("מרחב ציבורי קטמ+פת",   ["מרחב ציבורי"],                               PDF_DIR / "קטמונים ופת 05.2024.pdf"),
    # Initial proposals booklet
    ("תכנון מתחמי",          ["תכנון מתחמי"],                               PDF_DIR / "פרוייקטור אורנים_ תכנון מתחמי_09.2023.pdf"),
]


def _resolve_booklet(link: str):
    """Return (label, pdf_path) for the link, or (None, None) if no match."""
    for label, triggers, path in BOOKLET_DEFS:
        if all(t in link for t in triggers):
            return label, path
    return None, None

# --- Service mapping ------------------------------------------------------
# xlsx "שירות" string → app service_key (or "extra:<slug>" for new rows, or None for map-only)
SERVICE_MAP = {
    # Existing PARSER_KEYS
    "מעון יום": "maon",
    "גני ילדים": "gan",
    "גן ילדים": "gan",
    "רב גילאי": "gan",
    "בי\"ס יסודי": "yesodi",
    "יסודי": "yesodi",
    "בי\"ס על-יסודי": "al_yesodi",
    "על יסודי": "al_yesodi",
    "תחנה לבריאות המשפחה": "tipat_chalav",
    "מרפאה שכונתית": "clinic",
    "שירותי קהילה": "matnas",
    "משרדי שירותים חברתיים": "welfare_dept",
    "משרדי שירותים חברתיים ": "welfare_dept",
    "שירותי רווחה": "welfare_dept",
    "שירותי רווחה נוספים": "welfare_dept",
    "מוסדות רווחה נוספים": "welfare_dept",
    "מועדון נוער": "noar_club",
    "מועדון לקשיש": "elderly_club",
    "מרכז יום לקשיש": "elderly_day",
    "בית כנסת": "synagogue",
    "מקווה": "mikve",
    "אולם ספורט": "sport_hall",
    "ספרייה": "library",
    # extras (new rows in גוננים-only balance)
    "בריכת שחיה": "extra:pool",
    "מגרש ספורט": "extra:sport_field",
    "מגרש טניס": "extra:tennis",
    "מגרש כדורגל": "extra:soccer",
    "אולם מופעים": "extra:performance",
    "שירותי תרבות": "extra:culture",
    "מגן דוד אדום": "extra:emergency",
    "משטרה": "extra:emergency",
    "תחנת כיבוי אש": "extra:emergency",
}
EXTRA_LABELS = {
    "extra:pool": "בריכת שחיה",
    "extra:sport_field": "מגרש ספורט",
    "extra:tennis": "מגרש טניס",
    "extra:soccer": "מגרש כדורגל",
    "extra:performance": "אולם מופעים",
    "extra:culture": "שירותי תרבות",
    "extra:emergency": "חירום (מד\"א/משטרה/כיבוי)",
}

# --- Domain mapping -------------------------------------------------------
DOMAIN_MAP = {
    "פרוגרמה לצורכי ציבור": "programa",
    "מרחב ציבורי": "public_space",
    "תנועה": "transport",
}

# mavat_code filter per domain — pick ONLY relevant lots from a plan, not the whole plan footprint.
# Codes confirmed from data inspection:
#   400  = מבנים ומוסדות ציבור     410 = שטח לבניני ציבור
#   460  = פארק עירוני             670 = שטח ציבורי פתוח (שצ"פ)
#   830  = דרך מוצעת               860 = שביל
#   1670 = שטחים פתוחים ומבני ציבור (mixed)
#   1250/1300/1410/1480/1492/1550/1576/1578/1604 = הפרשה מבונה codes (treat as programa)
DOMAIN_MAVAT_CODES = {
    "programa": {400, 410, 450, 1250, 1300, 1410, 1480, 1492, 1550, 1576, 1578, 1604, 1670},
    "public_space": {460, 670, 1670, 400},  # primarily שצ"פ; 400 fallback when plan has only public-buildings plot
    "transport": {830, 860, 870},
}


# --- Helpers --------------------------------------------------------------
def project_id(area: str, proj_num, service: str, proj_name: str = "") -> str:
    # Includes proj_name to disambiguate rows where the xlsx has multiple distinct
    # projects under the same (area, num, service) — e.g. 5 different recommendations
    # share area='גוננים א-ו' num='' service='רחוב חדש'. Without proj_name they would
    # all collapse to one pid and only the first row's feature would be emitted.
    raw = f"{area}|{proj_num}|{service}|{proj_name}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


# --- PDF extraction -------------------------------------------------------
# Cache: {(pdf_path_str, page_idx): text}
_PDF_CACHE: dict = {}
_PDF_DOC_CACHE: dict = {}


def _open_pdf(path: Path):
    if path is None or not path.exists() or fitz is None:
        return None
    key = str(path)
    if key not in _PDF_DOC_CACHE:
        try:
            _PDF_DOC_CACHE[key] = fitz.open(str(path))
        except Exception:
            _PDF_DOC_CACHE[key] = None
    return _PDF_DOC_CACHE[key]


def _page_text(pdf_path: Path, page_idx0: int) -> str:
    """Return cleaned-up text for page (0-indexed). Empty string on failure."""
    if pdf_path is None or fitz is None:
        return ""
    key = (str(pdf_path), page_idx0)
    if key in _PDF_CACHE:
        return _PDF_CACHE[key]
    doc = _open_pdf(pdf_path)
    if doc is None or page_idx0 < 0 or page_idx0 >= doc.page_count:
        _PDF_CACHE[key] = ""
        return ""
    try:
        raw = doc[page_idx0].get_text("text") or ""
    except Exception:
        raw = ""
    # Trim whitespace + collapse heavy blank-line runs
    lines = [ln.rstrip() for ln in raw.splitlines()]
    out_lines = []
    blank = 0
    for ln in lines:
        if ln.strip():
            out_lines.append(ln)
            blank = 0
        else:
            blank += 1
            if blank <= 1:
                out_lines.append("")
    txt = "\n".join(out_lines).strip()
    _PDF_CACHE[key] = txt
    return txt


_PAGE_REF_RE = re.compile(r"עמ['׳]\s*([\d,\-]+)")
_SLIDE_RE = re.compile(r"שקופי(?:ות|ת)\s*([\d,\-]+)")

# Cache of "header → page index" indexes per PDF.
_BOOKLET_INDEX: dict[str, dict] = {}


# Hardcoded mapping for the "תכנון מתחמי" booklet — based on actual page contents:
#   מתחם 1 (אנטיגונוס) → pages 5-11
#   מתחם 2 (רשב"ג)     → pages 12-17 (core: 14)
#   מתחם 3 (ברניקי-פת) → pages 20-26 (core: 23)
#   מתחם 4 (הורקניה-מכוור) → pages 27-30 (core: 30)
METAHAM_PAGE_RANGES = {
    1: (5, 11, 8),    # range_start, range_end, "core" page
    2: (12, 17, 14),
    3: (20, 26, 23),
    4: (27, 30, 30),
}


def _build_booklet_index(pdf_path: Path) -> dict:
    """Scan all pages once. Currently we only need the total page count;
    page-number resolution uses the simple 1-indexed convention."""
    key = str(pdf_path)
    if key in _BOOKLET_INDEX:
        return _BOOKLET_INDEX[key]
    idx = {"n_pages": 0}
    doc = _open_pdf(pdf_path)
    if doc is None:
        _BOOKLET_INDEX[key] = idx
        return idx
    idx["n_pages"] = doc.page_count
    _BOOKLET_INDEX[key] = idx
    return idx


def _street_centroid(houses):
    """Return centroid of all building points on a street."""
    pts = []
    for hn, feat in houses:
        c = feature_centroid(feat)
        if c:
            pts.append(c)
    if not pts:
        return None
    return [sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts)]


def _find_street_in_text(text: str, searchable_streets: list[str]) -> str | None:
    """Find the longest street name from `searchable_streets` that appears as a token in text.
    Allow common Hebrew single-letter prefixes (מ ב ל ו ה ש כ) before the alias — e.g. 'מהורקניה'
    contains 'הורקניה' with the 'from' prefix. Token-aware to avoid 'פת' matching inside 'פתח'."""
    if not text:
        return None
    s = " " + str(text) + " "
    for street in searchable_streets:
        # Match either at word start (no Hebrew letter immediately before), OR after a single
        # prefix letter that is itself preceded by non-Hebrew. End of match must be at word end.
        pattern = (
            r"(?:(?<![א-ת])|(?<=[\s\-\.,\(\)/\"׳']))"
            r"[בלמוהשכ]?"
            + re.escape(street)
            + r"(?![א-ת])"
        )
        if re.search(pattern, s):
            return street
    return None


# --- Sub-neighborhood boundary geometry (cached for georef CP extraction).
_SUB_BOUNDARY_GEOM = None

def _get_sub_boundary_geom():
    """Return a shapely (Multi)LineString representing the outer edge of the
    4 גוננים sub-neighborhoods. Used as a snap-target for blue boundary
    polylines extracted from the תשריט PDF — these give dense control points
    for georeferencing."""
    global _SUB_BOUNDARY_GEOM
    if _SUB_BOUNDARY_GEOM is not None:
        return _SUB_BOUNDARY_GEOM
    if shape is None or unary_union is None:
        return None
    try:
        sub_geo = load_geojson(SUB_NEIGH_PATH)
    except Exception:
        return None
    polys = []
    target_names = set(AREA_TO_SUBNEIGH.values())  # {"גוננים", "רסקו - גבעת הורדים", ...}
    for f in sub_geo.get("features", []):
        if f["properties"].get("schn_nama") in target_names:
            try:
                polys.append(shape(f["geometry"]))
            except Exception:
                pass
    if not polys:
        return None
    union = unary_union(polys)
    _SUB_BOUNDARY_GEOM = union.boundary if union.geom_type != "Polygon" else union.exterior
    return _SUB_BOUNDARY_GEOM


def _build_pdf_to_wgs_transform(page, buildings_features, minahak_geom):
    """Build a high-accuracy PDF→WGS84 transform for the projector תשריט.

    Strategy (combines 3 sources of control points):
      1. **Street labels**: a rough affine from `<street_name>` spans → centers
         of all buildings on that street.
      2. **Per-chip nearest building**: each PDF span is matched not to the
         street's AVERAGE building but to the nearest one (eliminates the bias
         from long streets that cross multiple neighborhoods).
      3. **Blue boundary polylines** (NEW): the projector's תשריט draws the
         sub-neighborhood boundary in blue. We extract every vertex from those
         polylines, project them via the affine, and snap to the matching
         vertex on the real GIS sub-neighborhood polygons. This yields
         hundreds of additional control points distributed across the whole
         page — fixing local distortion that affine alone can't model.

    Finally, fits a Thin-Plate-Spline (RBF) interpolator over all CPs, which
    gives essentially zero residual on the control set and ~0-5m accuracy on
    interpolated points. Falls back to affine if scipy is unavailable.

    Returns: callable `transform(x, y) → (lon, lat)`, or None if too few CPs.
    """
    if fitz is None or _np is None:
        return None
    # 1) Page spans (text + position)
    spans = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for sp in line.get("spans", []):
                txt = (sp.get("text") or "").strip()
                if not txt:
                    continue
                bbox = sp.get("bbox")
                spans.append({
                    "text": txt,
                    "cx": (bbox[0] + bbox[2]) / 2,
                    "cy": (bbox[1] + bbox[3]) / 2,
                })
    # 2) Street → all buildings inside the minahak.
    from collections import defaultdict as _dd
    street_pts: dict[str, list] = _dd(list)
    for f in buildings_features:
        s = (f["properties"].get("street") or "").strip()
        g = f.get("geometry") or {}
        if not s or g.get("type") != "Point":
            continue
        cx, cy = g["coordinates"]
        if minahak_geom.contains(ShPoint(cx, cy)):
            street_pts[s].append((cx, cy))
    if not street_pts:
        return None
    street_avg = {s: (sum(p[0] for p in lst) / len(lst),
                     sum(p[1] for p in lst) / len(lst))
                  for s, lst in street_pts.items()}

    def _match_street(t):
        if t in street_pts:
            return t
        cands = [s for s in street_pts if s.endswith(" " + t) or s == t]
        return cands[0] if cands else None

    # 3) Pass 1: rough affine from street averages.
    initial_cps = []
    for sp in spans:
        real = _match_street(sp["text"])
        if real:
            rx, ry = street_avg[real]
            initial_cps.append((sp["cx"], sp["cy"], rx, ry))
    if len(initial_cps) < 6:
        return None

    def _fit_affine(cps):
        A = _np.array([[c[0], c[1], 1.0] for c in cps])
        rx = _np.array([c[2] for c in cps])
        ry = _np.array([c[3] for c in cps])
        cx, *_ = _np.linalg.lstsq(A, rx, rcond=None)
        cy, *_ = _np.linalg.lstsq(A, ry, rcond=None)
        return cx, cy

    cx0, cy0 = _fit_affine(initial_cps)

    def _aff0(x, y):
        return (cx0[0] * x + cx0[1] * y + cx0[2],
                cy0[0] * x + cy0[1] * y + cy0[2])

    # 4) Pass 2: per-chip nearest-building.
    import math as _math
    pass2_cps = []
    for sp in spans:
        real = _match_street(sp["text"])
        if not real:
            continue
        rx0, ry0 = _aff0(sp["cx"], sp["cy"])
        cand = min(street_pts[real],
                   key=lambda p: _math.hypot(p[0] - rx0, p[1] - ry0))
        pass2_cps.append((sp["cx"], sp["cy"], cand[0], cand[1]))

    cx1, cy1 = _fit_affine(pass2_cps)

    def _aff1(x, y):
        return (cx1[0] * x + cx1[1] * y + cx1[2],
                cy1[0] * x + cy1[1] * y + cy1[2])

    # 5) Pass 3: blue-boundary polylines as additional CPs.
    def _is_main_blue(c):
        return (isinstance(c, (tuple, list)) and len(c) >= 3
                and abs(c[0] - 0.182) < 0.03
                and abs(c[1] - 0.371) < 0.03
                and abs(c[2] - 0.674) < 0.03)

    boundary_cps = []
    boundary_geom = _get_sub_boundary_geom()
    if boundary_geom is not None and _shp_nearest_points is not None:
        try:
            drawings = page.get_drawings() or []
        except Exception:
            drawings = []
        for d in drawings:
            if not _is_main_blue(d.get("color")):
                continue
            items = d.get("items", [])
            if len(items) < 5:
                continue
            pts = []
            for it in items:
                op = it[0]
                args = it[1:]
                try:
                    if op == 'l' and len(args) >= 2:
                        for p in args:
                            pts.append((p.x, p.y))
                    elif op == 'c' and len(args) >= 4:
                        pts.append((args[0].x, args[0].y))
                        pts.append((args[3].x, args[3].y))
                except Exception:
                    continue
            for (px, py) in pts:
                rough = ShPoint(*_aff1(px, py))
                try:
                    _, snap = _shp_nearest_points(rough, boundary_geom)
                except Exception:
                    continue
                # Accept if rough→snap distance is small (≤50m)
                if rough.distance(snap) < 0.0005:
                    boundary_cps.append((px, py, snap.x, snap.y))

    combined = pass2_cps + boundary_cps

    # 6) Final fit — prefer TPS (scipy) for sub-meter accuracy, else affine.
    if _RBFInterpolator is not None and len(combined) >= 12:
        try:
            pts_pdf = _np.array([(c[0], c[1]) for c in combined])
            rx = _np.array([c[2] for c in combined])
            ry = _np.array([c[3] for c in combined])
            rbf_x = _RBFInterpolator(pts_pdf, rx, kernel='thin_plate_spline', smoothing=1.0)
            rbf_y = _RBFInterpolator(pts_pdf, ry, kernel='thin_plate_spline', smoothing=1.0)

            def _tps(x, y):
                arr = _np.array([[x, y]])
                return (float(rbf_x(arr)[0]), float(rbf_y(arr)[0]))

            # Quick residual sample for the log
            errs = []
            for cp in combined:
                rxp, ryp = _tps(cp[0], cp[1])
                errs.append((((rxp - cp[2]) ** 2 + (ryp - cp[3]) ** 2) ** 0.5) * 95000)
            errs.sort()
            print(f"      tashrir georef (TPS): {len(combined)} CPs "
                  f"({len(pass2_cps)} street + {len(boundary_cps)} boundary), "
                  f"median={errs[len(errs)//2]:.1f}m, max={errs[-1]:.1f}m")
            return _tps
        except Exception as e:
            print(f"      tashrir georef: TPS failed ({e}), falling back to affine")

    # Affine fallback (iterative outlier removal as before)
    cur = list(combined)
    for _ in range(3):
        cxF, cyF = _fit_affine(cur)
        residuals = []
        for cp in cur:
            rxp = cxF[0] * cp[0] + cxF[1] * cp[1] + cxF[2]
            ryp = cyF[0] * cp[0] + cyF[1] * cp[1] + cyF[2]
            err = ((rxp - cp[2]) ** 2 + (ryp - cp[3]) ** 2) ** 0.5
            residuals.append((err, cp))
        residuals.sort()
        median_err = residuals[len(residuals) // 2][0]
        threshold = max(0.0001, median_err * 2.5)
        keep = [cp for err, cp in residuals if err < threshold]
        if len(keep) == len(cur) or len(keep) < 6:
            break
        cur = keep
    cxF, cyF = _fit_affine(cur)
    errs = []
    for cp in cur:
        rxp = cxF[0] * cp[0] + cxF[1] * cp[1] + cxF[2]
        ryp = cyF[0] * cp[0] + cyF[1] * cp[1] + cyF[2]
        errs.append((((rxp - cp[2]) ** 2 + (ryp - cp[3]) ** 2) ** 0.5) * 95000)
    errs.sort()
    print(f"      tashrir georef (affine): {len(cur)} CPs, median={errs[len(errs)//2]:.1f}m, max={errs[-1]:.1f}m")

    def _aff(x, y):
        return (cxF[0] * x + cxF[1] * y + cxF[2],
                cyF[0] * x + cyF[1] * y + cyF[2])
    return _aff


# --- Tashrir (sal projects map) georeferencing ----------------------------
# The "סל פרויקטים תשריט גוננים" PDF has numbered markers placed on a map.
# We use street-name labels as control points, fit an affine transform from
# PDF coordinates to (lon, lat), and produce {project_num: [(lon, lat), ...]}.
# Each project_num may appear 2-3 times (one per domain on the same PDF).
TASHRIR_PDF = PDF_DIR / "סל פרויקטים_תשריט_גוננים.pdf"


def _build_tashrir_marker_index(buildings_features: list, minahak_geom_wkt: dict) -> dict:
    """Returns {project_num: [(lon, lat), ...]} of markers from the gonenim tashrir PDF.
    Returns {} if PDF or PyMuPDF unavailable, or if no usable control points."""
    if fitz is None or shape is None or _np is None or not TASHRIR_PDF.exists():
        return {}
    try:
        doc = fitz.open(str(TASHRIR_PDF))
        page = doc[0]
    except Exception:
        return {}

    # Build the high-accuracy transform (TPS + boundary CPs) once.
    minahak_geom = shape(minahak_geom_wkt)
    _transform = _build_pdf_to_wgs_transform(page, buildings_features, minahak_geom)
    if _transform is None:
        return {}

    # Re-collect spans (needed for numbered-marker extraction below).
    spans = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                txt = (span.get("text") or "").strip()
                if not txt:
                    continue
                bbox = span.get("bbox")
                spans.append({
                    "text": txt,
                    "cx": (bbox[0] + bbox[2]) / 2,
                    "cy": (bbox[1] + bbox[3]) / 2,
                })

    from collections import defaultdict as _dd

    # Transform all numbered spans to (lon, lat). Reject any outside the minahak (likely outliers).
    markers: dict[str, list] = _dd(list)
    for sp in spans:
        if re.fullmatch(r"\d{1,3}", sp["text"]):
            lon, lat = _transform(sp["cx"], sp["cy"])
            if minahak_geom.contains(ShPoint(lon, lat)):
                markers[sp["text"]].append((lon, lat))
    return dict(markers)


# Prefix in the תשריט labels → xlsx area name.
TASHRIR_PREFIX_TO_AREA = {
    "ר": "רסקו",
    "ג": "גוננים א-ו",
    "ק": "קטמונים ח-ט",
    "פ": "פת",
}


def _build_tashrir_typed_markers(buildings_features: list, minahak_geom_wkt: dict, roads_features: list | None = None) -> dict:
    """Return a richer index keyed by (area, type, num) → (lon, lat).

    The תשריט labels look like:
        "ג חציה 4"  → (גוננים א-ו, חציה,  4)
        "ק מעבר 5"  → (קטמונים ח-ט, מעבר, 5)
        "ר חציה 7"  → (רסקו,        חציה,  7)
        "פ מעבר 9"  → (פת,           מעבר,  9)
    Plus same-span variants like "5 ר חציה" or stuck-together pairs like
    "4 ק מעבר5 ק מעבר". We pair separated 'type' and 'digit' spans by
    spatial proximity (≤30px) and re-use the affine transform fit by the
    existing _build_tashrir_marker_index pipeline.
    """
    if fitz is None or shape is None or _np is None or not TASHRIR_PDF.exists():
        return {}
    try:
        doc = fitz.open(str(TASHRIR_PDF))
        page = doc[0]
    except Exception:
        return {}

    # 1) Collect all spans on the page (text + centroid).
    spans = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for sp in line.get("spans", []):
                txt = (sp.get("text") or "").strip()
                if not txt:
                    continue
                bbox = sp.get("bbox")
                spans.append({
                    "text": txt,
                    "cx": (bbox[0] + bbox[2]) / 2,
                    "cy": (bbox[1] + bbox[3]) / 2,
                })

    # 2) Build the high-accuracy transform (TPS + boundary CPs).
    minahak_geom = shape(minahak_geom_wkt)
    _transform = _build_pdf_to_wgs_transform(page, buildings_features, minahak_geom)
    if _transform is None:
        return {}

    # 3) Extract typed markers — by span text patterns.
    patternA = re.compile(r"^\s*(?P<num1>\d+)?\s*(?P<prefix>[רגקפ])\s+(?P<type>חציה|מעבר)\s*(?P<num2>\d+)?\s*$")
    patternB_type = re.compile(r"^\s*(?P<prefix>[רגקפ])\s+(?P<type>חציה|מעבר)\s*$")
    patternMulti  = re.compile(r"(?P<num>\d+)\s*(?P<prefix>[רגקפ])\s+(?P<type>חציה|מעבר)")

    markers_pdf: list[tuple] = []  # (area, type, num, cx, cy)
    consumed_indices: set[int] = set()

    for i, sp in enumerate(spans):
        t = sp["text"]
        # Multi-marker fragment (e.g. "4 ק מעבר5 ק מעבר")
        ms = list(patternMulti.finditer(t))
        if len(ms) >= 2 or (ms and "מעבר" in t and len(t) > 12):
            for m in ms:
                num = int(m.group("num"))
                area = TASHRIR_PREFIX_TO_AREA.get(m.group("prefix"))
                if area:
                    markers_pdf.append((area, m.group("type"), num, sp["cx"], sp["cy"]))
            consumed_indices.add(i)
            continue
        m = patternA.match(t)
        if m:
            num = m.group("num1") or m.group("num2")
            if num is None:
                continue
            num = int(num)
            area = TASHRIR_PREFIX_TO_AREA.get(m.group("prefix"))
            if area:
                markers_pdf.append((area, m.group("type"), num, sp["cx"], sp["cy"]))
            consumed_indices.add(i)

    # 4) Pair separated type-only spans with nearest digit span (≤30px).
    type_only_idx = [i for i, sp in enumerate(spans)
                     if i not in consumed_indices and patternB_type.match(sp["text"])]
    digit_idx = [i for i, sp in enumerate(spans)
                 if i not in consumed_indices and re.fullmatch(r"\d{1,3}", sp["text"])]
    used_digits: set[int] = set()
    import math as _math
    for ti in type_only_idx:
        tsp = spans[ti]
        m = patternB_type.match(tsp["text"])
        area = TASHRIR_PREFIX_TO_AREA.get(m.group("prefix"))
        if not area:
            continue
        best, best_d = None, 30.0
        for di in digit_idx:
            if di in used_digits:
                continue
            d = _math.hypot(spans[di]["cx"] - tsp["cx"], spans[di]["cy"] - tsp["cy"])
            if d < best_d:
                best_d = d
                best = di
        if best is not None:
            used_digits.add(best)
            num = int(spans[best]["text"])
            markers_pdf.append((area, m.group("type"), num, tsp["cx"], tsp["cy"]))

    # 5) Transform to (lon, lat). Keep only points inside the minahak (with a
    # small buffer for chips near the boundary — transform has 0-10m of drift
    # at the edges which can push otherwise-valid chips just outside).
    typed_index: dict[tuple, tuple] = {}
    # ~80m buffer in degrees at this latitude
    minahak_buffered = minahak_geom.buffer(0.00075)
    for area, typ, num, cx, cy in markers_pdf:
        lon, lat = _transform(cx, cy)
        if not minahak_buffered.contains(ShPoint(lon, lat)):
            continue
        key = (area, typ, num)
        if key in typed_index:
            # Duplicate: average with existing (helps cancel small text-position offsets)
            ex_lon, ex_lat = typed_index[key]
            typed_index[key] = ((ex_lon + lon) / 2, (ex_lat + lat) / 2)
        else:
            typed_index[key] = (lon, lat)

    # 6) ALSO extract the green passage POLYLINES drawn on the תשריט. We then
    # match each passage marker (X מעבר N) to the nearest path and store the
    # actual line geometry so the build can render the passage at its true
    # length instead of a synthetic 60m default.
    typed_index["__passage_lines__"] = _extract_passage_lines(
        page, markers_pdf, _transform, minahak_geom, roads_features
    )
    return typed_index


def _is_passage_green(c) -> bool:
    """Match any of the projector's passage greens — dark, medium, or yellow-green."""
    if not isinstance(c, (tuple, list)) or len(c) < 3:
        return False
    r, g, b = c[0], c[1], c[2]
    is_dark = abs(r - 0.15) < 0.10 and abs(g - 0.47) < 0.10 and abs(b - 0.23) < 0.10
    is_med  = abs(r - 0.40) < 0.10 and abs(g - 0.70) < 0.10 and abs(b - 0.42) < 0.10
    is_yellow_green = (0.50 <= r <= 0.65) and (0.70 <= g <= 0.85) and (0.15 <= b <= 0.35)
    return is_dark or is_med or is_yellow_green


def _extract_passage_lines(page, markers_pdf: list, transform_fn, minahak_geom,
                            roads_features: list | None = None) -> dict:
    """For each (area, type='מעבר', num) chip in the תשריט, generate a road-following
    LineString centered on the chip with bearing derived from the nearest green polyline.

    Algorithm (per chip):
      1. ANCHOR  ← transform_fn(chip_pdf_xy)
      2. BEARING ← principal-axis direction of nearest green polyline (PCA on its vertices,
                   projected to WGS84). If no nearby green polyline, BEARING = None.
      3. SNAP    ← snap ANCHOR to the nearest road within ~80m.
      4. WALK    ← from SNAP along the road most aligned with BEARING, ±80m default.
                   (Consumer clips to xlsx units_required.)

    Returns dict: (area, 'מעבר', num) → list of [lon, lat] coords (LineString).
    """
    out: dict[tuple, list] = {}
    try:
        drawings = page.get_drawings() or []
    except Exception:
        return out

    def _path_points(d):
        pts = []
        for it in d.get("items", []):
            op = it[0]
            args = it[1:]
            try:
                if op == 'l' and len(args) >= 2:
                    for p in args:
                        pts.append((p.x, p.y))
                elif op == 'c' and len(args) >= 4:
                    pts.append((args[0].x, args[0].y))
                    pts.append((args[3].x, args[3].y))
            except Exception:
                continue
        clean = []
        for p in pts:
            if not clean or (abs(clean[-1][0] - p[0]) > 0.01 or abs(clean[-1][1] - p[1]) > 0.01):
                clean.append(p)
        return clean

    green_paths = []
    for d in drawings:
        if _is_passage_green(d.get("color")) and len(d.get("items", [])) >= 1:
            pts = _path_points(d)
            if len(pts) >= 2:
                green_paths.append(pts)

    import math as _math

    try:
        from shapely.geometry import LineString as ShLine, Point as ShP
        from shapely.ops import substring as ShSubstring
    except Exception:
        ShLine = None

    DEFAULT_WALK_M = 400.0  # half-length; consumer clips to xlsx units_required.
    # Generous (800m total) so that after vertex-snap shrinks the line (often by 40-60%),
    # there's still enough length for the consumer to clip to spec — even for the longest
    # passage targets in the dataset (450m).

    # Same 80m buffer as the typed-index check so edge chips aren't rejected.
    minahak_buffered = minahak_geom.buffer(0.00075)

    for area, typ, num, cx, cy in markers_pdf:
        if typ != "מעבר":
            continue
        key = (area, "מעבר", num)
        if key in out:
            continue

        # Step 1: ANCHOR — chip position in WGS84.
        anchor_lon, anchor_lat = transform_fn(cx, cy)
        if not minahak_buffered.contains(ShPoint(anchor_lon, anchor_lat)):
            continue

        # Step 2: BEARING — principal axis of nearest green polyline.
        # Use a hybrid score (min + mean distance) so a long polyline that merely grazes
        # the chip loses to a short polyline centered nearby.
        best_path = None
        best_score = float("inf")
        for path in green_paths:
            dists = [_math.hypot(p[0] - cx, p[1] - cy) for p in path]
            min_d = min(dists)
            if min_d > 80.0:
                continue
            mean_d = sum(dists) / len(dists)
            score = min_d * 0.3 + mean_d * 0.7
            if score < best_score:
                best_score = score
                best_path = path

        bearing_lon_dir, bearing_lat_dir = None, None
        if best_path and len(best_path) >= 2:
            xs = [p[0] for p in best_path]
            ys = [p[1] for p in best_path]
            mx = sum(xs) / len(xs)
            my = sum(ys) / len(ys)
            sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
            sx2 = sum((x - mx) ** 2 for x in xs)
            sy2 = sum((y - my) ** 2 for y in ys)
            if (sx2 + sy2) > 1e-9:
                theta = 0.5 * _math.atan2(2 * sxy, sx2 - sy2)
                bx_pdf = _math.cos(theta)
                by_pdf = _math.sin(theta)
                # Convert PDF-direction to WGS84-direction via a small probe step.
                step = 5.0
                lon1, lat1 = transform_fn(cx - step * bx_pdf, cy - step * by_pdf)
                lon2, lat2 = transform_fn(cx + step * bx_pdf, cy + step * by_pdf)
                bdx = lon2 - lon1
                bdy = lat2 - lat1
                mag = _math.hypot(bdx, bdy)
                if mag > 1e-12:
                    bearing_lon_dir = bdx / mag
                    bearing_lat_dir = bdy / mag

        # Step 3: Emit a straight 3-point line centered on the chip ANCHOR (no road snap).
        # Pedestrian passages often cut BETWEEN streets rather than along existing roads,
        # so snapping to the nearest car-road would mis-place case 2 passages onto a
        # parallel street. The downstream consumer (build_features loop) performs its own
        # snap_point_to_nearest_road with a tight tolerance (~25m) which leaves off-road
        # vertices in place — exactly the behavior we want here.
        line_coords = None
        if bearing_lon_dir is not None:
            cos_lat = max(0.1, _math.cos(_math.radians(anchor_lat)))
            half_deg = (DEFAULT_WALK_M / 2.0) / 111000.0 / cos_lat
            line_coords = [
                [round(anchor_lon - half_deg * bearing_lon_dir, 7),
                 round(anchor_lat - half_deg * bearing_lat_dir, 7)],
                [round(anchor_lon, 7), round(anchor_lat, 7)],
                [round(anchor_lon + half_deg * bearing_lon_dir, 7),
                 round(anchor_lat + half_deg * bearing_lat_dir, 7)],
            ]

        if line_coords and len(line_coords) >= 2:
            out[key] = line_coords
    return out


# --- Online geocoding (Nominatim) -----------------------------------------
# We persist results to GEOCODE_CACHE_PATH so re-runs are fast and identical.
# Set NOMINATIM_ENABLED=False to skip online lookups (offline build).
NOMINATIM_ENABLED = True
NOMINATIM_USER_AGENT = "oranim-projector-builder/1.0"

_nominatim_session_calls = 0


def _load_geocode_cache() -> dict:
    if GEOCODE_CACHE_PATH.exists():
        try:
            with GEOCODE_CACHE_PATH.open(encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_geocode_cache(cache: dict) -> None:
    try:
        with GEOCODE_CACHE_PATH.open("w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        # Also mirror to oranim-map/data/
        other = DATA_DIRS[0] / GEOCODE_CACHE_PATH.name
        with other.open("w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _query_nominatim(query: str) -> tuple[float, float] | None:
    """Single-shot Nominatim search. Returns (lat, lon) or None.
    Respects 1-req/sec rate limit per Nominatim policy."""
    if not NOMINATIM_ENABLED:
        return None
    global _nominatim_session_calls
    import urllib.parse
    import urllib.request
    import time as _t
    if _nominatim_session_calls > 0:
        _t.sleep(1.2)
    _nominatim_session_calls += 1
    url = (
        "https://nominatim.openstreetmap.org/search?"
        f"q={urllib.parse.quote(query)}&format=json&limit=1&accept-language=he"
    )
    req = urllib.request.Request(url, headers={"User-Agent": NOMINATIM_USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read())
            if data:
                # Sanity check: result must be in/near Jerusalem (lat ~31.7, lon ~35.2).
                lat, lon = float(data[0]["lat"]), float(data[0]["lon"])
                if 31.5 <= lat <= 32.0 and 34.9 <= lon <= 35.4:
                    return (lat, lon)
    except Exception:
        return None
    return None


def _geocode_via_cache_or_nominatim(project_name: str, sub_neigh: str, cache: dict) -> dict | None:
    """Try (1) the cache, then (2) Nominatim with a few query variations."""
    if not project_name:
        return None
    key = f"{project_name}|{sub_neigh}"
    # Cache: explicit hit (with lat/lon) → use; explicit null → don't re-query.
    if key in cache:
        entry = cache[key]
        if isinstance(entry, dict) and entry.get("lat") is not None and entry.get("lon") is not None:
            return {"type": "Point", "coordinates": [float(entry["lon"]), float(entry["lat"])]}
        return None
    # Online lookup: a few variants. Stop at first hit.
    candidates = [
        f"{project_name} {sub_neigh} ירושלים",
        f"{project_name} ירושלים",
        f"{project_name}",
    ]
    for q in candidates:
        result = _query_nominatim(q)
        if result:
            cache[key] = {"lat": result[0], "lon": result[1], "source": f"nominatim:{q}"}
            return {"type": "Point", "coordinates": [result[1], result[0]]}
    # Record miss so we don't re-query.
    cache[key] = {"lat": None, "lon": None, "source": "nominatim_miss"}
    return None


def _roads_centroid(road_features: list) -> list | None:
    """Compute centroid across all road LineString points for a street name."""
    pts = []
    for f in road_features:
        g = f.get("geometry") or {}
        t = g.get("type")
        if t == "LineString":
            pts.extend(g["coordinates"])
        elif t == "MultiLineString":
            for line in g["coordinates"]:
                pts.extend(line)
    if not pts:
        return None
    return [sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts)]


# --- Structured parser for the 3 "מרחב ציבורי" booklets ----------------
# Each recommendation page contains structured fields:
#   "מיקום: ..." (location), "חשיבות: ..." (importance),
#   "ההמלצה: ..." (recommendation), "ייעודי קרקע: ...", "זמינות לעבודה: ..."
# We parse them into a dict keyed by (area_name_in_xlsx, recommendation_num).
PUBLIC_SPACE_BOOKLETS = [
    ("גוננים א-ו", PDF_DIR / "גוננים מרחב ציבורי_05.2024.pdf"),
    ("רסקו",       PDF_DIR / "רסקו מרחב ציבורי_17.4.24.pdf"),
    # קטמונים ופת share one booklet — pages mix both areas; we register both.
    ("קטמונים ח-ט", PDF_DIR / "קטמונים ופת 05.2024.pdf"),
    ("פת",         PDF_DIR / "קטמונים ופת 05.2024.pdf"),
]
_PS_FIELD_LABELS = {
    "מיקום": "location",
    "חשיבות": "importance",
    "חשיבות המעבר": "importance",
    "חשיבות המיקום": "importance",
    "ההמלצה": "recommendation",
    "ההצעה": "recommendation",
    "המלצה": "recommendation",
    "מאפיין ייחודי": "unique_feature",
    "ייעודי קרקע": "land_use",
    "מידות": "dimensions",
    "זמינות לעבודה": "availability",
    "סטטוס": "status",
    "הצורך": "need",
    "ההזדמנות": "opportunity",
}

# Icon-label keywords that appear as standalone phrases above the body text on each
# recommendation page. Extracted as `categories: [...]` for popup badge display.
# Order matters: longer/more-specific phrases first to avoid partial-overlap mis-matches.
_PS_CATEGORY_LABELS = [
    "הנגשה מכאנית",
    "הסדרה סטטוטורית",
    "הצעה מתחמית",
    "פינת ישיבה",
    "יצירת מקום",
    "מעבר חדש",
    "שימור",
    "תצפית",
    "שדרוג",
    "הסדרה",
    "הנגשה",
]


def _parse_public_space_pages(pdf_path: Path) -> dict:
    """Returns {recommendation_num_str: {title, location, importance, recommendation, land_use,
    dimensions, availability, page}} for all numbered recommendation pages in the PDF."""
    if fitz is None or not pdf_path.exists():
        return {}
    try:
        doc = fitz.open(str(pdf_path))
    except Exception:
        return {}
    field_pattern = "|".join(re.escape(k) for k in _PS_FIELD_LABELS.keys())
    out: dict = {}
    for i in range(doc.page_count):
        try:
            text = doc[i].get_text("text") or ""
        except Exception:
            continue
        # Only pages that look like recommendation pages
        if "מיקום" not in text and "ההמלצה" not in text and "ההצעה" not in text:
            continue
        # Skip if not in המלצות section (headers say "פרק ג'" or similar)
        # Identify the recommendation number — the title line is "<title text>N" (Hebrew RTL).
        head = text[:400]
        # Two patterns:
        #   "1. שדרוג ..."   (number first, English-style)
        #   ". שדרוג ... 1"  (Hebrew RTL — number ends the title line; may be followed
        #                     directly by the next icon-label without whitespace)
        m = re.search(r"\.\s+([א-ת][^\n]{2,60}?)(\d+)(?=\s|$|[א-ת])", head)
        if m:
            num = m.group(2).strip()
            title = m.group(1).strip()
        else:
            m = re.search(r"\b(\d+)\s*\.\s+([א-ת][^\n]{2,80})", head)
            if m:
                num = m.group(1).strip()
                title = m.group(2).strip()
            else:
                continue
        # Extract structured fields. We require labels at LINE START — mid-line labels
        # in this PDF are caused by RTL text-extraction artifacts and produce garbled values.
        # Some fields like "מידות" / "סטטוס" are in mixed-line layouts that we drop here;
        # they're better consumed via dedicated regex on the full text below.
        fields = {}
        chunks = re.split(rf"(?:^|\n)\s*({field_pattern})\s*:\s*", text)
        # chunks[0] is text before first label; the rest alternate: label, value, label, value...
        for ci in range(1, len(chunks) - 1, 2):
            label = chunks[ci].strip()
            value = chunks[ci + 1] if ci + 1 < len(chunks) else ""
            # Take lines until: blank line OR a line that contains another colon (likely a
            # mid-line field label for a different field — RTL extraction artifact).
            value_lines = []
            for ln in value.split("\n"):
                ln_stripped = ln.strip()
                if not ln_stripped:
                    if value_lines:
                        break
                    continue
                # If a line contains ANY of our known labels followed by ":" (mid-line),
                # stop here — that's a different field starting.
                if any(other_label + ":" in ln for other_label in _PS_FIELD_LABELS):
                    if value_lines:
                        break
                    # Drop the leading mid-line garbage and stop the rest
                    continue
                if len(ln_stripped) < 3:
                    continue
                if re.match(r"^[\d\s\.,/]+$", ln_stripped):
                    continue
                value_lines.append(ln_stripped)
                # 1-2 lines is usually enough for these fields
                if len(" ".join(value_lines)) > 200:
                    break
            value = " ".join(value_lines).strip()
            value = re.sub(r"\s+", " ", value)
            canonical = _PS_FIELD_LABELS.get(label, label)
            if canonical not in fields and value:
                fields[canonical] = value
        # Mid-line fields (RTL artifact: PDF extraction puts them at end of a mixed-content line).
        # For "מידות:" and "סטטוס:" the value is interleaved with adjacent text — capture the
        # FULL line containing the label and remove the label, leaving the raw dimension text.
        for mid_label in ("מידות", "סטטוס"):
            canonical = _PS_FIELD_LABELS.get(mid_label, mid_label)
            if canonical in fields:
                continue
            for line in text.splitlines():
                if mid_label + ":" not in line:
                    continue
                # Strip the label itself
                cleaned = line.replace(mid_label + ":", "").strip()
                cleaned = re.sub(r"\s+", " ", cleaned)
                if cleaned and len(cleaned) > 2:
                    fields[canonical] = cleaned
                    break

        # Extract icon-categories: standalone short phrases that appear between the title
        # and the first field, indicating intervention types (שדרוג / פינת ישיבה / תצפית etc).
        # Take the chunk before the first field-label colon, and scan for known labels.
        first_label_pos = len(text)
        for label in _PS_FIELD_LABELS:
            pos = text.find(label + ":")
            if pos > 0 and pos < first_label_pos:
                first_label_pos = pos
        # The PDF often splits multi-word icon labels across lines (e.g. "פינת\nישיבה",
        # "הנגשה\nמכאנית"). Replace internal whitespace runs with single spaces in the
        # search haystack so multi-word labels match. Also include the FULL page text —
        # some icons appear after the field-block (small icon row at bottom).
        head_for_cats = re.sub(r"\s+", " ", text)
        cats: list[str] = []
        for cat in _PS_CATEGORY_LABELS:
            if re.search(r"(?<![א-ת])" + re.escape(cat) + r"(?![א-ת])", head_for_cats):
                if cat not in cats:
                    cats.append(cat)
        # De-dup: if "הנגשה מכאנית" matched, drop the bare "הנגשה".
        if "הנגשה מכאנית" in cats and "הנגשה" in cats:
            cats.remove("הנגשה")
        if "הסדרה סטטוטורית" in cats and "הסדרה" in cats:
            cats.remove("הסדרה")

        rec = {"page": i + 1, "title": title, "categories": cats, **fields}
        # Keep the FIRST occurrence per number (the canonical recommendation page,
        # not the second-page detail).
        if num not in out:
            out[num] = rec
    return out


def _build_public_space_lookup() -> dict:
    """Builds a {(area, rec_num): rec_dict} lookup from all 3 booklets."""
    lookup: dict = {}
    for area, path in PUBLIC_SPACE_BOOKLETS:
        recs = _parse_public_space_pages(path)
        for num, rec in recs.items():
            key = (area, num)
            if key not in lookup:  # first booklet wins for shared (קטמונים+פת)
                lookup[key] = {**rec, "_area": area, "_pdf": path.name}
    return lookup


# --- Parser for the כרטסת שטחים חומים PDFs (programa cards) ---
# Each page is a structured card with these labeled fields:
#   רחוב, גוש/חלקה, תב"ע/מגרש, שטח המגרש (דונם), ייעוד ע"פ תב"ע,
#   שימושים מותרים [למבנים ומוסדות ציבור], מ"ר מבני ציבור, בעלות,
#   בינוי קיים במתחם, פוטנציאל פרוגרמתי, המלצות/הליך נדרש לתכנון
KARTESET_BOOKLETS = [
    PDF_DIR / "כרטסת שטחים חומים מרחב רסקו-גוננים א-ו ופת 01052024.pdf",
    PDF_DIR / "קטמון ח-ט - כרטסת שטחים חומים 06.062024.pdf",
]
_KARTESET_FIELD_LABELS = {
    "רחוב": "address",
    "גוש/חלקה": "gush_helka",
    "גוש / חלקה": "gush_helka",
    "תב\"ע/מגרש": "taba_lot",
    "תב\"ע / מגרש": "taba_lot",
    "שטח המגרש": "lot_area",
    "ייעוד ע\"פ תב\"ע": "land_use",
    "ייעוד עפ\"י תב\"ע": "land_use",
    "שימושים מותרים": "allowed_uses",
    "שימושים מותרים למבנים ומוסדות ציבור": "allowed_uses",
    "מ\"ר מבני ציבור": "public_sqm",
    "בעלות": "ownership",
    "בינוי קיים במתחם": "existing_built",
    "פוטנציאל פרוגרמתי": "programmatic_potential",
    "המלצות": "recommendations",
    "המלצות/הליך נדרש לתכנון": "recommendations",
    "הליך נדרש לתכנון": "recommendations",
}


def _parse_karteset_page(text: str) -> dict:
    """Extract structured fields from one karteset page. Uses per-field regexes to be
    resilient to PDF line wrapping (e.g. 'ייעוד ע\"\\n פ\\nתב\"ע:')."""
    fields: dict = {}
    # Normalize whitespace runs to single space; preserve newlines as soft markers
    norm = re.sub(r"[ \t]+", " ", text)
    # Strip the AVIV copyright line and the cover-style headers
    norm = re.sub(r"כל הזכויות שמורות[^\n]*", "", norm)
    norm = re.sub(r"אין להעתיק[^\n]*", "", norm)

    # ---- title: line right above the first labeled field ------------------
    m_t = re.search(r"\n\s*([^\n:]+?(?:\d+|[א-ת]+))\s*\n\s*רחוב", norm)
    if m_t:
        fields["title"] = m_t.group(1).strip()

    # ---- per-field regexes ------------------------------------------------
    def _grab(pattern, flags=re.DOTALL):
        m = re.search(pattern, norm, flags)
        if not m:
            return None
        v = m.group(1)
        v = re.sub(r"\s+", " ", v).strip()
        v = v.strip(" -–/:.")
        return v if v else None

    fields["address"] = _grab(r"רחוב\s*:?\s*([^\n]+)")
    # גוש/חלקה: <gush> / <helka>  (with optional newlines between)
    m_gh = re.search(r"גוש\s*/\s*חלקה\s*:?\s*(\d+)[\s/]+(\d+(?:\s*,\s*\d+)*)", norm)
    if m_gh:
        fields["gush_helka"] = f"גוש {m_gh.group(1)} / חלקה {m_gh.group(2).strip()}"
    # תב"ע/מגרש: <plan_number> / <lot>
    m_tl = re.search(r"תב[\"׳']ע\s*/\s*מגרש\s*:?\s*([\d\-/א-ת]+)[\s/]+(\d+(?:\s*,\s*\d+)*)", norm)
    if m_tl:
        fields["taba_lot"] = f"{m_tl.group(1).strip()} / מגרש {m_tl.group(2).strip()}"
    # שטח המגרש (דונם): X
    fields["lot_area"] = _grab(r"שטח\s*המגרש\s*\(?\s*דונם\s*:?\s*\)?\s*:?\s*([0-9.]+)\s*(?:דונם)?")
    # ייעוד ע"פ תב"ע: <value>  — the PDF often line-wraps the label itself
    # ("ייעוד ע\"\n פ\nתב\"ע:"); allow up to ~20 chars (incl. newlines) before "תב"ע".
    fields["land_use"] = _grab(r"ייעוד.{0,20}תב[\"׳']ע\s*:?\s*([^\n]+?)(?:\n|שימוש|מ\"ר|בעלות|בינוי|פוטנציאל|המלצ)", re.DOTALL)
    # שימושים מותרים [...]: multi-line up to next label
    fields["allowed_uses"] = _grab(r"שימושים\s*מותרים[^:]*:\s*(.+?)(?:\n\s*(?:מ\"ר|בעלות|בינוי|פוטנציאל|המלצ))")
    # מ"ר מבני ציבור: <value>
    fields["public_sqm"] = _grab(r"מ[\"׳']ר\s*מבני\s*ציבור\s*:?\s*(.+?)(?:\n\s*(?:בעלות|בינוי|פוטנציאל|המלצ))")
    # בעלות: <value>
    fields["ownership"] = _grab(r"בעלות\s*:?\s*([^\n]+?)(?:\n|בינוי|פוטנציאל|המלצ)")
    # בינוי קיים במתחם: multi-line
    fields["existing_built"] = _grab(r"בינוי\s*קיים\s*במתחם\s*:?\s*(.+?)(?:\n\s*(?:פוטנציאל|המלצ))")
    # פוטנציאל פרוגרמתי: multi-line
    fields["programmatic_potential"] = _grab(r"פוטנציאל\s*פרוגרמתי\s*:?\s*(.+?)(?:\n\s*(?:המלצ|הליך|\*|$))")
    # המלצות / הליך נדרש לתכנון: rest of text
    fields["recommendations"] = _grab(r"(?:המלצות[^:]*?|הליך\s*נדרש[^:]*?)\s*:?\s*(.+?)(?:\n\s*\*|$)")

    # Drop None values
    fields = {k: v for k, v in fields.items() if v}
    return fields


def _build_karteset_lookup() -> dict:
    """Returns {(pdf_filename, page_num_1indexed): {fields}} for ALL karteset pages."""
    if fitz is None:
        return {}
    out: dict = {}
    for path in KARTESET_BOOKLETS:
        if not path.exists():
            continue
        try:
            doc = fitz.open(str(path))
        except Exception:
            continue
        for i in range(doc.page_count):
            try:
                text = doc[i].get_text("text") or ""
            except Exception:
                continue
            # Skip non-content pages (no field labels)
            if "גוש" not in text and "תב" not in text:
                continue
            parsed = _parse_karteset_page(text)
            # Require at least 2 fields to count as a real card
            if len([k for k in parsed if k != "title"]) >= 2:
                out[(path.name, i + 1)] = parsed
    return out


# Path → (xlsx area name) mapping for the "תיק שכונה" / "דוח מסכם" booklets.
# Each is a per-neighborhood narrative deck with per-project slides whose titles
# look like "שקופית 35: רבי צדוק | סיכום והמלצות".
TIK_SHCHUNAH_PDFS = [
    # (area_name_in_xlsx, pdf_filename, label)
    ("גוננים א-ו",       PDF_DIR / "25-12-29 גוננים א-ו - תיק שכונה.pdf",       "תיק שכונה — גוננים א-ו"),
    ("רסקו",             PDF_DIR / "25-09-30 רסקו - תיק שכונה.pptx.pdf",        "תיק שכונה — רסקו"),
    # The ח-ט+פת file covers both sub-neighborhoods; we register both areas.
    ("קטמונים ח-ט",      PDF_DIR / "גוננים ח-ט ופת- דוח מסכם.pdf",              "דוח מסכם — קטמונים ח-ט ופת"),
    ("פת",               PDF_DIR / "גוננים ח-ט ופת- דוח מסכם.pdf",              "דוח מסכם — קטמונים ח-ט ופת"),
]

# Slides that are GENERIC (statistics / overview) — skip these because their titles
# are not project names.
_TIK_GENERIC_TITLE_KEYWORDS = (
    "מצב קיים", "מצב מתוכנן", "סכמת תנועה", "תחזיות תנועה", "ניתוח", "סקר",
    "התחדשות עירונית", "מיקום במרחב", "עוצמת שימושים", "מוקדי עניין",
    "רשת ", "סקר", "פערים והזדמנויות", "תכנון מאקרו", "תכנון מיקרו",
    "ניתוח SWOT", "כניסה", "תודה", "סיכום מסקנות", "סיכום המלצות",
    "מסלולי תחבורה", "יעדים ומסלולים", "תוספת מקומות חניה",
    "קווים נכנסים", "תחנות רק", "הליכה אסטרטגית", "אופניים אסטרטגית",
    "תח\"צ אסטרטגית", "רחובות אסטרטגית",
)


def _tik_title_is_project(title: str) -> bool:
    """Heuristic: is the slide title a project-specific title (vs. a general overview)?

    Project titles look like:
        "רבי צדוק | סיכום והמלצות"
        "אליעזר הגדול | חיבור לכביש הברך"
        "ניקנור | מצב מוצע"
        "חוות הנוער הציוני"
    """
    if not title:
        return False
    t = title.strip()
    # Strip leading "שקופית NN: " or "שקופית NN " prefix
    t = re.sub(r"^שקופית\s*\d+\s*[:\-]?\s*", "", t)
    if not t or len(t) < 3:
        return False
    if any(k in t for k in _TIK_GENERIC_TITLE_KEYWORDS):
        return False
    # If the title has no Hebrew letters, skip
    if not any('א' <= c <= 'ת' for c in t):
        return False
    return True


def _tik_extract_project_name(title: str) -> str:
    """Pull out the canonical project name from a slide title.

    "שקופית 35: רבי צדוק | סיכום והמלצות" → "רבי צדוק"
    "שקופית 14: חוות הנוער הציוני"          → "חוות הנוער הציוני"
    "שקופית 35: רבי צדוק| סיכום והמלצות"   → "רבי צדוק"
    """
    if not title:
        return ""
    t = re.sub(r"^שקופית\s*\d+\s*[:\-]?\s*", "", title.strip())
    # Split on '|' — first part is the project name
    if "|" in t:
        t = t.split("|", 1)[0]
    return t.strip()


def _build_tik_shchunah_index() -> dict:
    """Index per-project narrative slides from the "תיק שכונה" / "דוח מסכם" decks.

    Returns:
        dict (area_xlsx, normalized_name) → list of {
            "pdf_name": str, "page": int (1-indexed), "title": str,
            "label": str, "text": str (extracted slide body)
        }
    """
    if fitz is None:
        return {}
    out: dict = defaultdict(list)
    for area, path, label in TIK_SHCHUNAH_PDFS:
        if not path.exists():
            print(f"      [tik] missing: {path.name}")
            continue
        try:
            doc = fitz.open(str(path))
        except Exception as e:
            print(f"      [tik] open error {path.name}: {e}")
            continue
        toc = doc.get_toc() or []
        kept = 0
        # Track seen titles per file to avoid duplicate entries
        for entry in toc:
            try:
                lvl, title, page = entry[0], entry[1], entry[2]
            except Exception:
                continue
            if lvl < 2 and len(toc) > 5:
                continue  # skip section headers when subentries exist
            if not _tik_title_is_project(title):
                continue
            proj_name = _tik_extract_project_name(title)
            if not proj_name:
                continue
            page_idx = max(0, min(page - 1, doc.page_count - 1))
            try:
                text = doc[page_idx].get_text("text") or ""
            except Exception:
                text = ""
            # Squash huge whitespace runs; keep up to ~800 chars for popup
            text_clean = re.sub(r"[ \t]+", " ", text)
            text_clean = re.sub(r"\n{3,}", "\n\n", text_clean).strip()
            out[(area, _norm_shp_name(proj_name))].append({
                "pdf_name": path.name,
                "page": page,
                "title": title.strip(),
                "label": label,
                "text": text_clean[:1200],
            })
            kept += 1
        print(f"      [tik] {path.name}: indexed {kept} project slides")
        doc.close()
    return dict(out)


def _find_tik_match(area: str, project_name: str, tik_index: dict) -> dict | None:
    """Find a tik-shchunah slide that matches the xlsx row by area+project name.

    Match precedence:
      1) Exact normalized name match within the same area.
      2) Substring (slide title contains xlsx name or vice versa) within same area.
         Picks the longest overlap.
    """
    if not project_name:
        return None
    norm_xlsx = _norm_shp_name(project_name)
    direct = tik_index.get((area, norm_xlsx))
    if direct:
        return direct[0]
    best = None
    best_score = 0
    for (a, sn), entries in tik_index.items():
        if a != area or not sn:
            continue
        if sn in norm_xlsx or norm_xlsx in sn:
            score = min(len(sn), len(norm_xlsx))
            # Require a minimum overlap to avoid spurious 3-char matches.
            if score >= 4 and score > best_score:
                best_score = score
                best = entries[0]
    return best


def _format_karteset_excerpt(rec: dict, page: int) -> str:
    """Same '#N — title' + 'CATS:' format as the public-space parser, so the JS
    popup renderer treats it uniformly."""
    title = rec.get("title", "כרטיס פרויקט")
    parts = [f"#{page} — {title}"]
    for label_he, key in [
        ("🏠 כתובת",         "address"),
        ("📋 גוש/חלקה",      "gush_helka"),
        ("📑 תב\"ע/מגרש",   "taba_lot"),
        ("📐 שטח המגרש",     "lot_area"),
        ("🏛 ייעוד",          "land_use"),
        ("✓ שימושים מותרים", "allowed_uses"),
        ("🏗 מ\"ר ציבור",   "public_sqm"),
        ("🔑 בעלות",          "ownership"),
        ("🏢 בינוי קיים",     "existing_built"),
        ("🎯 פוטנציאל",       "programmatic_potential"),
        ("💡 המלצות",         "recommendations"),
    ]:
        v = rec.get(key)
        if v:
            parts.append(f"{label_he}: {v}")
    return "\n".join(parts)


def _format_structured_excerpt(rec: dict) -> str:
    """Format the parsed recommendation as a clean string for popup display.
    The popup builder detects this format by the leading '#N — title' line and
    renders fields as styled rows. Categories are emitted on a special CATS: line
    so the JS can render them as badges."""
    title = rec.get("title")
    parts = []
    if title:
        parts.append(f"#{rec.get('page', '?')} — {title}")
    cats = rec.get("categories") or []
    if cats:
        parts.append(f"CATS: {' | '.join(cats)}")
    for label_he, key in [
        ("📍 מיקום", "location"),
        ("⭐ חשיבות", "importance"),
        ("🎯 הצורך", "need"),
        ("💡 ההזדמנות", "opportunity"),
        ("✓ המלצה", "recommendation"),
        ("✨ מאפיין ייחודי", "unique_feature"),
        ("🗺 ייעודי קרקע", "land_use"),
        ("📐 מידות", "dimensions"),
        ("📌 סטטוס", "status"),
        ("⏱ זמינות", "availability"),
    ]:
        v = rec.get(key)
        if v:
            parts.append(f"{label_he}: {v}")
    return "\n".join(parts)


def _parse_pdf_for_location(excerpt: str) -> dict:
    """Parse the structured fields the כרטסת PDFs always include:
       'רחוב:NAME[ NUMBER]' and 'גוש/חלקה: GUSH / PARCELS'.
       Returns {'street', 'house_num', 'gush', 'parcels': [..]}."""
    result = {"street": None, "house_num": None, "gush": None, "parcels": []}
    if not excerpt:
        return result
    # רחוב: NAME [optional number]. The PDFs use "רחוב:NAME" with no space, sometimes with number.
    m = re.search(r"רחוב\s*:\s*([^\n]+)", excerpt)
    if m:
        line = m.group(1).strip()
        # Try to split "<street> <num>"
        m2 = re.match(r"^([א-ת'\"\s\.]+?)\s*(\d{1,3})\s*$", line)
        if m2:
            result["street"] = m2.group(1).strip()
            result["house_num"] = m2.group(2)
        else:
            result["street"] = line
    # גוש/חלקה may span multiple lines: "גוש/חלקה:\n30185\n /\n178,203"
    # Capture from "גוש/חלקה:" until "תב\"ע" or "שטח" or end.
    m3 = re.search(r"גוש\s*/\s*חלקה\s*:\s*([\s\S]+?)(?:תב|שטח|בעלות|בינוי|$)", excerpt)
    if m3:
        block = m3.group(1)
        # Pull "GUSH / HELKAS" pairs. Be permissive about whitespace.
        m4 = re.search(r"(\d{4,5})\s*/\s*([\d,\-\s]+?)(?:\n|$|/)", block)
        if m4:
            result["gush"] = m4.group(1)
            helkas_raw = m4.group(2)
            for part in re.split(r"[,\s]+", helkas_raw):
                part = part.strip()
                if part.isdigit():
                    result["parcels"].append(part)
                elif "-" in part:
                    try:
                        a, b = part.split("-")
                        a_i, b_i = int(a), int(b)
                        if a_i <= b_i and (b_i - a_i) <= 30:
                            result["parcels"].extend(str(x) for x in range(a_i, b_i + 1))
                    except ValueError:
                        pass
    return result


def _try_address_lookup(project_name, buildings_by_addr, buildings_by_street) -> dict | None:
    """Project names often look like 'יצחק קצנלסון 3' or 'דוד שמעוני 22, ירושלים'.
    Try to extract street + house_num and match against buildings.geojson. Returns Point geometry or None.
    Falls back to nearest house number on the same street, then to street centroid."""
    if not project_name:
        return None
    s = str(project_name)
    s = re.sub(r",\s*ירושלים\s*$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    m = re.match(r"^([א-ת'\"\s\.\-]+?)\s+(\d{1,3})(?:[\-\s].*)?$", s)
    if not m:
        return None
    street_n = re.sub(r"^רח['׳]\s*", "", m.group(1).strip()).strip()
    house_num = m.group(2)
    # Try exact match first
    for candidate in [(street_n, house_num),
                      (street_n.replace("'", "׳"), house_num),
                      (street_n.replace("׳", "'"), house_num)]:
        feats = buildings_by_addr.get(candidate)
        if feats:
            c = feature_centroid(feats[0])
            if c:
                return {"type": "Point", "coordinates": c}
    # Fallback A: nearest house number on the same street.
    # House numbers are usually two parallel sequences (even/odd). Pick the closest.
    target = int(house_num) if house_num.isdigit() else None
    if target is not None:
        for street_variant in [street_n, street_n.replace("'", "׳"), street_n.replace("׳", "'")]:
            houses = buildings_by_street.get(street_variant)
            if not houses:
                continue
            # Sort by absolute distance from target, prefer same parity (same side of street)
            same_parity = [h for h in houses if h[0].isdigit() and int(h[0]) % 2 == target % 2]
            other_parity = [h for h in houses if h[0].isdigit() and int(h[0]) % 2 != target % 2]
            for pool in (same_parity, other_parity):
                if not pool:
                    continue
                pool_sorted = sorted(pool, key=lambda h: abs(int(h[0]) - target))
                # Skip if the nearest is too far (>20 numbers off — different segment)
                if abs(int(pool_sorted[0][0]) - target) <= 20:
                    c = feature_centroid(pool_sorted[0][1])
                    if c:
                        return {"type": "Point", "coordinates": c}
    return None


def _parse_page_token(token: str) -> list[int]:
    """Parse '13' / '13-15' / '13,28' / '54-56' into list of integer page numbers (1-indexed)."""
    out = []
    for part in token.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                a_i, b_i = int(a), int(b)
                if a_i <= b_i and (b_i - a_i) <= 8:
                    out.extend(range(a_i, b_i + 1))
            except ValueError:
                pass
        else:
            try:
                out.append(int(part))
            except ValueError:
                pass
    return out


def extract_pdf_excerpt(pdf_link: str | None) -> str:
    """Resolve a 'קישורים' string like 'עמ' 13 חוברת מרחב ציבורי גוננים' to
    raw text from the relevant PDF pages. Returns empty string if not resolvable."""
    if not pdf_link or not isinstance(pdf_link, str):
        return ""
    s = pdf_link.strip()

    label, pdf_path = _resolve_booklet(s)
    if pdf_path is None or not pdf_path.exists():
        return ""

    booklet_idx = _build_booklet_index(pdf_path)
    n_pages = booklet_idx["n_pages"]

    pages0: list[int] = []  # 0-indexed page indices

    # 1) "עמ' N" / "עמ' 54-56" / "עמ' 9,28"
    m = _PAGE_REF_RE.search(s)
    if m:
        for pn in _parse_page_token(m.group(1)):
            if 0 <= pn - 1 < n_pages:
                pages0.append(pn - 1)

    # 2) "שקופית N" / "שקופיות 5-6"
    if not pages0:
        m2 = _SLIDE_RE.search(s)
        if m2:
            for pn in _parse_page_token(m2.group(1)):
                if 0 <= pn - 1 < n_pages:
                    pages0.append(pn - 1)

    # 3) "מתחם מס' N" — only meaningful for the תכנון מתחמי booklet
    if not pages0:
        m3 = re.search(r"מתחם\s*מס['׳]?\s*(\d+)", s)
        if m3 and label == "תכנון מתחמי":
            idx_n = int(m3.group(1))
            if idx_n in METAHAM_PAGE_RANGES:
                start, end, core = METAHAM_PAGE_RANGES[idx_n]
                for pn in [core, core + 1, core - 1, core + 2, core - 2]:
                    if start <= pn <= end and 0 <= pn - 1 < n_pages:
                        pages0.append(pn - 1)
                        if len(pages0) >= 4:
                            break

    # 4) Fallback for booklets with no page reference (e.g., 2-page בית שוויץ note)
    if not pages0:
        # Take all pages if the booklet is short, else first 3.
        if n_pages <= 4:
            pages0 = list(range(n_pages))
        else:
            pages0 = [0, 1, 2]

    # Preserve order while deduping
    seen = set()
    ordered = []
    for p in pages0:
        if p not in seen:
            seen.add(p)
            ordered.append(p)
    pieces = []
    for p in ordered[:5]:
        t = _page_text(pdf_path, p)
        if t:
            page_label = f"[עמ' {p + 1}]"
            pieces.append(f"{page_label}\n{t}")
    text = "\n\n— — —\n\n".join(pieces)
    if len(text) > 1800:
        text = text[:1800] + " …"
    return text


def normalize_chumash(chumash_text, year_completion) -> int | None:
    """Map xlsx 'חומש לביצוע' to 1/2/3 or None."""
    if not chumash_text:
        return None
    s = str(chumash_text).strip()
    # Direct digit
    m = re.search(r"חומש\s*(\d)", s)
    if m:
        return int(m.group(1))
    # Hebrew ordinals
    if "ראשון" in s:
        return 1
    if "שני" in s:
        return 2
    if "שלישי" in s:
        return 3
    if "מימוש" in s or "רלוונטי" in s:
        return None
    # Fallback to completion year
    try:
        y = int(year_completion)
        if y <= 2030:
            return 1
        if y <= 2035:
            return 2
        return 3
    except (TypeError, ValueError):
        return None


PLAN_NUM_RE = re.compile(r"\b(101-\d{7})\b")


def extract_plan_number(raw):
    """Extract canonical plan number 101-XXXXXXX or return None."""
    if not raw:
        return None
    m = PLAN_NUM_RE.search(str(raw))
    return m.group(1) if m else None


MIGRASH_RE = re.compile(r"(\d+(?:[.,\-]\d+)*)\s*/\s*(\d+)")


def extract_migrash_pairs(raw) -> list[tuple[str, str]]:
    """Return list of (helka_string, gush) tuples from a 'מזהה מיקום' value.

    Handles two formats commonly seen in the xlsx:
      1) "<helkas>/<gush>"     e.g. "13,17-19/30169"
      2) "גוש: <gush>, חלקות: <helka_list>"  e.g. "גוש: 30169, חלקות: 13,17,18,22,24"
    """
    if not raw:
        return []
    s = str(raw)
    out = []
    # Format 1: digit-slash-digit
    for m in MIGRASH_RE.finditer(s):
        out.append((m.group(1), m.group(2)))
    if out:
        return out
    # Format 2: "גוש: <gush>... חלקות: <helkas>"
    m_g = re.search(r"גוש\s*:?\s*(\d{3,5})", s)
    m_h = re.search(r"חלקות?\s*:?\s*([\d\s,\-]+)", s)
    if m_g and m_h:
        gush = m_g.group(1)
        helkas = m_h.group(1).strip().rstrip(",").strip()
        if helkas:
            out.append((helkas, gush))
    return out


def coerce_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def coerce_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# --- Geometry helpers (no shapely dep — keep stdlib) ---------------------
def polygon_centroid(rings):
    """Centroid of a Polygon (list of rings, each list of [x,y])."""
    total_a, cx, cy = 0.0, 0.0, 0.0
    if not rings or not rings[0]:
        return None
    outer = rings[0]
    for i in range(len(outer) - 1):
        x0, y0 = outer[i]
        x1, y1 = outer[i + 1]
        a = x0 * y1 - x1 * y0
        total_a += a
        cx += (x0 + x1) * a
        cy += (y0 + y1) * a
    if abs(total_a) < 1e-12:
        # degenerate — average vertices
        xs = [p[0] for p in outer]
        ys = [p[1] for p in outer]
        return [sum(xs) / len(xs), sum(ys) / len(ys)]
    total_a *= 0.5
    return [cx / (6 * total_a), cy / (6 * total_a)]


def feature_centroid(feat):
    g = feat.get("geometry") or {}
    t = g.get("type")
    coords = g.get("coordinates") or []
    if t == "Point":
        return list(coords)
    if t == "Polygon":
        return polygon_centroid(coords)
    if t == "MultiPolygon":
        # use centroid of largest polygon (simple bbox area proxy)
        best = None
        best_area = -1
        for poly in coords:
            if not poly or not poly[0]:
                continue
            xs = [p[0] for p in poly[0]]
            ys = [p[1] for p in poly[0]]
            area = (max(xs) - min(xs)) * (max(ys) - min(ys))
            if area > best_area:
                best_area = area
                best = poly
        return polygon_centroid(best) if best else None
    if t in ("LineString", "MultiLineString"):
        # rough midpoint
        if t == "LineString":
            pts = coords
        else:
            pts = [p for line in coords for p in line]
        if not pts:
            return None
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        return [sum(xs) / len(xs), sum(ys) / len(ys)]
    return None


def feature_display_point(geom):
    """Pick a [lon, lat] that is GUARANTEED to be inside (or on) the feature.

    Why this exists: a polygon's geometric centroid can fall OUTSIDE the
    polygon when the shape is concave (L-shapes, donut-like lots, multi-lot
    MultiPolygons whose pieces are spread apart). The result on the map was
    markers floating in the road / open space, away from the lot.

    Strategy:
      • Point/LineString → use centroid/midpoint (existing logic).
      • Polygon          → use shapely.representative_point (always inside).
      • MultiPolygon     → use representative_point of the LARGEST piece by
                           actual polygon area (not bbox). For multi-lot plans
                           this lands on the biggest contiguous lot, not in
                           the gap between pieces.
    """
    if not geom:
        return None
    t = geom.get("type")
    if t == "Point":
        return list(geom["coordinates"])
    # If shapely is unavailable, fall back to the existing logic.
    if shape is None:
        return feature_centroid({"geometry": geom})
    try:
        g = shape(geom)
    except Exception:
        return feature_centroid({"geometry": geom})
    try:
        if t == "MultiPolygon" and hasattr(g, "geoms"):
            biggest = max((p for p in g.geoms if not p.is_empty), key=lambda p: p.area, default=None)
            if biggest is None:
                return None
            pt = biggest.representative_point()
            return [round(pt.x, 7), round(pt.y, 7)]
        if t == "Polygon":
            pt = g.representative_point()
            return [round(pt.x, 7), round(pt.y, 7)]
        if t in ("LineString", "MultiLineString"):
            pt = g.interpolate(0.5, normalized=True)
            return [round(pt.x, 7), round(pt.y, 7)]
    except Exception:
        pass
    return feature_centroid({"geometry": geom})


def make_rect_polygon(lon: float, lat: float, w_m: float, h_m: float) -> dict:
    """Build a small rectangular Polygon centered on (lon, lat). Used for elongated
    features (passages, crossings, road segments) where direction matters."""
    import math
    dlat = (h_m / 2) / 111000.0
    dlon = (w_m / 2) / (111000.0 * max(0.1, math.cos(math.radians(lat))))
    ring = [
        [lon - dlon, lat - dlat],
        [lon + dlon, lat - dlat],
        [lon + dlon, lat + dlat],
        [lon - dlon, lat + dlat],
        [lon - dlon, lat - dlat],
    ]
    return {"type": "Polygon", "coordinates": [ring]}


def make_circle_polygon(lon: float, lat: float, radius_m: float, n: int = 24) -> dict:
    """Build an approximate circle as an n-vertex polygon."""
    import math
    cos_lat = max(0.1, math.cos(math.radians(lat)))
    dlat_per_m = 1.0 / 111000.0
    dlon_per_m = 1.0 / (111000.0 * cos_lat)
    ring = []
    for i in range(n):
        angle = 2 * math.pi * i / n
        dx = radius_m * math.cos(angle)
        dy = radius_m * math.sin(angle)
        ring.append([lon + dx * dlon_per_m, lat + dy * dlat_per_m])
    ring.append(ring[0])
    return {"type": "Polygon", "coordinates": [ring]}


def buffer_linestring(line_coords: list, width_m: float, lat0: float) -> dict:
    """Buffer a LineString [(lon, lat), ...] into a Polygon of the given width.
    Used to generate REAL street/crossing/passage polygons from roads.geojson —
    follows the actual road shape instead of a generic rectangle."""
    if shape is None:
        return None
    try:
        from shapely.geometry import LineString as ShLine
        ls = ShLine(line_coords)
        # Convert width meters to degrees (approximate at this latitude)
        import math
        cos_lat = max(0.1, math.cos(math.radians(lat0)))
        # Use lon-degrees as the buffer unit (pretend isotropic — for small distances OK)
        width_deg = (width_m / 2) / (111000.0 * cos_lat)
        polygon = ls.buffer(width_deg, cap_style=2, join_style=2)  # flat caps, mitred joins
        if polygon.is_empty:
            return None
        return mapping(polygon)
    except Exception:
        return None


def find_nearest_road_full(point: tuple, roads_features: list, max_dist_deg: float = 0.0008) -> list | None:
    """Find the nearest road LineString and return its FULL coordinates (no clip).
    Use when the caller wants to clip the road to a specific length itself.
    Returns coords list or None."""
    if shape is None:
        return None
    try:
        from shapely.geometry import LineString as ShLine, Point as P
        target = P(point[0], point[1])
        best = None
        best_dist = float("inf")
        for f in roads_features:
            g = f.get("geometry") or {}
            t = g.get("type")
            if t == "LineString":
                lines = [g["coordinates"]]
            elif t == "MultiLineString":
                lines = g["coordinates"]
            else:
                continue
            for coords in lines:
                if len(coords) < 2:
                    continue
                ls = ShLine(coords)
                d = ls.distance(target)
                if d < best_dist:
                    best_dist = d
                    best = ls
        if best is None or best_dist > max_dist_deg:
            return None
        return list(best.coords)
    except Exception:
        return None


def find_nearest_road_segment(point: tuple, roads_features: list, max_dist_deg: float = 0.0008) -> list | None:
    """Given a (lon, lat), find the nearest road LineString and return a clipped
    portion (~30m) around the nearest point on it. Returns coords list or None."""
    if shape is None:
        return None
    try:
        from shapely.geometry import LineString as ShLine, Point as P
        target = P(point[0], point[1])
        best = None
        best_dist = float("inf")
        for f in roads_features:
            g = f.get("geometry") or {}
            t = g.get("type")
            if t == "LineString":
                lines = [g["coordinates"]]
            elif t == "MultiLineString":
                lines = g["coordinates"]
            else:
                continue
            for coords in lines:
                if len(coords) < 2: continue
                ls = ShLine(coords)
                d = ls.distance(target)
                if d < best_dist:
                    best_dist = d
                    best = ls
        if best is None or best_dist > max_dist_deg:
            return None
        # Project the point onto the line and take a ±15m slice
        proj = best.project(target, normalized=False)
        # 30m in degrees
        import math
        cos_lat = max(0.1, math.cos(math.radians(point[1])))
        delta = 15.0 / 111000.0 / cos_lat
        start = max(0, proj - delta)
        end = min(best.length, proj + delta)
        try:
            from shapely.ops import substring
            seg = substring(best, start, end)
            return list(seg.coords)
        except Exception:
            # Fallback: use full line
            return list(best.coords)
    except Exception:
        return None


# --- Units → geometry-kind classifier --------------------------------------
# The xlsx 'יחידות' column tells us what the project is MEASURED in. We map the
# unit to a geometry hint:
#   • length units (מ' / מ"א / ק"מ) → "line"  (LineString)
#   • area units   (מ"ר / דונם)     → "area"  (Polygon)
#   • count units  (כיתות / קומפלט / יח') → "point" (single feature)
# Reference: ה"ר = הולכי רגל (pedestrians) — used in service names like
# "מעבר ה"ר" or "שביל ה"ר".
_UNITS_LINE  = {"מ'", "מ׳", "מ", "מ\"א", "מ׳׳א", "מא", "ק\"מ", "קמ"}
_UNITS_AREA  = {"מ\"ר", "מ׳׳ר", "מר", "דונם", "דונמים"}
_UNITS_POINT = {"כיתות", "קומפלט", "יח'", "יחידות", "יח׳"}


def units_geometry_kind(units_label) -> str | None:
    """Return 'line' / 'area' / 'point' / None based on the units column."""
    if units_label is None:
        return None
    s = str(units_label).strip().strip("_")
    if not s:
        return None
    # Normalize Hebrew quote variants
    norm = s.replace("׳", "'").replace("״", '"')
    if norm in _UNITS_LINE:  return "line"
    if norm in _UNITS_AREA:  return "area"
    if norm in _UNITS_POINT: return "point"
    # Fuzzy fallback: any 'מ' followed by ' or "א or "ר etc.
    if norm in {"מ'"} or norm.startswith("מ'") or norm.startswith("מ\""):
        return "area" if "ר" in norm else "line"
    return None


def snap_point_to_nearest_road(lon: float, lat: float, roads_features: list,
                               max_dist_deg: float = 0.0015) -> tuple[float, float] | None:
    """Project (lon, lat) onto the nearest road LineString and return the point
    that lies ON the road. Used for crosswalks so the marker sits on the street,
    not inside the building.

    max_dist_deg ≈ 0.0015 ≈ 150m at this latitude — a generous catch radius for
    cases where the centroid of a parcel sits a fair way from the curb.
    """
    if shape is None:
        return None
    try:
        from shapely.geometry import LineString as ShLine, Point as P
        target = P(lon, lat)
        best = None
        best_dist = float("inf")
        for f in roads_features:
            g = f.get("geometry") or {}
            t = g.get("type")
            if t == "LineString":
                lines = [g["coordinates"]]
            elif t == "MultiLineString":
                lines = g["coordinates"]
            else:
                continue
            for coords in lines:
                if len(coords) < 2:
                    continue
                ls = ShLine(coords)
                d = ls.distance(target)
                if d < best_dist:
                    best_dist = d
                    best = ls
        if best is None or best_dist > max_dist_deg:
            return None
        proj_d = best.project(target, normalized=False)
        snapped = best.interpolate(proj_d)
        return (round(snapped.x, 7), round(snapped.y, 7))
    except Exception:
        return None


def road_bearing_at(lon: float, lat: float, roads_features: list,
                    max_dist_deg: float = 0.0015) -> float | None:
    """Compass bearing (0=N, clockwise) of the nearest road segment at the
    projection of (lon, lat). Returned in [0, 360). The crosswalk renderer
    uses (bearing - 90°) to lay the rectangle ACROSS the road. None if no
    road within max_dist_deg."""
    if shape is None:
        return None
    try:
        import math
        from shapely.geometry import LineString as ShLine, Point as P
        target = P(lon, lat)
        best = None
        best_dist = float("inf")
        for f in roads_features:
            g = f.get("geometry") or {}
            t = g.get("type")
            if t == "LineString":
                lines = [g["coordinates"]]
            elif t == "MultiLineString":
                lines = g["coordinates"]
            else:
                continue
            for coords in lines:
                if len(coords) < 2:
                    continue
                ls = ShLine(coords)
                d = ls.distance(target)
                if d < best_dist:
                    best_dist = d
                    best = ls
        if best is None or best_dist > max_dist_deg:
            return None
        proj_d = best.project(target, normalized=False)
        cos_lat = max(0.1, math.cos(math.radians(lat)))
        delta_deg = 5.0 / 111000.0 / cos_lat  # 5m on each side
        start = max(0.0, proj_d - delta_deg)
        end = min(best.length, proj_d + delta_deg)
        if end <= start:
            return None
        p1 = best.interpolate(start)
        p2 = best.interpolate(end)
        dx_m = (p2.x - p1.x) * 111000.0 * cos_lat
        dy_m = (p2.y - p1.y) * 111000.0
        bearing = math.degrees(math.atan2(dx_m, dy_m))  # compass: 0=N, 90=E
        if bearing < 0:
            bearing += 360.0
        return round(bearing, 1)
    except Exception:
        return None


def polygon_area_m2(geom: dict) -> float | None:
    """Area in m² of a Polygon/MultiPolygon GeoJSON (lon/lat) by projecting
    to Israel TM (EPSG:2039). Returns None if shapely/pyproj unavailable or
    geometry isn't polygonal."""
    if not geom or shape is None or _ITM_TRANSFORMER is None:
        return None
    if geom.get("type") not in ("Polygon", "MultiPolygon"):
        return None
    try:
        from shapely.ops import transform as _shp_transform
        g = shape(geom)
        g_itm = _shp_transform(
            lambda x, y, z=None: _ITM_TRANSFORMER.transform(x, y), g
        )
        return float(g_itm.area)
    except Exception:
        return None


def make_stadium_polygon(lon: float, lat: float, length_m: float, width_m: float,
                          horizontal: bool = False) -> dict:
    """Stadium shape (rectangle with rounded ends) — matches the legend's dark-green oval
    used for 'מעבר להולכי רגל'. If horizontal=True, length is along east-west."""
    import math
    cos_lat = max(0.1, math.cos(math.radians(lat)))
    dlat_per_m = 1.0 / 111000.0
    dlon_per_m = 1.0 / (111000.0 * cos_lat)
    half_l = length_m / 2.0
    r = width_m / 2.0
    # Build along Y (vertical) by default; rotate by swapping if horizontal.
    pts = []
    # Top semicircle (12 points)
    for i in range(13):
        angle = math.pi * (1 - i / 12.0)  # π → 0
        x = r * math.cos(angle)
        y = half_l + r * math.sin(angle)
        pts.append((x, y))
    # Bottom semicircle
    for i in range(13):
        angle = math.pi * (i / 12.0)  # 0 → π
        x = r * math.cos(angle + math.pi)
        y = -half_l + r * math.sin(angle + math.pi)
        pts.append((x, y))
    if horizontal:
        pts = [(y, x) for (x, y) in pts]
    ring = [[lon + dx * dlon_per_m, lat + dy * dlat_per_m] for (dx, dy) in pts]
    ring.append(ring[0])
    return {"type": "Polygon", "coordinates": [ring]}


# Synthetic polygon SHAPE per (domain, service) — built so each marker resembles its
# matching item in the תשריט legend.  Returns (geometry, marker_kind) where marker_kind
# is read by the JS layer to vary fill/stroke (e.g. 'hollow_circle' = no fill, only stroke).
def _make_synthetic_polygon(lon: float, lat: float, domain: str, service_he: str) -> tuple[dict, str]:
    s = (service_he or "").strip()

    # ---- TRANSPORT domain --------------------------------------------------
    if domain == "transport":
        if "מעבר חציה" in s or "חציה" in s:
            # Crosswalk = exact POINT. Frontend draws a small dot.
            return {"type": "Point", "coordinates": [lon, lat]}, "crossing_point"
        if "מעבר ה" in s or "הולכי" in s:
            # Pedestrian passage = a Point fallback when no road segment was found.
            # Frontend treats it like a passage marker; ideal source has LineString.
            return {"type": "Point", "coordinates": [lon, lat]}, "passage_point"
        if "רחוב חדש" in s or "רחוב" in s or "תח" in s or "תחנ" in s:
            # Road segment / bus stop — thin horizontal rectangle
            return make_rect_polygon(lon, lat, 22.0, 5.0), "road_segment"
        return make_circle_polygon(lon, lat, 6.0), "small_circle"

    # ---- PUBLIC SPACE domain ----------------------------------------------
    if domain == "public_space":
        if "מעבר ה" in s or "מעבר הולכי" in s:
            return {"type": "Point", "coordinates": [lon, lat]}, "passage_point"
        if "מעבר חציה" in s or "חציה" in s:
            return {"type": "Point", "coordinates": [lon, lat]}, "crossing_point"
        if "מתחם" in s:
            # Area-level recommendation — bigger red square (matches "מתחם" in legend)
            return make_rect_polygon(lon, lat, 28.0, 24.0), "metaham_square"
        if "פרויקט בשצ" in s or "פרויקט בשצ\"פ" in s or s.startswith("פרויקט"):
            # Hollow green circle (matches "פרויקט בשצ"פ" in legend — outline only)
            return make_circle_polygon(lon, lat, 11.0), "hollow_circle"
        if "שצ" in s or "גינ" in s or "פארק" in s or "שביל" in s:
            # שצ"פ — green filled rectangle (matches legend "שצ"פ")
            return make_rect_polygon(lon, lat, 22.0, 18.0), "shatzap_rect"
        # default public-space → hollow circle
        return make_circle_polygon(lon, lat, 9.0), "hollow_circle"

    # ---- PROGRAMA (שב"צ) domain ------------------------------------------
    if "מתחם" in s:
        return make_rect_polygon(lon, lat, 28.0, 24.0), "metaham_square"
    # שב"צ עתידי — brown filled rectangle (matches legend "שב"צ עתידי")
    return make_rect_polygon(lon, lat, 22.0, 18.0), "shavatz_future_rect"


def union_geometry(features):
    """Union features' polygons via shapely (proper boolean union), avoiding
    Leaflet evenodd holes from naive concat. See `feedback_multipolygon_evenodd.md`."""
    if not features:
        return None
    if shape is not None and unary_union is not None:
        try:
            geoms = []
            for f in features:
                g = f.get("geometry")
                if not g:
                    continue
                geoms.append(shape(g))
            if not geoms:
                return None
            merged = unary_union(geoms)
            if merged.is_empty:
                return None
            return mapping(merged)
        except Exception:
            pass
    # Fallback: naive concat
    polys = []
    for f in features:
        g = f.get("geometry") or {}
        if g.get("type") == "Polygon":
            polys.append(g["coordinates"])
        elif g.get("type") == "MultiPolygon":
            polys.extend(g["coordinates"])
    if not polys:
        return features[0].get("geometry")
    if len(polys) == 1:
        return {"type": "Polygon", "coordinates": polys[0]}
    return {"type": "MultiPolygon", "coordinates": polys}


# --- Main -----------------------------------------------------------------
def load_geojson(path: Path):
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _norm_shp_name(s) -> str:
    """Normalize a project name for matching: collapse punctuation/whitespace."""
    if s is None:
        return ""
    s = str(s)
    # Replace Hebrew quote-like marks + ASCII punct with space
    s = re.sub(r"[\s\-_'\"׳״,()\.]+", " ", s)
    return s.strip()


def build_projector_official_index() -> tuple[dict, list]:
    """Load the 3 official projector shapefiles (already in WGS84 GeoJSON form).

    Returns:
        index: dict (shp_sub, normalized_name) → list of features (geom + props).
               Used to look up xlsx rows by area+name.
        bus_stops: list of layer-2 bus-stop Point features (not present in xlsx).
                   These will be emitted as standalone projector features.
    """
    index: dict[tuple[str, str], list] = defaultdict(list)
    bus_stops: list[dict] = []
    for p in PROJECTOR_OFFICIAL_PATHS:
        if not p.exists():
            print(f"      [shp] missing: {p.name}")
            continue
        try:
            fc = load_geojson(p)
        except Exception as e:
            print(f"      [shp] load error {p.name}: {e}")
            continue
        layer_tag = p.stem.replace("projector_official_", "")
        n_index = 0
        for f in fc.get("features", []):
            prop = f.get("properties") or {}
            sub = prop.get("שכונה")
            name = prop.get("שם_פרוייקט")
            if not sub:
                continue
            f.setdefault("properties", {})["_shp_layer"] = layer_tag
            # Bus-stop points (layer2): mahut='תחנה חדשה' — emit as standalone features.
            if layer_tag == "layer2" and prop.get("מהות") in {"תחנה חדשה"}:
                bus_stops.append(f)
                continue
            if not name:
                continue
            key = (str(sub), _norm_shp_name(name))
            index[key].append(f)
            n_index += 1
        print(f"      [shp] {p.name}: indexed {n_index}, bus_stops={len(bus_stops) if layer_tag=='layer2' else '-'}")
    return dict(index), bus_stops


def find_official_shp_match(area: str, project_name: str, shp_index: dict) -> dict | None:
    """Find an official shapefile feature that matches the xlsx row by area+name.

    Match precedence:
      1) Exact normalized name match within the same sub-neighborhood.
      2) Substring match (xlsx name contains shp name, or vice versa) within same sub.
    Returns the FIRST matched feature (or None). When layer1 + layer3 both have the
    same project, layer3 should win because it has richer attributes (status, חומש,
    parking). We achieve that by inserting layer3 first in PROJECTOR_OFFICIAL_PATHS.
    Actually we want layer3 first so it's in shp_index first. Update PATHS order if needed.
    """
    shp_sub = AREA_TO_SHP_SUB.get(area)
    if not shp_sub or not project_name:
        return None
    norm_xlsx = _norm_shp_name(project_name)
    # Exact match
    direct = shp_index.get((shp_sub, norm_xlsx))
    if direct:
        # Prefer layer3 (richer attrs) if there are multiple hits.
        for f in direct:
            if f.get("properties", {}).get("_shp_layer") == "layer3":
                return f
        return direct[0]
    # Substring match
    best: dict | None = None
    best_score = 0
    for (sub, sn), fs in shp_index.items():
        if sub != shp_sub or not sn:
            continue
        if sn in norm_xlsx or norm_xlsx in sn:
            # Prefer longer overlap (= more specific match)
            score = min(len(sn), len(norm_xlsx))
            if score > best_score:
                best_score = score
                best = fs[0]
                # Within the candidate group, prefer layer3 too
                for f in fs:
                    if f.get("properties", {}).get("_shp_layer") == "layer3":
                        best = f
                        break
    return best


def main():
    print(f"[1/6] Reading xlsx: {XLSX_PATH.name}")
    if not XLSX_PATH.exists():
        sys.exit(f"FATAL: xlsx not found at {XLSX_PATH}")
    # OneDrive may keep the file locked while Excel is open.
    # Fall back to a copy in temp if a direct open fails with PermissionError.
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    except PermissionError:
        import shutil
        tmp = Path(r"C:\temp\oranim_extract") / "projector_xlsx_copy.xlsx"
        tmp.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(XLSX_PATH, tmp)
        print(f"      (xlsx locked - reading copy at {tmp})")
        wb = openpyxl.load_workbook(tmp, data_only=True)
    ws = wb[SHEET]

    # Read header from row 2 (row 1 is meta — "רשימה"/"הזנה חופשית" labels)
    header = [c.value for c in ws[2]]
    h = {name: idx for idx, name in enumerate(header)}
    print(f"      xlsx headers ({len(header)} cols): {[str(c) for c in header]}")

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(c is None for c in row):
            continue
        if row[h["אזור"]] not in GONENIM_AREAS:
            continue
        rows.append(row)
    from collections import Counter as _Ca
    by_area = _Ca(r[h["אזור"]] for r in rows)
    print(f"      total rows in minahak גוננים: {len(rows)}")
    for a, c in by_area.most_common():
        print(f"        {a}: {c}")

    print("[2/6] Loading reference geojsons …")
    landuse = load_geojson(LANDUSE_PATH)
    plans = load_geojson(PLANS_PATH)
    future_shavaz = load_geojson(FUTURE_SHAVAZ_PATH)
    minahak = load_geojson(MINAHAK_PATH)
    sub_neigh = load_geojson(SUB_NEIGH_PATH)
    migrash_panui = load_geojson(MIGRASH_PANUI_PATH) if MIGRASH_PANUI_PATH.exists() else {"features": []}
    buildings = load_geojson(BUILDINGS_PATH) if BUILDINGS_PATH.exists() else {"features": []}
    roads = load_geojson(ROADS_PATH) if ROADS_PATH.exists() else {"features": []}

    # Build the minahak-union polygon (union of the 4 sub-neighborhoods).
    # Used to filter buildings/roads so street-centroid lookups don't pick up the same
    # street name from a different city area.
    minahak_union_geom = None
    if MINAHAK_UNION_PATH.exists() and shape is not None:
        try:
            mu = load_geojson(MINAHAK_UNION_PATH)
            if mu.get("features"):
                minahak_union_geom = shape(mu["features"][0]["geometry"])
        except Exception:
            minahak_union_geom = None
    if minahak_union_geom is not None:
        print(f"      minahak union bounds: {minahak_union_geom.bounds}")
    # Build area → centroid lookup
    sub_centroid_by_area: dict[str, list] = {}
    for f in sub_neigh.get("features", []):
        name = f["properties"].get("schn_nama")
        if not name:
            continue
        c = feature_centroid(f)
        if c:
            for area, sn_name in AREA_TO_SUBNEIGH.items():
                if sn_name == name:
                    sub_centroid_by_area[area] = c
                    break
    print(f"      sub-neighborhood centroids: {sorted(sub_centroid_by_area.keys())}")

    # Index landuse by pl_number
    landuse_by_plnum: dict[str, list] = defaultdict(list)
    for f in landuse["features"]:
        pn = f["properties"].get("pl_number")
        if pn:
            landuse_by_plnum[pn].append(f)
    # Index future_shavaz by (helka, gush). To avoid matching the same MIGRASH value across
    # different sub-neighborhoods (e.g. a פת feature matching a קטמונים lot), we ALSO key by
    # sub-neighborhood derived from each lot's geometric containment.
    fshavaz_by_migrash: dict[str, list] = defaultdict(list)
    fshavaz_by_migrash_in_sub: dict[tuple, list] = defaultdict(list)
    # Pre-load sub-neighborhood polygons keyed by xlsx area name.
    sub_polys: dict[str, object] = {}
    if shape is not None and sub_neigh.get("features"):
        for sf in sub_neigh["features"]:
            schn = sf["properties"].get("schn_nama")
            for area_name, target_schn in AREA_TO_SUBNEIGH.items():
                if schn == target_schn:
                    sub_polys[area_name] = shape(sf["geometry"])
    for f in future_shavaz["features"]:
        mig = f["properties"].get("MIGRASH")
        if not mig:
            continue
        if minahak_union_geom is not None and shape is not None:
            try:
                geom = shape(f["geometry"])
                if not minahak_union_geom.intersects(geom):
                    continue
                # Index per sub-neighborhood that the lot intersects
                for area_name, poly in sub_polys.items():
                    try:
                        if poly.intersects(geom):
                            fshavaz_by_migrash_in_sub[(area_name, str(mig))].append(f)
                    except Exception:
                        pass
            except Exception:
                pass
        fshavaz_by_migrash[str(mig)].append(f)
    # Index plans by plan_name (for overlap)
    plans_by_name: dict[str, dict] = {}
    for f in plans["features"]:
        pn = f["properties"].get("plan_name")
        if pn:
            plans_by_name[pn] = f

    # Index migrash_panui by (gush, parcel) — for direct lot lookup using xlsx 'מזהה מיקום' / PDF excerpt 'גוש/חלקה'.
    migpanui_by_gush_parcel: dict[tuple, dict] = {}
    for f in migrash_panui.get("features", []):
        prop = f["properties"]
        gush = prop.get("planning_block")
        parcel = prop.get("planning_parcel")
        if gush and parcel:
            migpanui_by_gush_parcel[(str(gush), str(parcel))] = f

    # NEW: index full Jerusalem-parcels layer (10K+ חלקות) by (gush, helka).
    # This is the authoritative source for resolving 'מזהה מיקום' = "<helkas>/<gush>"
    # to actual lot polygons. Used to position pedestrian passages across multiple lots.
    parcels_by_gush_helka: dict[tuple, list] = defaultdict(list)
    if PARCELS_PATH.exists():
        try:
            parcels = load_geojson(PARCELS_PATH)
            for f in parcels.get("features", []):
                prop = f["properties"]
                gush = str(prop.get("GUSH_NO") or "").strip()
                helka = str(prop.get("HELKA_SHOW") or "").strip()
                if gush and helka:
                    parcels_by_gush_helka[(gush, helka)].append(f)
            print(f"      parcels indexed: {len(parcels.get('features', []))} (in {len(parcels_by_gush_helka)} (gush,helka) keys)")
        except Exception as e:
            print(f"      parcels load failed: {e}")

    # Load yiud_karka_kayam — existing-land-use polygons from the city DB.
    # Used to detect inner שצ"פ polygons inside metaham compounds (so the
    # compound polygon can be rendered as just an outline + the שצ"פים inside
    # rendered prominently as the actual project subjects).
    yiud_kayam_shatzap_features: list = []  # list of {geom_shape, area_m2, descr}
    if YIUD_KAYAM_PATH.exists() and shape:
        try:
            yk_data = load_geojson(YIUD_KAYAM_PATH)
            # שצ"פ-like Descr values (filter for open-space lots only).
            _SHATZAP_DESCR_KEYWORDS = ("שטח צבורי פתוח", "שטח ציבורי פתוח", "שטח פרטי פתוח",
                                      "פרטי פתוח", "שטחים פתוחים", "שפ\"פ", "פס ירק",
                                      "שטח ירק", "שטח עיצוב נופי")
            for f in yk_data.get("features", []):
                d = str(f["properties"].get("Descr") or "")
                if not any(kw in d for kw in _SHATZAP_DESCR_KEYWORDS):
                    continue
                try:
                    g = shape(f["geometry"])
                    if not g.is_valid:
                        g = g.buffer(0)
                    if g.is_empty:
                        continue
                    yiud_kayam_shatzap_features.append({
                        "geom": g,
                        "descr": d,
                        "raw": f["geometry"],
                    })
                except Exception:
                    continue
            print(f"      yiud_karka שצ\"פ indexed: {len(yiud_kayam_shatzap_features)} polygons")
        except Exception as e:
            print(f"      yiud_karka load failed: {e}")

    # Index buildings by (street_normalized, house_num) — for address-based lookup from project_name.
    def _norm_street(s: str) -> str:
        if not s:
            return ""
        # Strip common suffixes/prefixes; collapse whitespace
        s = re.sub(r"רח['׳]\s*", "", str(s))
        s = re.sub(r"\s+", " ", s).strip()
        return s

    # Helper: only keep features whose centroid/representative point is inside the minahak union.
    # This prevents "טשרניחובסקי" from being matched to its segment in the wrong neighborhood.
    def _inside_minahak(feat) -> bool:
        if minahak_union_geom is None or ShPoint is None:
            return True  # no filter available — accept all
        g = feat.get("geometry")
        if not g:
            return False
        if g["type"] == "Point":
            return minahak_union_geom.contains(ShPoint(g["coordinates"][0], g["coordinates"][1]))
        # For polygon/line: take a representative point
        try:
            return minahak_union_geom.intersects(shape(g))
        except Exception:
            return False

    buildings_by_addr: dict[tuple, list] = defaultdict(list)
    buildings_by_street: dict[str, list] = defaultdict(list)
    n_inside = 0
    for f in buildings.get("features", []):
        prop = f["properties"]
        st = _norm_street(prop.get("street") or "")
        hn = prop.get("house_num")
        if not (st and hn):
            continue
        if not _inside_minahak(f):
            continue
        n_inside += 1
        buildings_by_addr[(st, str(hn))].append(f)
        buildings_by_street[st].append((str(hn), f))
    print(f"      buildings inside minahak: {n_inside}")

    # Roads index: street_name → list of road features (LineStrings) — minahak-clipped.
    roads_by_street: dict[str, list] = defaultdict(list)
    n_roads_inside = 0
    for f in roads.get("features", []):
        st = _norm_street(f["properties"].get("street") or "")
        if not st:
            continue
        if not _inside_minahak(f):
            continue
        n_roads_inside += 1
        roads_by_street[st].append(f)
    print(f"      roads inside minahak: {n_roads_inside}")

    # Official projector shapefiles (transport layer): name+area indexed for direct lookup.
    print("      [shp] loading official projector shapefiles …")
    shp_official_index, shp_bus_stops = build_projector_official_index()
    shp_used_keys: set = set()  # track which shp features got attached to an xlsx row
    print(f"      [shp] index keys: {len(shp_official_index)}, standalone bus stops: {len(shp_bus_stops)}")

    # Master index of street names that appear in our data (buildings ∪ roads).
    # Many streets have a prefix ("דוד שמעוני", "הרב הרצוג", "יוסי בן יועזר") — the projector text
    # often says only the last word ("שמעוני"). Add suffix-based aliases so those match.
    all_street_names = set(buildings_by_street.keys()) | set(roads_by_street.keys())
    # alias_to_canonicals: alias_str → list of canonical street names that contain it as suffix.
    alias_to_canonicals: dict[str, list[str]] = defaultdict(list)
    for canonical in all_street_names:
        if len(canonical) < 4:
            continue
        if not any('א' <= c <= 'ת' for c in canonical):
            continue
        # Always include the canonical name itself
        alias_to_canonicals[canonical].append(canonical)
        # Add last word as alias if it's >=4 chars (avoid 'בן', 'דרך', 'הרב')
        words = canonical.split()
        if len(words) > 1:
            for k in range(1, min(3, len(words))):
                suffix = " ".join(words[-k:])
                if len(suffix) >= 4 and not any(suffix.startswith(p) for p in ("הרב ", "רבי ", "דרך ", "מעלה ")):
                    alias_to_canonicals[suffix].append(canonical)
    # Sort alias keys: longest first (more specific match wins)
    searchable_streets = sorted(alias_to_canonicals.keys(), key=lambda s: -len(s))
    print(f"      streets indexed: {len(all_street_names)} unique names, {len(searchable_streets)} aliases")

    minahak_centroid = feature_centroid(minahak["features"][0]) if minahak["features"] else None
    if not minahak_centroid:
        sys.exit("FATAL: cannot compute minahak_gonen centroid")
    print(f"      minahak centroid: {minahak_centroid}")

    # Build the structured public-space booklets lookup: (area, rec_num) → fields.
    public_space_lookup = _build_public_space_lookup()
    print(f"      public-space booklet recommendations: {len(public_space_lookup)} (across 3 PDFs)")

    # Build the karteset cards lookup keyed by (pdf_filename, page_num).
    karteset_lookup = _build_karteset_lookup()
    print(f"      karteset cards: {len(karteset_lookup)} pages (across 2 PDFs)")

    # Build the per-project narrative index from the "תיק שכונה" / "דוח מסכם" decks.
    # Keyed by (area, normalized_name) → slide info { pdf_name, page, title, label, text }.
    tik_index = _build_tik_shchunah_index()
    print(f"      tik-shchunah: {len(tik_index)} project-slide keys (across {len(TIK_SHCHUNAH_PDFS)} PDFs)")

    # Build the tashrir-marker index (gonenim only — the PDF covers only that area).
    tashrir_markers = _build_tashrir_marker_index(buildings.get("features", []), minahak["features"][0]["geometry"])
    if tashrir_markers:
        n_total = sum(len(v) for v in tashrir_markers.values())
        print(f"      tashrir markers: {len(tashrir_markers)} unique numbers, {n_total} total markers")

    # Build the TYPED tashrit index: (area, type, num) → (lon, lat) for crossings + passages.
    # This is the authoritative position source for these transport features.
    # Use the FULL minahak-gonenim union (covers all 4 sub-neighborhoods: רסקו,
    # גוננים א-ו, קטמונים ח-ט, פת). The legacy minahak_gonen.geojson is a strict
    # subset that excludes רסקו, which would silently mis-locate any chip outside it.
    _minahak_for_tashrir = minahak["features"][0]["geometry"]
    try:
        _mu_full = load_geojson(MINAHAK_UNION_PATH) if MINAHAK_UNION_PATH.exists() else None
        if _mu_full and _mu_full.get("features"):
            _minahak_for_tashrir = _mu_full["features"][0]["geometry"]
    except Exception:
        pass
    tashrir_typed = _build_tashrir_typed_markers(
        buildings.get("features", []),
        _minahak_for_tashrir,
        roads.get("features", []),
    )
    if tashrir_typed:
        from collections import Counter as _Cnt
        by_at = _Cnt((k[0], k[1]) for k in tashrir_typed.keys())
        print(f"      tashrit typed markers: {len(tashrir_typed)} keys")
        for (a, t), n in by_at.most_common():
            print(f"        {a:14s} {t}: {n}")

    # NOTE: We used to extract the green-line drawings from the PDF and use them
    # as passage geometry, then snap each vertex to the nearest road. That was
    # consistently off (the PDF→WGS84 affine has 5-15m drift, varies by region)
    # so it's no longer wired in. The tashrit *labels* (point markers) are still
    # used as a position anchor when xlsx has no parcels match — see the
    # geocoding block for the lookup against `tashrir_typed[(area, type, num)]`.

    print("[3/6] Building features …")
    geocode_cache = _load_geocode_cache()
    print(f"      geocode cache: {sum(1 for v in geocode_cache.values() if isinstance(v, dict) and v.get('lat'))} entries")
    out_features = []
    seen_pids: set[str] = set()
    overlap_idx: dict[str, dict] = {}
    summary: dict[str, dict] = defaultdict(lambda: {"by_key": defaultdict(lambda: {"units": 0.0, "overlap_units": 0.0, "projects": []}), "extras": defaultdict(lambda: {"units": 0.0, "overlap_units": 0.0, "projects": []})})
    match_stats = defaultdict(lambda: defaultdict(int))
    duplicate_count = 0

    for row in rows:
        area = row[h["אזור"]]
        proj_num = row[h["מס פרויקט"]]
        proj_name = row[h["שם פרויקט"]]
        # Apply name override BEFORE any name-based lookups (overrides, notes, etc.)
        _orig_name = str(proj_name or "").strip()
        _name_override = MANUAL_NAME_OVERRIDES.get((_orig_name, area, str(proj_num or "").strip()))
        if _name_override:
            proj_name = _name_override
        domain_he = row[h["תחום"]]
        service_he = (row[h["שירות"]] or "").strip()
        units_label = row[h["יחידות"]]
        units_required = row[h["יחידות נדרשות לביצוע"]]
        description = row[h["תיאור הפרויקט"]]
        plan_raw = row[h["מספר תכנית (אם קיים)"]]
        year_planning = coerce_int(row[h["שנת תכנון נדרשת"]])
        year_completion = row[h["צפי לשנת השלמת ביצוע"]]
        chumash_text = row[h["חומש לביצוע"]]
        managing = row[h["גורם מנהל"]]
        obstacles = row[h["חסמים/פעולות נדרשות"]] if "חסמים/פעולות נדרשות" in h else None
        related_projects = row[h["פרויקטים נלווים"]] if "פרויקטים נלווים" in h else None
        location_id = row[h["מזהה מיקום"]]
        pdf_link = row[h["קישורים"]]
        _pdf_override = MANUAL_PDF_OVERRIDES.get((str(proj_name or "").strip(), area))
        if _pdf_override:
            pdf_link = _pdf_override
        plan_status = row[h["סטטוס"]]
        estimate = coerce_float(row[h["אומדן"]])

        domain = DOMAIN_MAP.get(domain_he, "other")
        service_key = SERVICE_MAP.get(service_he)
        chumash = normalize_chumash(chumash_text, year_completion)
        plan_number = extract_plan_number(plan_raw)
        pid = project_id(area, proj_num, service_he, str(proj_name or "").strip())
        if pid in seen_pids:
            duplicate_count += 1
            # Same (area, num, svc, name) → merge any distinct obstacles into the
            # already-emitted feature so we don't silently drop info.
            obs_new = str(obstacles).strip() if obstacles else ""
            if obs_new:
                existing_feat = next((f for f in out_features
                                      if f["properties"].get("project_id") == pid), None)
                if existing_feat:
                    obs_old = existing_feat["properties"].get("obstacles") or ""
                    parts = [p.strip() for p in obs_old.split(";") if p.strip()] if obs_old else []
                    if obs_new not in parts:
                        parts.append(obs_new)
                        existing_feat["properties"]["obstacles"] = "; ".join(parts)
            continue
        seen_pids.add(pid)

        # Skip features explicitly marked for removal (no map presence at all).
        _sk_name = str(proj_name or "").strip()
        _sk_svc = str(service_he or "").strip()
        if ((_sk_name, area) in SKIP_FEATURES
            or (_sk_name, area, _sk_svc) in SKIP_FEATURES):
            continue

        # Compute PDF excerpt early so geocoding can use it for street/gush lookup.
        pdf_excerpt = extract_pdf_excerpt(pdf_link)

        # For public_space recommendations, see if we have a structured page in the booklet.
        # ONLY match when project_num is a plain integer (e.g. "1", "2"). Strings like
        # "מעבר ה\"ר מס' 6" use a SEPARATE numbering scheme for the crossings table —
        # they must NOT match the main numbered recommendations.
        public_space_rec = None
        if domain == "public_space" and proj_num is not None:
            num_str = str(proj_num).strip()
            if re.fullmatch(r"\d{1,3}", num_str):
                public_space_rec = public_space_lookup.get((area, num_str))
        if public_space_rec:
            structured_excerpt = _format_structured_excerpt(public_space_rec)
            if structured_excerpt:
                pdf_excerpt = structured_excerpt

        # For programa features: try the כרטסת card matching the pdf_link page reference.
        # pdf_link looks like "עמ' 18 כרטסת שטחים חומים רסקו, גוננים א-ו, פת" or
        # "עמ' 5 קטמונים כרטסת שטחים חומים".
        karteset_rec = None
        karteset_page = None
        if domain == "programa" and pdf_link and "כרטסת" in str(pdf_link):
            m_pg = re.search(r"עמ['׳]?\s*(\d+)", str(pdf_link))
            if m_pg:
                page_num = int(m_pg.group(1))
                # Determine which karteset PDF this references
                target_pdf = None
                if "קטמונים" in str(pdf_link):
                    target_pdf = "קטמון ח-ט - כרטסת שטחים חומים 06.062024.pdf"
                else:
                    target_pdf = "כרטסת שטחים חומים מרחב רסקו-גוננים א-ו ופת 01052024.pdf"
                rec = karteset_lookup.get((target_pdf, page_num))
                if rec:
                    karteset_rec = rec
                    karteset_page = page_num
                    pdf_excerpt = _format_karteset_excerpt(rec, page_num)

        # --- Geocode --------------------------------------------------------
        geometry = None
        match_quality = None
        geometry_source = "minahak_centroid"
        matched_features = []
        # Carry shapefile attributes (filled in below if a match is found).
        shp_enrich: dict = {}

        # HARD OVERRIDE: features whose xlsx location_id points to a wrong parcel
        # OR whose passage geometry can't be derived from the chip+bearing logic.
        # Override values are either:
        #   - [lon, lat] list/tuple → Point
        #   - {"type": "LineString" | "Polygon" | "MultiPolygon" | ..., "coordinates": [...]}
        #     → use as-is
        #   - {"plan_name": "101-NNNNNNN"} → look up the polygon from plans.geojson at
        #     build time (cleaner than inlining large polygons)
        # Lookup precedence: (name, area, service, project_num) → (name, area, service) → (name, area) → None.
        # The 4-tuple form is for cases where multiple xlsx rows share name+area+service
        # but differ by project_num (e.g. "מעבר ה"ר מגן חביבה" appears for 3 different
        # מס' values, each a different physical passage).
        # The 3-tuple form is for cases where multiple xlsx rows share name+area but
        # differ by service (e.g. "מתחם לוריא" has 6 services, each needing its own
        # location).
        _name = str(proj_name or "").strip()
        _service = str(service_he or "").strip()
        _pn = str(proj_num or "").strip()
        override = (MANUAL_LOCATION_OVERRIDES.get((_name, area, _service, _pn))
                    or MANUAL_LOCATION_OVERRIDES.get((_name, area, _service))
                    or MANUAL_LOCATION_OVERRIDES.get((_name, area)))
        if override:
            if isinstance(override, dict) and override.get("plan_name"):
                pn_ref = override["plan_name"]
                ref_feat = next((f for f in plans.get("features", [])
                                 if f.get("properties", {}).get("plan_name") == pn_ref), None)
                if ref_feat and ref_feat.get("geometry"):
                    geometry = ref_feat["geometry"]
                    match_quality = "manual_override"
                    geometry_source = "manual_override"
                else:
                    print(f"      [override] plan_name={pn_ref!r} not found in plans.geojson — skipping override for {proj_name!r}")
            elif isinstance(override, dict) and override.get("road_name"):
                # Collect all road segments whose `street` matches and union them
                # via shapely. Divided roads have parallel ft/tf pairs; drop `tf` to
                # avoid drawing both lanes. Then linemerge to chain adjacent segments.
                # Optional `cut_at: [lon, lat]` truncates the road at the projection
                # of that point — keeps the LONGER portion by default.
                rn = override["road_name"]
                cut_at = override.get("cut_at")
                from shapely.geometry import LineString as _ShL, MultiLineString as _ShML, Point as _ShP
                from shapely.ops import linemerge as _linemerge, substring as _substring
                ls = []
                for rf in roads.get("features", []):
                    rp = rf.get("properties") or {}
                    if rp.get("street") != rn:
                        continue
                    if rp.get("oneway") == "tf":
                        continue  # parallel lane of a divided road — paired with an 'ft'
                    rg = rf.get("geometry") or {}
                    if rg.get("type") == "LineString" and len(rg["coordinates"]) >= 2:
                        ls.append(_ShL(rg["coordinates"]))
                    elif rg.get("type") == "MultiLineString":
                        for ln in rg["coordinates"]:
                            if len(ln) >= 2:
                                ls.append(_ShL(ln))
                if ls:
                    from shapely.ops import unary_union as _uu
                    # Step 1: merge adjacent endpoints into long chunks.
                    merged_obj = _linemerge(_ShML(ls))
                    chunks = list(merged_obj.geoms) if merged_obj.geom_type == "MultiLineString" else [merged_obj]
                    # Step 2: drop parallel-lane duplicates by buffer-overlap ratio.
                    # If ≥70% of a candidate sits within ~20m of a previously-kept
                    # chunk, treat as a parallel duplicate. (Some divided streets
                    # have lanes 15-20m apart, so we need a generous buffer.)
                    BUFFER_DEG = 20.0 / 111000.0   # ~20m at this latitude
                    chunks.sort(key=lambda l: -l.length)
                    kept = []
                    for cand in chunks:
                        is_dup = False
                        for kp in kept:
                            try:
                                buf = kp.buffer(BUFFER_DEG)
                                inter = cand.intersection(buf)
                                inside_len = inter.length if hasattr(inter, "length") else 0
                                if cand.length > 0 and inside_len / cand.length > 0.7:
                                    is_dup = True
                                    break
                            except Exception:
                                pass
                        if not is_dup:
                            kept.append(cand)
                    # Optional cut: keep only the longer portion around cut_at point
                    if cut_at and kept:
                        cut_pt = _ShP(cut_at[0], cut_at[1])
                        cut_kept = []
                        for line in kept:
                            try:
                                proj_d = line.project(cut_pt, normalized=False)
                                # Two sides of the cut
                                lo = _substring(line, 0, proj_d)
                                hi = _substring(line, proj_d, line.length)
                                # Keep whichever side is longer
                                cut_kept.append(lo if lo.length >= hi.length else hi)
                            except Exception:
                                cut_kept.append(line)
                        kept = [c for c in cut_kept if not c.is_empty]
                    # Optional `extra_lines`: append manually-defined LineStrings to
                    # the road_name pull (e.g. missing connecting segments).
                    for extra in (override.get("extra_lines") or []):
                        if extra and len(extra) >= 2:
                            try:
                                kept.append(_ShL(extra))
                            except Exception:
                                pass
                    if not kept:
                        geometry = None
                    elif len(kept) == 1:
                        geometry = mapping(kept[0])
                    else:
                        geometry = mapping(_ShML(kept))
                    match_quality = "manual_override"
                    geometry_source = "manual_override_road_name"
                else:
                    print(f"      [override] road_name={rn!r} not found in roads.geojson — skipping override for {proj_name!r}")
            elif isinstance(override, dict) and override.get("gush_helka"):
                # "{gush}/{helka}" → single parcel polygon
                # "{gush}/{helka1},{helka2},..." → union of multiple parcels (MultiPolygon)
                gh = str(override["gush_helka"])
                if "/" in gh:
                    g_, hs = gh.split("/", 1)
                    g_ = g_.strip().lstrip("0")
                    helka_list = [h.strip().lstrip("0") for h in hs.split(",") if h.strip()]
                    geoms = []
                    missing = []
                    for h_ in helka_list:
                        matched = parcels_by_gush_helka.get((g_, h_))
                        if matched and matched[0].get("geometry"):
                            try:
                                geoms.append(shape(matched[0]["geometry"]))
                            except Exception:
                                pass
                        else:
                            missing.append(h_)
                    if missing:
                        print(f"      [override] gush_helka={gh!r}: missing helkas {missing} — skipping")
                    elif geoms:
                        from shapely.ops import unary_union as _uu
                        union = _uu(geoms) if len(geoms) > 1 else geoms[0]
                        geometry = mapping(union)
                        match_quality = "manual_override"
                        geometry_source = "manual_override"
                    else:
                        print(f"      [override] gush_helka={gh!r}: no parcels found — skipping")
                else:
                    print(f"      [override] gush_helka={gh!r} must be 'gush/helka' format — skipping")
            elif isinstance(override, dict) and override.get("area_polygon"):
                # Use the whole sub-neighborhood polygon for area-wide recommendations
                # (e.g. תו חניה, sub-neighborhood-level concepts with no specific lot).
                sn_name = AREA_TO_SUBNEIGH.get(area)
                sn_feat = next((f for f in sub_neigh.get("features", [])
                                if f.get("properties", {}).get("schn_nama") == sn_name), None)
                if sn_feat and sn_feat.get("geometry"):
                    geometry = sn_feat["geometry"]
                    match_quality = "manual_override"
                    geometry_source = "manual_override_area_polygon"
                    pass
                else:
                    print(f"      [override] area_polygon: sub_neighborhood {sn_name!r} not found for area {area!r}")
            elif isinstance(override, dict) and override.get("type"):
                geometry = override
                match_quality = "manual_override"
                geometry_source = "manual_override"
            else:
                geometry = {"type": "Point", "coordinates": list(override)}
                match_quality = "manual_override"
                geometry_source = "manual_override"

        # Step 0 (PRIORITY): official projector shapefile match by area+name.
        # The shapefile is authoritative for transport-domain features (roads, passages,
        # bus stops). It also covers many pedestrian-passage / public-space items, so
        # we try matching for transport AND public_space — but NOT for programa (would
        # be too risky for false positives on building names).
        # SKIP this step if a MANUAL_LOCATION_OVERRIDES already set the geometry —
        # the override is authoritative and the shp match would clobber it (a row like
        # "מעבר חציה מס' 3" name='הנוטרים' would otherwise match the entire הנוטרים
        # street polyline by name, hiding the user-provided crosswalk point).
        if not geometry and domain in ("transport", "public_space") and shp_official_index:
            shp_feat = find_official_shp_match(area, proj_name, shp_official_index)
            if shp_feat and shp_feat.get("geometry"):
                geometry = shp_feat["geometry"]
                match_quality = "shapefile"
                geometry_source = f"projector_official_{shp_feat['properties'].get('_shp_layer','')}"
                sp = shp_feat["properties"]
                # Mark this shp feature as used so we don't also emit it as a standalone.
                shp_used_keys.add((sp.get("שכונה"), _norm_shp_name(sp.get("שם_פרוייקט"))))
                # Collect the official-source enrichment fields (only non-empty/non-default).
                for k_src, k_dst in [
                    ("היררכיה", "shp_hierarchy"),
                    ("סטריות", "shp_oneway"),
                    ("מיתון", "shp_calming"),
                    ("מס_נתיבים", "shp_lanes"),
                    ("רוחב_זכות_דרך", "shp_row_m"),
                    ("סוג_שביל", "shp_path_kind"),
                    ("אלמנט", "shp_element"),
                    ("מהות", "shp_essence"),
                    ("חומש", "shp_chumash_text"),
                    ("שנת_תכנון", "shp_year_planning"),
                    ("שנת_ביצוע", "shp_year_execution"),
                    ("סטטוס", "shp_status"),
                    ("גורם_מנהל", "shp_managing"),
                    ("Park_north", "shp_park_n"),
                    ("Park_south", "shp_park_s"),
                    ("Park_east", "shp_park_e"),
                    ("Park_west", "shp_park_w"),
                ]:
                    v = sp.get(k_src)
                    if v not in (None, "", "--"):
                        shp_enrich[k_dst] = v

        if not geometry and plan_number and plan_number in landuse_by_plnum:
            all_lots = landuse_by_plnum[plan_number]
            # Filter to lots whose mavat_code matches this domain's relevant codes.
            # Avoids "polygon covers whole plan" when the rec is for a specific service.
            allowed_codes = DOMAIN_MAVAT_CODES.get(domain, set())
            relevant = [f for f in all_lots if f["properties"].get("mavat_code") in allowed_codes] if allowed_codes else []
            if relevant:
                matched_features = relevant
                geometry = union_geometry(relevant)
                match_quality = "plan_number"
                geometry_source = "landuse_xplan_filtered"
            else:
                # No domain-relevant plot found inside the plan → use plan polygon
                # (rather than centroid; user can see the plan area).
                matched_features = all_lots
                geometry = union_geometry(all_lots)
                match_quality = "plan_number"
                geometry_source = "landuse_xplan_full"
        elif not geometry and plan_number and plan_number in plans_by_name:
            matched_features = [plans_by_name[plan_number]]
            geometry = plans_by_name[plan_number].get("geometry")
            match_quality = "plan_number"
            geometry_source = "plans"
        elif not geometry:
            # Step 1: try גוש/חלקה pairs from xlsx 'מזהה מיקום' against the full Jerusalem
            # parcels layer (10K+ חלקות). For passages spanning multiple parcels, union them all.
            mig_pairs = extract_migrash_pairs(location_id)
            if mig_pairs:
                hits_parcels = []
                helkas_resolved = []
                for helka_str, gush in mig_pairs:
                    helka_singles: list[str] = []
                    for part in re.split(r"[,\s]+", helka_str):
                        if "-" in part:
                            try:
                                a, b = part.split("-")
                                a_i, b_i = int(a), int(b)
                                if a_i <= b_i and (b_i - a_i) <= 40:
                                    helka_singles.extend(str(x) for x in range(a_i, b_i + 1))
                            except ValueError:
                                pass
                        elif part.isdigit():
                            helka_singles.append(part)
                    for hl in helka_singles:
                        cands = parcels_by_gush_helka.get((str(gush), hl), [])
                        for c in cands:
                            hits_parcels.append(c)
                            helkas_resolved.append(f"{gush}/{hl}")
                if hits_parcels:
                    is_crossing = "מעבר חציה" in (service_he or "") or "חציה" in (service_he or "")
                    if is_crossing:
                        # Crossings are POINT-LIKE — don't union all parcels (would span the
                        # whole road). Use the centroid of the union as the anchor, then SNAP
                        # the point to the nearest road so it sits on the street, not inside a
                        # building.
                        union_geom = union_geometry(hits_parcels)
                        if union_geom:
                            c = feature_centroid({"geometry": union_geom})
                            if c:
                                snapped = snap_point_to_nearest_road(c[0], c[1], roads.get("features", []))
                                final = list(snapped) if snapped else c
                                geometry = {"type": "Point", "coordinates": final}
                                match_quality = "parcels"
                                geometry_source = "parcels_centroid_crossing" + ("_snapped" if snapped else "")
                                matched_features = hits_parcels
                    else:
                        matched_features = hits_parcels
                        geometry = union_geometry(hits_parcels)
                        match_quality = "parcels"
                        geometry_source = "parcels_gush_helka"
                else:
                    # Step 1b: try migrash_panui (smaller, vacant-only)
                    hits_panui = []
                    for helka_str, gush in mig_pairs:
                        for part in re.split(r"[,\s\-]+", helka_str):
                            if part.isdigit():
                                feat = migpanui_by_gush_parcel.get((str(gush), part))
                                if feat:
                                    hits_panui.append(feat)
                    if hits_panui:
                        matched_features = hits_panui
                        geometry = union_geometry(hits_panui)
                        match_quality = "migrash"
                        geometry_source = "migrash_panui"
                    else:
                        # Step 1c: future_shavaz MIGRASH (legacy, area-restricted)
                        for helka, gush in mig_pairs:
                            cand = (
                                fshavaz_by_migrash_in_sub.get((area, helka))
                                or fshavaz_by_migrash_in_sub.get((area, f"{helka}/{gush}"))
                                or fshavaz_by_migrash_in_sub.get((area, f"{helka}-{gush}"))
                            )
                            if cand:
                                matched_features = cand
                                geometry = union_geometry(cand)
                                match_quality = "migrash"
                                geometry_source = "future_shavaz"
                                break

            # Step 2: if still no geometry, try address parsing from project_name (street + number).
            if not geometry:
                addr_geom = _try_address_lookup(proj_name, buildings_by_addr, buildings_by_street)
                if addr_geom:
                    geometry = addr_geom
                    match_quality = "address"
                    geometry_source = "buildings_address"

            # Step 3: parse the PDF excerpt for רחוב/גוש/חלקה (the כרטסת always lists these).
            if not geometry and pdf_excerpt:
                pdf_loc = _parse_pdf_for_location(pdf_excerpt)
                # Step 3a: gush+parcel → migrash_panui exact lot
                if pdf_loc["gush"] and pdf_loc["parcels"]:
                    for parcel in pdf_loc["parcels"]:
                        feat = migpanui_by_gush_parcel.get((pdf_loc["gush"], parcel))
                        if feat:
                            geometry = feat.get("geometry")
                            match_quality = "migrash"
                            geometry_source = "migrash_panui_pdf"
                            break
                # Step 3b: street + house_num → buildings exact / nearest
                if not geometry and pdf_loc["street"]:
                    pseudo = pdf_loc["street"] + (" " + pdf_loc["house_num"] if pdf_loc["house_num"] else "")
                    addr_geom = _try_address_lookup(pseudo, buildings_by_addr, buildings_by_street)
                    if addr_geom:
                        geometry = addr_geom
                        match_quality = "address"
                        geometry_source = "buildings_pdf_address"
                # Step 3c: street centroid (when no house number) — average of all buildings on that street
                if not geometry and pdf_loc["street"]:
                    s_norm = pdf_loc["street"].strip()
                    for st_variant in [s_norm, s_norm.replace("'", "׳"), s_norm.replace("׳", "'")]:
                        houses = buildings_by_street.get(st_variant)
                        if houses:
                            c = _street_centroid(houses)
                            if c:
                                geometry = {"type": "Point", "coordinates": c}
                                match_quality = "street"
                                geometry_source = "buildings_street_centroid"
                                break

            # Step 3.5: cache + online geocoding (Nominatim) for landmarks not in our local data.
            if not geometry:
                geo = _geocode_via_cache_or_nominatim(proj_name, area, geocode_cache)
                if geo:
                    geometry = geo
                    match_quality = "geocoded"
                    geometry_source = "nominatim_cache"

            # Step 4: scan project_name + description + structured PDF location for any known
            # street name (longest match wins). Helps "מעבר ה"ר מגן חביבה", "פת יעקב 3" etc.
            if not geometry:
                ps_loc = (public_space_rec or {}).get("location", "") if public_space_rec else ""
                hay = " ".join([str(proj_name or ""), str(description or ""), str(ps_loc)])
                alias = _find_street_in_text(hay, searchable_streets)
                if alias:
                    # alias may map to multiple canonical streets — pick the first one with buildings/roads.
                    candidates = alias_to_canonicals.get(alias, [alias])
                    for canonical in candidates:
                        houses = buildings_by_street.get(canonical)
                        if houses:
                            c = _street_centroid(houses)
                            if c:
                                geometry = {"type": "Point", "coordinates": c}
                                match_quality = "street"
                                geometry_source = "buildings_street_named"
                                break
                        roads_for_street = roads_by_street.get(canonical)
                        if roads_for_street:
                            c = _roads_centroid(roads_for_street)
                            if c:
                                geometry = {"type": "Point", "coordinates": c}
                                match_quality = "street"
                                geometry_source = "roads_centroid"
                                break

        # Step 5: tashrir map markers (only for גוננים א-ו — the PDF covers only that area).
        # The PDF shows numbered markers per project. Use proj_num to look up coordinates.
        if not geometry and area == "גוננים א-ו" and tashrir_markers:
            num_str = str(proj_num).strip() if proj_num is not None else ""
            # The xlsx project_num may be "1" or "מתחם 1" — extract trailing digits
            m = re.search(r"(\d{1,3})\s*$", num_str)
            if m:
                hits = tashrir_markers.get(m.group(1)) or []
                if hits:
                    # Multiple markers per number (one per domain on PDF) — pick the centroid.
                    avg_lon = sum(p[0] for p in hits) / len(hits)
                    avg_lat = sum(p[1] for p in hits) / len(hits)
                    geometry = {"type": "Point", "coordinates": [avg_lon, avg_lat]}
                    match_quality = "tashrir"
                    geometry_source = "tashrir_marker"

        if not geometry:
            # Fallback: sub-neighborhood centroid (more accurate than minahak-wide)
            sub_c = sub_centroid_by_area.get(area)
            if sub_c:
                geometry = {"type": "Point", "coordinates": sub_c}
                geometry_source = "sub_neighborhood_centroid"
            else:
                geometry = {"type": "Point", "coordinates": minahak_centroid}
                geometry_source = "minahak_centroid"
            match_quality = match_quality or None

        # --- Overlap detection ---------------------------------------------
        overlap_plan_names = []
        overlap_features = []
        if match_quality == "plan_number" and plan_number in plans_by_name:
            plan_feat = plans_by_name[plan_number]
            pp = plan_feat["properties"]
            shavatz_in_prog = pp.get("shavatz_in_prog") or 0
            units_in = pp.get("units_in") or 0
            try:
                shavatz_val = float(shavatz_in_prog) if shavatz_in_prog not in ("", None) else 0
            except (TypeError, ValueError):
                shavatz_val = 0
            try:
                units_in_val = float(units_in) if units_in not in ("", None) else 0
            except (TypeError, ValueError):
                units_in_val = 0
            # Treat as overlap if the plan has any incoming public-building program
            # OR any incoming residential units (programa rec applies in either case).
            if shavatz_val > 0 or units_in_val > 0:
                overlap_plan_names.append(plan_number)
                overlap_features.append(pp.get("plan_name_he") or plan_number)

        # --- Build feature -------------------------------------------------
        # Convert MultiPolygon geometry centroid to point if it's the whole big
        # plan (would be confusing to color a 30-dunam polygon as a single rec).
        # Strategy:
        #   • domain=='transport' polygons → centroid Point (then post-process
        #     turns it into a road/passage LineString as appropriate).
        #   • LINEAR features (מעבר ה"ר / מעבר חציה / שביל / רחוב חדש) regardless
        #     of domain → also centroid Point. Otherwise a Polygon from parcels
        #     stays a Polygon and renders as a circleMarker instead of a line.
        _svc_linear = service_he and (
            "מעבר ה" in service_he or "הולכי" in service_he
            or "חציה" in service_he
            or "שביל" in service_he
            or "רחוב חדש" in service_he
        )
        if (domain == "transport" or _svc_linear) and geometry and geometry.get("type") in ("Polygon", "MultiPolygon") and match_quality != "manual_override":
            # Convert polygon-from-parcel to centroid Point for transport features
            # (they'll be re-rendered as small line/rect by _make_synthetic_polygon).
            # Skip manual_override — those polygons (e.g. area_polygon for sub-neighborhood
            # scope) are intentional and should NOT be collapsed.
            c = feature_centroid({"geometry": geometry})
            if c:
                geometry = {"type": "Point", "coordinates": c}

        # Final safety: if the computed geometry falls OUTSIDE the מינה"ק boundary, snap
        # the feature to the sub-neighborhood centroid (the original area name from xlsx).
        # This can happen when a street name extends across multiple parts of the city.
        # Skip for manual overrides — the user explicitly chose the coords, even if
        # they're on the boundary or just outside the minahak union.
        if (minahak_union_geom is not None and geometry and ShPoint is not None
                and shape is not None and match_quality != "manual_override"):
            try:
                g_test = shape(geometry)
                rep_pt = g_test.representative_point() if hasattr(g_test, "representative_point") else None
                if rep_pt and not minahak_union_geom.contains(rep_pt):
                    sub_c = sub_centroid_by_area.get(area)
                    if sub_c:
                        geometry = {"type": "Point", "coordinates": sub_c}
                        match_quality = "approx"
                        geometry_source = "sub_neighborhood_centroid"
            except Exception:
                pass

        # Convert Point/MultiPoint geometries to a synthetic Polygon shaped per
        # (domain, service) — and remember `marker_kind` for the JS layer to vary
        # fill/stroke per legend symbol.
        marker_kind = None
        if geometry and geometry.get("type") == "Point":
            lon, lat = geometry["coordinates"][0], geometry["coordinates"][1]
            # Decide geometry per (domain, service):
            #   - Crosswalks (מעבר חציה)     → stay a Point (frontend renders a small dot)
            #   - Pedestrian passages (מעבר ה"ר / הולכי) → become a LineString along the nearest road
            #   - New roads / streets         → LineString along the nearest road
            #   - Other (programa / שצ"פ)    → stay a Point (frontend renders a colored ring)
            road_geom = None
            s = (service_he or "")
            if (domain in ("transport", "public_space") and roads.get("features")
                    and match_quality != "manual_override"):
                # Tashrit-label FALLBACK: use the labeled position from the projector's
                # planning map ONLY when we don't already have a high-quality position
                # from xlsx (parcels / plan_number / shapefile). The PDF→WGS84 affine
                # transform has 5-15m of drift, so we prefer xlsx-derived positions.
                # Manual overrides skip this block entirely — their geometry is final.
                proj_num_int = None
                try:
                    pn_match = re.search(r"(\d+)", str(proj_num)) if proj_num is not None else None
                    if pn_match:
                        proj_num_int = int(pn_match.group(1))
                except Exception:
                    proj_num_int = None
                tashrit_typ = None
                if "מעבר חציה" in s or ("חציה" in s and "מעבר" not in s.replace("חציה", "")):
                    tashrit_typ = "חציה"
                elif "מעבר ה" in s or "הולכי" in s:
                    tashrit_typ = "מעבר"
                # Only override when the current position is a low-quality fallback.
                LOW_QUALITY_SOURCES = {
                    "minahak_centroid",
                    "sub_neighborhood_centroid",
                    "nominatim_cache",
                    "tashrir_marker",
                }
                if (tashrit_typ and proj_num_int is not None and tashrir_typed
                        and geometry_source in LOW_QUALITY_SOURCES):
                    tkey = (area, tashrit_typ, proj_num_int)
                    if tkey in tashrir_typed:
                        t_lon, t_lat = tashrir_typed[tkey]
                        geometry = {"type": "Point", "coordinates": [t_lon, t_lat]}
                        lon, lat = t_lon, t_lat
                        geometry_source = f"tashrit_label_{tashrit_typ}"

                if "מעבר חציה" in s or "חציה" in s:
                    # Crosswalk: KEEP as Point (precise location); snap onto the
                    # nearest road so it sits on the street, not on a building.
                    marker_kind = "crossing_point"
                    snapped = snap_point_to_nearest_road(lon, lat, roads.get("features", []))
                    if snapped:
                        geometry = {"type": "Point", "coordinates": [snapped[0], snapped[1]]}
                        lon, lat = snapped[0], snapped[1]
                        if not geometry_source.endswith("_snapped"):
                            geometry_source = (geometry_source or "") + "_snapped"
                elif "מעבר ה" in s or "הולכי" in s:
                    # Passage geometry priority:
                    #   1) The actual GREEN LINE drawn on the תשריט (now that the new
                    #      per-chip georef gives 17m median accuracy, we can use the
                    #      drawn line — the chip is just a legend pointer, the line
                    #      is the real feature). Each vertex gets snapped to the
                    #      nearest road within 25m for road-alignment precision.
                    #   2) Else: a segment from `find_nearest_road_segment` using the
                    #      xlsx anchor and `units_required` length.
                    tashrit_line = None
                    # Skip the tashrit-line override when xlsx already supplied a
                    # geographic identifier (location_id) or we have a high-quality
                    # match (parcels/plan_number/shapefile/manual_override). The
                    # chip-derived line is a fallback for chip-only features; using
                    # it on top of a real parcel polygon would replace a true polygon
                    # with an approximate LineString.
                    HIGH_QUALITY = {"parcels", "plan_number", "shapefile", "manual_override"}
                    if not location_id and match_quality not in HIGH_QUALITY:
                        passage_lines = (tashrir_typed.get("__passage_lines__")
                                         if isinstance(tashrir_typed, dict) else None) or {}
                        if proj_num_int is not None:
                            tashrit_line = passage_lines.get((area, "מעבר", proj_num_int))
                    if tashrit_line and len(tashrit_line) >= 2:
                        # Snap each vertex to the nearest road (within ~25m). Vertices
                        # in open courtyards stay as-is (passage runs between buildings).
                        snapped_pts = []
                        for pt in tashrit_line:
                            sp = snap_point_to_nearest_road(pt[0], pt[1], roads.get("features", []), max_dist_deg=0.00025)
                            snapped_pts.append([round(sp[0], 7), round(sp[1], 7)] if sp else pt)
                        # Dedupe consecutive identical vertices that may collapse.
                        dedup = []
                        for p in snapped_pts:
                            if not dedup or (abs(dedup[-1][0] - p[0]) > 1e-7 or abs(dedup[-1][1] - p[1]) > 1e-7):
                                dedup.append(p)
                        # Clip to xlsx units_required when it's a linear length —
                        # the PDF's drawn line is approximate, xlsx is the spec value.
                        u_norm_t = (str(units_label or "")
                                   .replace("׳", "'").replace("״", '"').strip().strip("_"))
                        target_len_m = None
                        if (isinstance(units_required, (int, float)) and units_required > 0
                                and u_norm_t in {"מ'", "מ\"א", "מ", "מא"}):
                            target_len_m = max(20.0, min(float(units_required), 400.0))
                        elif (isinstance(units_required, (int, float)) and units_required > 0
                              and u_norm_t in {"ק\"מ", "קמ"}):
                            target_len_m = max(20.0, min(float(units_required) * 1000.0, 1000.0))
                        if target_len_m and len(dedup) >= 2 and shape is not None:
                            try:
                                import math
                                from shapely.geometry import LineString as ShLine
                                from shapely.ops import substring
                                ls = ShLine(dedup)
                                # Approx degree→meter at this latitude
                                lat0 = (dedup[0][1] + dedup[-1][1]) / 2
                                cos_lat = max(0.1, math.cos(math.radians(lat0)))
                                cur_len_m = ls.length * 111000 * cos_lat
                                if cur_len_m > target_len_m * 1.10:
                                    # Too long: clip ±target/2 around the midpoint
                                    mid_d = ls.length / 2
                                    half = target_len_m / 2 / 111000 / cos_lat
                                    sub = substring(ls, max(0, mid_d - half), min(ls.length, mid_d + half))
                                    if not sub.is_empty:
                                        dedup = [[round(x, 7), round(y, 7)] for (x, y) in sub.coords]
                                # If too SHORT, we leave it (don't synthesize extra path)
                            except Exception:
                                pass
                        if len(dedup) >= 2:
                            road_geom = {"type": "LineString", "coordinates": dedup}
                            marker_kind = "passage_line"
                            geometry_source = "tashrit_line_מעבר"
                    # Fallback: full road, clipped to units_required length.
                    seg = None if road_geom else find_nearest_road_full((lon, lat), roads["features"], 0.0008)
                    if seg:
                        try:
                            import math
                            cos_lat = max(0.1, math.cos(math.radians(lat)))
                            from shapely.geometry import LineString as ShLine, Point as P
                            from shapely.ops import substring
                            ls = ShLine(seg)
                            tp = ls.project(P(lon, lat))
                            # Length from units_required (meters) when available;
                            # fall back to 30m.
                            length_m = 30.0
                            u_norm = (str(units_label or "")
                                      .replace("׳", "'").replace("״", '"').strip().strip("_"))
                            if (isinstance(units_required, (int, float)) and units_required > 0
                                    and u_norm in {"מ'", "מ\"א", "מ", "מא"}):
                                length_m = max(20.0, min(float(units_required), 400.0))
                            delta = length_m / 2.0 / 111000.0 / cos_lat
                            sub = substring(
                                ls,
                                max(0, tp - delta),
                                min(ls.length, tp + delta),
                            )
                            if not sub.is_empty:
                                road_geom = mapping(sub)
                                marker_kind = "passage_line"
                        except Exception:
                            road_geom = None
                elif "רחוב חדש" in s or "תח" in s:
                    seg = find_nearest_road_full((lon, lat), roads["features"], 0.0008)
                    if seg:
                        try:
                            import math
                            cos_lat = max(0.1, math.cos(math.radians(lat)))
                            from shapely.geometry import LineString as ShLine, Point as P
                            from shapely.ops import substring
                            ls = ShLine(seg)
                            tp = ls.project(P(lon, lat))
                            # Length from xlsx units_required if linear; else 60m default.
                            length_m = 60.0
                            u_norm = (str(units_label or "")
                                      .replace("׳", "'").replace("״", '"').strip().strip("_"))
                            if (isinstance(units_required, (int, float)) and units_required > 0
                                    and u_norm in {"מ'", "מ\"א", "מ", "מא"}):
                                length_m = max(20.0, min(float(units_required), 400.0))
                            elif u_norm in {"ק\"מ", "קמ"} and isinstance(units_required, (int, float)):
                                length_m = max(20.0, min(float(units_required) * 1000.0, 1000.0))
                            delta = length_m / 2.0 / 111000.0 / cos_lat
                            sub = substring(ls, max(0, tp - delta), min(ls.length, tp + delta))
                            if not sub.is_empty:
                                road_geom = mapping(sub)
                                marker_kind = "road_segment"
                        except Exception:
                            road_geom = {"type": "LineString", "coordinates": seg}
                            marker_kind = "road_segment"
                else:
                    # Generic units-based rule: if the row is measured in length units
                    # (מ' / מ"א / ק"מ) → treat as a linear feature along the nearest road.
                    # Catches services we haven't enumerated explicitly: שדרוג רחובות,
                    # שביל אופניים, גשר, חיבור לשצ"פ (when מ"א), etc.
                    if units_geometry_kind(units_label) == "line":
                        seg = find_nearest_road_full((lon, lat), roads["features"], 0.0010)
                        if seg:
                            try:
                                import math
                                cos_lat = max(0.1, math.cos(math.radians(lat)))
                                from shapely.geometry import LineString as ShLine, Point as P
                                from shapely.ops import substring
                                ls = ShLine(seg)
                                tp = ls.project(P(lon, lat))
                                # Use the required length from units_required when it
                                # is a positive number — otherwise default to 60m.
                                length_m = 60.0
                                if isinstance(units_required, (int, float)) and units_required > 0:
                                    # Treat ק"מ as kilometers
                                    u_norm = (str(units_label or "")
                                              .replace("׳", "'").replace("״", '"').strip().strip("_"))
                                    if u_norm in {"ק\"מ", "קמ"}:
                                        length_m = float(units_required) * 1000.0
                                    else:
                                        length_m = float(units_required)
                                    # Clamp to a sane on-map length
                                    length_m = max(20.0, min(length_m, 400.0))
                                delta = length_m / 2.0 / 111000.0 / cos_lat
                                sub = substring(
                                    ls,
                                    max(0, tp - delta),
                                    min(ls.length, tp + delta),
                                )
                                if not sub.is_empty:
                                    # Choose marker_kind based on service hint
                                    if "שביל" in s or "אופניים" in s:
                                        marker_kind = "bike_line"
                                    elif "גשר" in s:
                                        marker_kind = "bridge_line"
                                    elif "שדרוג" in s and "רחוב" in s:
                                        marker_kind = "road_segment"
                                    else:
                                        marker_kind = "passage_line"
                                    road_geom = mapping(sub)
                            except Exception:
                                road_geom = None
            if road_geom:
                geometry = road_geom
            elif marker_kind == "crossing_point":
                # Geometry stays a Point. No synthetic polygon for crossings.
                pass
            else:
                geometry, marker_kind = _make_synthetic_polygon(lon, lat, domain, service_he)
        elif geometry and geometry.get("type") == "MultiPoint":
            # Bus stops (תח"צ) keep MultiPoint geometry and get marker_kind=small_circle;
            # the row is split into separate Point features below (one per coord)
            # so each renders independently via the existing Point bus-stop path.
            # Everything else gets the synthetic polygon treatment.
            if service_he and "תח\"צ" in service_he:
                marker_kind = "small_circle"
            else:
                polys = []
                for pt in geometry["coordinates"]:
                    p, mk = _make_synthetic_polygon(pt[0], pt[1], domain, service_he)
                    polys.append(p["coordinates"])
                    marker_kind = mk  # all use same kind
                geometry = {"type": "MultiPolygon", "coordinates": polys}
        else:
            # Real lot polygons: assign marker_kind based on domain/service for consistent styling
            if domain == "programa":
                marker_kind = "metaham_square" if service_he and "מתחם" in service_he else "shavatz_future_rect"
            elif domain == "public_space":
                if service_he and ("מתחם" in service_he):
                    marker_kind = "metaham_square"
                elif service_he and "גשר" in service_he:
                    # Bridge service → bridge_line (dark-blue styling)
                    marker_kind = "bridge_line"
                else:
                    # Detect passages via service text OR project name (some rows
                    # have service="הנגשה" but a name like "שדרוג והנגשה מעבר ה\"ר…").
                    _passage_kw = ("מעבר", "ה\"ר", "הולכי", "שביל", "חיבור")
                    _svc_match = service_he and any(k in service_he for k in _passage_kw)
                    _name_match = proj_name and any(k in proj_name for k in _passage_kw)
                    # Road-like services in public_space domain (e.g. ציר הרצוג):
                    # surface them as road_segment so the styling matches transport.
                    _road_kw = ("שדרוג רחובות", "רחוב חדש", "תח\"צ", "שיפור רמת שירות", "רחוב")
                    _road_svc_match = service_he and any(k in service_he for k in _road_kw)
                    if _svc_match or _name_match:
                        _gt = (geometry or {}).get("type") or ""
                        if _gt in ("LineString", "MultiLineString"):
                            marker_kind = "passage_line"
                        elif _gt in ("Polygon", "MultiPolygon"):
                            marker_kind = "passage_polygon"
                        else:
                            marker_kind = "passage_point"
                    elif _road_svc_match:
                        marker_kind = "road_segment"
                    else:
                        marker_kind = "shatzap_rect"
            elif domain == "transport":
                marker_kind = "road_segment"
            else:
                marker_kind = "shavatz_future_rect"

        # ---- Polygon-render eligibility flags (Phase-1 visual refinements) ----
        # #1 עיבוי שטחים חומים: programa polygon, NOT הפרשה מבונה, area > 1 dunam.
        # #2 שצ"פ/גינה: public_space polygon when service mentions שצ"פ or גינה.
        # #3 Crossings perpendicular to road: capture road bearing for crossing_point.
        _HM_CODES = {1250, 1300, 1410, 1480, 1492, 1550, 1576, 1578, 1604}
        is_hafrasha_mevuna = False
        if domain == "programa":
            # Aggregate every text field that might mention הפרשה מבונה / הקצאה /
            # שטחי הפרשה. Catches more cases than just land_use alone.
            _texts = [
                (karteset_rec or {}).get("land_use") or "",
                (karteset_rec or {}).get("recommendations") or "",
                (karteset_rec or {}).get("programmatic_potential") or "",
                (karteset_rec or {}).get("allowed_uses") or "",
                (karteset_rec or {}).get("existing_built") or "",
                description or "",
            ]
            _hm_patterns = ("הפרשה מבונה", "הקצאת מבנה", "שטחי הפרשה", "הקצאה מבונה")
            for _t in _texts:
                if any(pat in _t for pat in _hm_patterns):
                    is_hafrasha_mevuna = True
                    break
            if not is_hafrasha_mevuna and matched_features:
                if any(
                    (f.get("properties") or {}).get("mavat_code") in _HM_CODES
                    for f in matched_features
                ):
                    is_hafrasha_mevuna = True

        lot_area_m2 = None
        is_eivuy_brown = False
        is_shatzap_polygon = False
        is_metaham_polygon = False
        # Metaham-master-plan: a "תכנון אב מתחמית" recommendation that should be
        # rendered as a light-brown polygon covering the entire compound (not a
        # small marker). Triggered by service text or project name keywords.
        _metaham_kw = ("תכנון אב", "תוכנית אב", "מתחמית", "מתחמי להתחדשות")
        _is_metaham = (
            (service_he and any(k in service_he for k in _metaham_kw))
            or (proj_name and any(k in proj_name for k in _metaham_kw))
            or (proj_name and "מתחם" in proj_name and service_he and "מתחם" in service_he)
        )
        _force_point = (str(proj_name or "").strip(), area) in FORCE_POINT_FEATURES
        if geometry and geometry.get("type") in ("Polygon", "MultiPolygon"):
            lot_area_m2 = polygon_area_m2(geometry)
            if _force_point:
                pass  # user-opt-out — render as centroid point
            elif _is_metaham and (lot_area_m2 or 0) > 1000:
                is_metaham_polygon = True
            elif domain == "programa" and not is_hafrasha_mevuna and (lot_area_m2 or 0) > 1000:
                is_eivuy_brown = True
            elif domain == "public_space" and service_he and (
                "שצ" in service_he or "גינ" in service_he
            ):
                is_shatzap_polygon = True

        # For metaham compounds, detect inner שצ"פ polygons from yiud_karka_kayam
        # so the frontend can render them prominently (compound = faint outline,
        # שצ"פים inside = colored fill per category). Filters by centroid-in-polygon
        # + min area. Categories + extra trail features come from per-compound
        # overrides INNER_SHATZAP_CATEGORIES / INNER_TRAILS_BY_COMPOUND.
        inner_shatzaps = []  # list of {category, label, area_m2, geometry}
        inner_trails = []    # list of {label, kind, geometry}
        if is_metaham_polygon and yiud_kayam_shatzap_features and shape:
            try:
                compound_shape = shape(geometry)
                if not compound_shape.is_valid:
                    compound_shape = compound_shape.buffer(0)
                cat_overrides = INNER_SHATZAP_CATEGORIES.get(
                    (str(proj_name or "").strip(), area), {}
                )
                for s in yiud_kayam_shatzap_features:
                    if not compound_shape.intersects(s["geom"]):
                        continue
                    if not compound_shape.contains(s["geom"].centroid):
                        continue
                    a = polygon_area_m2(s["raw"])
                    if a < 100:
                        continue
                    # Match to nearest category override (within ~30m — tighter
                    # to avoid mis-matching small extra polygons not labeled in PDF)
                    cx, cy = s["geom"].centroid.x, s["geom"].centroid.y
                    cat_info = None
                    best_dist = 999.0
                    for ((ox, oy), info) in cat_overrides.items():
                        d = ((ox - cx) ** 2 + (oy - cy) ** 2) ** 0.5
                        if d < best_dist and d < 0.00027:  # ~30m
                            best_dist = d
                            cat_info = info
                    inner_shatzaps.append({
                        "category": (cat_info or {}).get("category", "מוצלח"),
                        "label": (cat_info or {}).get("label"),
                        "area_m2": round(a, 1),
                        "geometry": s["raw"],
                    })
                # Inner trails for this compound (e.g. yellow path on PDF p18)
                for trail in INNER_TRAILS_BY_COMPOUND.get(
                    (str(proj_name or "").strip(), area), []
                ):
                    inner_trails.append(trail)
            except Exception as e:
                print(f"      [inner_features] failed for {proj_name!r}: {e}")

        inner_shatzap_count = len(inner_shatzaps)
        inner_shatzap_total_m2 = round(sum(x["area_m2"] for x in inner_shatzaps), 1)
        inner_trail_count = len(inner_trails)

        road_bearing_deg = None
        if marker_kind == "crossing_point" and geometry and geometry.get("type") == "Point":
            cx, cy = geometry["coordinates"][:2]
            road_bearing_deg = road_bearing_at(cx, cy, roads.get("features", []))

        props = {
            "project_id": pid,
            "domain": domain,
            "domain_he": domain_he,
            "service_he": service_he,
            "service_key": service_key,
            "project_name": proj_name,
            "project_num": str(proj_num) if proj_num is not None else None,
            "description": description,
            "units_label": units_label,
            "units_required": units_required if isinstance(units_required, (int, float)) else (str(units_required) if units_required else None),
            "units_kind": units_geometry_kind(units_label),
            "plan_number": plan_number,
            "plan_raw": plan_raw,
            "plan_status": plan_status,
            "year_planning": year_planning,
            "year_execution": year_completion,
            "chumash": chumash,
            "chumash_raw": chumash_text,
            "sub_neighborhood": area,
            "location_id": str(location_id) if location_id is not None else None,
            "managing_body": managing,
            "estimate_nis": estimate,
            "pdf_link": pdf_link,
            "pdf_excerpt": pdf_excerpt,
            "match_quality": match_quality,
            "overlap_plan_names": overlap_plan_names,
            "geometry_source": geometry_source,
            "marker_kind": marker_kind,
            "lot_area_m2": round(lot_area_m2, 1) if lot_area_m2 is not None else None,
            "is_eivuy_brown": is_eivuy_brown,
            "is_shatzap_polygon": is_shatzap_polygon,
            "is_metaham_polygon": is_metaham_polygon,
            "inner_shatzaps": inner_shatzaps if is_metaham_polygon else [],
            "inner_trails": inner_trails if is_metaham_polygon else [],
            "inner_shatzap_count": inner_shatzap_count if is_metaham_polygon else 0,
            "inner_shatzap_total_m2": inner_shatzap_total_m2 if is_metaham_polygon else 0,
            "compound_principles": COMPOUND_PRINCIPLES.get(
                (str(proj_name or "").strip(), area)
            ) if is_metaham_polygon else None,
            "crosswalk_table_row": CROSSWALK_TABLE_DATA.get((area, str(proj_num or "").strip())),
            # Master-plan compound *proposals* (a specific service category) —
            # rendered in pink to distinguish from generic metaham polygons.
            # Covers the 4 compounds from "תכנון מתחמי 09.2023" PDF + מתחם הפסגה
            # in קטמונים ח-ט (separate "מצגת מתחם הפסגה" PDF) — all 5 share the
            # same xlsx service "הצעה להכנת תוכנית אב מתחמית להתחדשות".
            "is_metaham_pdf_tichnun": is_metaham_polygon and isinstance(service_he, str) and (
                "הצעה להכנת תוכנית אב מתחמית" in service_he
                or service_he.strip() == "מתחם, שצ\"פ"
            ),
            "road_bearing_deg": road_bearing_deg,
            "obstacles": str(obstacles).strip() if obstacles else None,
            "related_projects_text": str(related_projects).strip() if related_projects else None,
        }
        # Notes lookup precedence: (name, area, service) → (name, area)
        _nk_name = str(proj_name or "").strip()
        _note = (MANUAL_NOTES.get((_nk_name, area, str(service_he or "").strip()))
                 or MANUAL_NOTES.get((_nk_name, area)))
        if _note:
            props["manual_note"] = _note
        # Merge any enrichment fields from the official shapefile match (transport rows).
        if shp_enrich:
            props.update(shp_enrich)

        # Surface the structured public-space recommendation fields at the top level
        # so the popup can render a prominent banner (recommendation + categories +
        # land-use + dimensions). The icon labels above the page (שדרוג / פינת ישיבה
        # / תצפית / הנגשה etc.) come through as `categories`.
        if public_space_rec:
            props["ps_title"]          = public_space_rec.get("title")
            props["ps_categories"]     = public_space_rec.get("categories") or []
            props["ps_recommendation"] = public_space_rec.get("recommendation")
            props["ps_land_use"]       = public_space_rec.get("land_use")
            props["ps_dimensions"]     = public_space_rec.get("dimensions")
            props["ps_status"]         = public_space_rec.get("status")
            props["ps_location"]       = public_space_rec.get("location")
            props["ps_unique_feature"] = public_space_rec.get("unique_feature")
            props["ps_page"]           = public_space_rec.get("page")

        # Surface the structured "כרטסת שטחים חומים" (brown-land card) fields when
        # the programa row has a matching card. ייעוד ע"פ תב"ע + the recommendation
        # block are highlighted in the popup banner; the rest are placed in details.
        if karteset_rec:
            props["kar_address"]              = karteset_rec.get("address")
            props["kar_gush_helka"]           = karteset_rec.get("gush_helka")
            props["kar_taba_lot"]             = karteset_rec.get("taba_lot")
            props["kar_lot_area"]             = karteset_rec.get("lot_area")
            props["kar_land_use"]             = karteset_rec.get("land_use")
            props["kar_allowed_uses"]         = karteset_rec.get("allowed_uses")
            props["kar_public_sqm"]           = karteset_rec.get("public_sqm")
            props["kar_ownership"]            = karteset_rec.get("ownership")
            props["kar_existing_built"]       = karteset_rec.get("existing_built")
            props["kar_programmatic_potential"] = karteset_rec.get("programmatic_potential")
            props["kar_recommendations"]      = karteset_rec.get("recommendations")
            props["kar_page"]                 = karteset_page

        # Attach a "תיק שכונה" / "דוח מסכם" slide excerpt if one matches by area+name.
        tik_match = _find_tik_match(area, proj_name, tik_index)
        if tik_match:
            props["tik_title"] = tik_match["title"]
            props["tik_pdf_name"] = tik_match["pdf_name"]
            props["tik_page"] = tik_match["page"]
            props["tik_label"] = tik_match["label"]
            # Trim narrative text for popup (full text already capped at 1200 chars)
            props["tik_excerpt"] = tik_match["text"]

        # Reliable [lon, lat] for marker placement (always inside the polygon).
        props["display_point"] = feature_display_point(geometry)

        # Split MultiPoint bus stops into separate Point features so each renders
        # independently via the existing Point + small_circle path. Each split
        # feature gets a unique project_id suffix; project_name stays the same so
        # they share a popup and group in the sibling index.
        if (geometry and geometry.get("type") == "MultiPoint"
                and props.get("marker_kind") == "small_circle"):
            base_id = props.get("project_id") or ""
            for i, pt in enumerate(geometry["coordinates"]):
                sub_props = dict(props)
                sub_props["project_id"] = f"{base_id}-{i+1}" if base_id else None
                sub_geom = {"type": "Point", "coordinates": list(pt)}
                sub_props["display_point"] = feature_display_point(sub_geom)
                out_features.append({"type": "Feature", "geometry": sub_geom, "properties": sub_props})
        else:
            out_features.append({"type": "Feature", "geometry": geometry, "properties": props})

        overlap_idx[pid] = {
            "plan_number": plan_number,
            "match_quality": match_quality,
            "overlap_plan_names": overlap_plan_names,
            "overlap_features": overlap_features,
        }

        # --- Summary aggregation -------------------------------------------
        # Only count toward balance if units_required is a number AND service_key is set.
        if service_key and isinstance(units_required, (int, float)):
            sub = summary[area]
            bucket = sub["by_key"][service_key] if not service_key.startswith("extra:") else sub["extras"][service_key]
            bucket["units"] += float(units_required)
            if overlap_plan_names:
                bucket["overlap_units"] += float(units_required)
            bucket["projects"].append({
                "project_id": pid,
                "project_name": proj_name,
                "units_required": units_required,
                "chumash": chumash,
                "plan_number": plan_number,
                "plan_status": plan_status,
                "estimate_nis": estimate,
                "overlap": bool(overlap_plan_names),
            })

        match_stats[domain][match_quality or "none"] += 1

    if duplicate_count:
        print(f"      skipped {duplicate_count} duplicate project_id rows")

    # --- Emit standalone shapefile-only features ----------------------------
    # 1) Bus stops (layer2 points): not present in the xlsx as rows.
    # 2) Any layer1/3 polylines that no xlsx row matched.
    # These are still legitimate transport recommendations from the projector team.
    SHP_AREA_TO_AREA = {v: k for k, v in AREA_TO_SHP_SUB.items()}
    standalone_bus = 0
    for f in shp_bus_stops:
        sp = f.get("properties") or {}
        shp_sub = sp.get("שכונה")
        area = SHP_AREA_TO_AREA.get(shp_sub)
        if not area:
            continue  # outside gonenim (e.g. גבעת מרדכי)
        proj_name = sp.get("שם_פרוייקט") or sp.get("מהות") or "תחנת אוטובוס"
        pid = project_id(area, sp.get("id"), "extra:bus_stop")
        chumash_text = sp.get("חומש") or ""
        m_ch = re.search(r"(\d)", str(chumash_text))
        chumash_val = int(m_ch.group(1)) if m_ch else None
        props = {
            "project_id": pid,
            "domain": "transport",
            "domain_he": "תנועה",
            "service_he": sp.get("מהות") or "תחנת אוטובוס",
            "service_key": "extra:bus_stop",
            "project_name": proj_name,
            "project_num": str(sp.get("id") or ""),
            "description": sp.get("פירוט"),
            "units_label": "קומפלט",
            "units_required": 1,
            "plan_number": None,
            "plan_raw": None,
            "plan_status": sp.get("סטטוס"),
            "year_planning": sp.get("שנת_תכנון"),
            "year_execution": sp.get("שנת_ביצוע"),
            "chumash": chumash_val,
            "chumash_raw": chumash_text,
            "sub_neighborhood": area,
            "location_id": (f"{sp.get('גוש')}/{sp.get('חלקה')}"
                            if sp.get("גוש") and sp.get("חלקה") else None),
            "managing_body": sp.get("גורם_מנהל"),
            "estimate_nis": None,
            "pdf_link": sp.get("לינק_1") or sp.get("לינק_2"),
            "pdf_excerpt": None,
            "match_quality": "shapefile",
            "overlap_plan_names": [],
            "geometry_source": "projector_official_layer2",
            "marker_kind": "small_circle",
            "is_shp_only": True,
        }
        props["display_point"] = feature_display_point(f.get("geometry"))
        out_features.append({"type": "Feature", "geometry": f.get("geometry"), "properties": props})
        standalone_bus += 1
    print(f"      [shp] emitted {standalone_bus} standalone bus-stop features")

    # 2) Polylines from layer1/3 that no xlsx row matched.
    standalone_poly = 0
    for (shp_sub, norm_name), feats in shp_official_index.items():
        if (shp_sub, norm_name) in shp_used_keys:
            continue
        area = SHP_AREA_TO_AREA.get(shp_sub)
        if not area:
            continue  # skip גבעת מרדכי (outside gonenim)
        # Prefer layer3 (richer) when both exist
        feat = None
        for f in feats:
            if f.get("properties", {}).get("_shp_layer") == "layer3":
                feat = f
                break
        feat = feat or feats[0]
        sp = feat.get("properties") or {}
        proj_name = sp.get("שם_פרוייקט")
        if not proj_name:
            continue
        # Skip if no geometry (1 layer3 feature has null geometry)
        if not feat.get("geometry") or not feat["geometry"].get("coordinates"):
            continue
        pid = project_id(area, sp.get("id"), "transport_shp")
        chumash_text = sp.get("חומש") or ""
        m_ch = re.search(r"(\d)", str(chumash_text))
        chumash_val = int(m_ch.group(1)) if m_ch else None
        # Build a project name + description that surfaces the source attrs
        desc_parts = []
        for k in ("מהות", "פירוט", "אלמנט"):
            v = sp.get(k)
            if v:
                desc_parts.append(str(v))
        # Derive a service_he label per shapefile type
        mahut = sp.get("מהות") or ""
        sug = sp.get("סוג_שביל") or ""
        if "אופניים" in mahut or "אופניים" in sug:
            svc_he = "שביל אופניים"
        elif "חיבור" in mahut or "רחוב חדש" in mahut:
            svc_he = "רחוב חדש"
        elif "מיתון" in mahut:
            svc_he = "מיתון תנועה"
        elif "מעבר" in mahut or "ה ר" in mahut or "הולכי" in mahut:
            svc_he = "מעבר הולכי רגל"
        else:
            svc_he = mahut or "פרוייקט תנועה"
        props = {
            "project_id": pid,
            "domain": "transport",
            "domain_he": "תנועה",
            "service_he": svc_he,
            "service_key": "extra:transport",
            "project_name": proj_name,
            "project_num": str(sp.get("id") or ""),
            "description": " · ".join(desc_parts) if desc_parts else None,
            "units_label": "קומפלט",
            "units_required": 1,
            "plan_number": None,
            "plan_raw": None,
            "plan_status": sp.get("סטטוס"),
            "year_planning": sp.get("שנת_תכנון"),
            "year_execution": sp.get("שנת_ביצוע"),
            "chumash": chumash_val,
            "chumash_raw": chumash_text,
            "sub_neighborhood": area,
            "location_id": None,
            "managing_body": sp.get("גורם_מנהל"),
            "estimate_nis": None,
            "pdf_link": sp.get("לינק_1") or sp.get("לינק_2"),
            "pdf_excerpt": None,
            "match_quality": "shapefile",
            "overlap_plan_names": [],
            "geometry_source": f"projector_official_{sp.get('_shp_layer', '')}",
            "marker_kind": "road_segment",
            "is_shp_only": True,
        }
        # Carry the same enrich fields as the matched case so the popup can show them.
        for k_src, k_dst in [
            ("היררכיה", "shp_hierarchy"),
            ("סטריות", "shp_oneway"),
            ("מיתון", "shp_calming"),
            ("מס_נתיבים", "shp_lanes"),
            ("רוחב_זכות_דרך", "shp_row_m"),
            ("סוג_שביל", "shp_path_kind"),
            ("אלמנט", "shp_element"),
            ("מהות", "shp_essence"),
            ("Park_north", "shp_park_n"),
            ("Park_south", "shp_park_s"),
            ("Park_east", "shp_park_e"),
            ("Park_west", "shp_park_w"),
        ]:
            v = sp.get(k_src)
            if v not in (None, "", "--"):
                props[k_dst] = v
        # Attach tik-shchunah excerpt for standalone shp features too.
        tik_match = _find_tik_match(area, proj_name, tik_index)
        if tik_match:
            props["tik_title"] = tik_match["title"]
            props["tik_pdf_name"] = tik_match["pdf_name"]
            props["tik_page"] = tik_match["page"]
            props["tik_label"] = tik_match["label"]
            props["tik_excerpt"] = tik_match["text"]
        props["display_point"] = feature_display_point(feat.get("geometry"))
        out_features.append({"type": "Feature", "geometry": feat.get("geometry"), "properties": props})
        standalone_poly += 1
    print(f"      [shp] emitted {standalone_poly} standalone polyline features (unmatched in xlsx)")

    # Emit EXTRA_FEATURES — crosswalks/items not in xlsx but present in PDF tables.
    # Each has Point geometry from user-provided coords + crosswalk_table_row data.
    extra_emitted = 0
    for ext_pid, ext_def in EXTRA_FEATURES.items():
        if not ext_def or not ext_def.get("geometry"):
            continue
        ext_props = dict(ext_def.get("properties") or {})
        ext_props["project_id"] = ext_pid
        ext_props.setdefault("geometry_source", "extra_features_manual")
        ext_props.setdefault("match_quality", "manual_override")
        ext_props.setdefault("marker_kind", "crossing_point")
        ext_props.setdefault("domain", "transport")
        ext_props.setdefault("domain_he", "תנועה")
        if "display_point" not in ext_props:
            ext_props["display_point"] = feature_display_point(ext_def["geometry"])
        out_features.append({"type": "Feature", "geometry": ext_def["geometry"], "properties": ext_props})
        extra_emitted += 1
    if extra_emitted:
        print(f"      [extra] emitted {extra_emitted} EXTRA_FEATURES not in xlsx")

    print("[4/6] Match-quality distribution by domain:")
    for dom, stats in match_stats.items():
        total = sum(stats.values())
        line = ", ".join(f"{k}={v}" for k, v in sorted(stats.items()))
        print(f"      {dom:14s} ({total:3d}): {line}")

    # Phase-1 visual flags coverage (sanity check).
    n_eivuy = sum(1 for f in out_features if f["properties"].get("is_eivuy_brown"))
    n_shatzap = sum(1 for f in out_features if f["properties"].get("is_shatzap_polygon"))
    n_metaham = sum(1 for f in out_features if f["properties"].get("is_metaham_polygon"))
    n_inner_shatzap = sum(1 for f in out_features if f["properties"].get("inner_shatzaps"))
    inner_counts = [f["properties"].get("inner_shatzap_count", 0) for f in out_features if f["properties"].get("inner_shatzaps")]
    n_bearing = sum(1 for f in out_features if f["properties"].get("road_bearing_deg") is not None)
    n_crossing = sum(1 for f in out_features if f["properties"].get("marker_kind") == "crossing_point")
    print(f"      polygon-render flags: is_eivuy_brown={n_eivuy}, "
          f"is_shatzap_polygon={n_shatzap}, is_metaham_polygon={n_metaham} "
          f"(inner_שצ\"פ on {n_inner_shatzap}/{n_metaham}, total {sum(inner_counts)} sub-polys), "
          f"road_bearing_deg={n_bearing}/{n_crossing} crossings")

    # Resolve RELATED_FEATURES — second pass after all features are built.
    # Builds (name, area) → first matching feature index so we can attach a
    # compact reference (project_id + display_point) to each related entry.
    _feat_index = {}
    for f in out_features:
        prop = f["properties"]
        key = (str(prop.get("project_name") or "").strip(), prop.get("sub_neighborhood") or "")
        if key not in _feat_index:
            _feat_index[key] = prop
    n_related = 0
    for f in out_features:
        prop = f["properties"]
        my_key = (str(prop.get("project_name") or "").strip(), prop.get("sub_neighborhood") or "")
        targets = RELATED_FEATURES.get(my_key) or []
        if not targets:
            continue
        rel_list = []
        for t_name, t_area in targets:
            t = _feat_index.get((t_name, t_area))
            if t:
                rel_list.append({
                    "project_id": t.get("project_id"),
                    "name": t.get("project_name"),
                    "area": t.get("sub_neighborhood"),
                    "display_point": t.get("display_point"),
                })
        if rel_list:
            prop["related_features"] = rel_list
            n_related += 1
    print(f"      related_features attached to {n_related} features")

    # שביל גוננים — trim overlapping sub-segments so each child feature owns
    # its piece. The trail mentions itself in the children's descriptions; we
    # detect those, subtract their buffered geometries from the trail, and
    # append a "(חלק משביל גוננים)" note to each child for popup discoverability.
    try:
        from shapely.geometry import shape as _shape, mapping as _mapping
        from shapely.ops import unary_union as _uu
        SHVIL_NAME = "שביל גוננים"
        SHVIL_BUFFER_M = 10.0  # ~10m around child geometries
        BUFFER_DEG = SHVIL_BUFFER_M / 111000.0
        shvil_feat = next((f for f in out_features
                           if (f.get("properties") or {}).get("project_name") == SHVIL_NAME), None)
        if shvil_feat:
            shvil_geom = _shape(shvil_feat["geometry"])
            children = []
            for f in out_features:
                prop = f.get("properties") or {}
                if prop.get("project_name") == SHVIL_NAME:
                    continue
                desc = str(prop.get("description") or "")
                if SHVIL_NAME in desc and f.get("geometry"):
                    children.append(f)
            trimmed_len_m = None
            if children:
                child_geoms = []
                for c in children:
                    try:
                        cg = _shape(c["geometry"])
                        child_geoms.append(cg.buffer(BUFFER_DEG))
                    except Exception:
                        pass
                if child_geoms:
                    union = _uu(child_geoms)
                    trimmed = shvil_geom.difference(union)
                    if not trimmed.is_empty and hasattr(trimmed, "length") and trimmed.length > 0:
                        shvil_feat["geometry"] = _mapping(trimmed)
                        shvil_feat["properties"]["display_point"] = feature_display_point(shvil_feat["geometry"])
                        trimmed_len_m = trimmed.length * 111000
                # Append the "part of שביל גוננים" note to each child.
                NOTE_MARK = "חלק משביל גוננים"
                for c in children:
                    prop = c["properties"]
                    existing_note = prop.get("manual_note") or ""
                    if NOTE_MARK not in existing_note:
                        prop["manual_note"] = (existing_note + " " + NOTE_MARK).strip() if existing_note else NOTE_MARK
                if trimmed_len_m is not None:
                    print(f"      שביל גוננים: trimmed {len(children)} overlapping child segments "
                          f"(remaining trail length: {trimmed_len_m:.0f}m)")
    except Exception as e:
        print(f"      [warn] שביל גוננים trim failed: {e}")

    # Convert defaultdicts in summary
    summary_out = {}
    for sub_name, sub in summary.items():
        summary_out[sub_name] = {
            "by_key": {k: dict(v) for k, v in sub["by_key"].items()},
            "extras": [
                {"key": k, "label": EXTRA_LABELS.get(k, k), **v}
                for k, v in sub["extras"].items()
            ],
        }
    # Add a flat "extras" view per service_key to make it easy in JS
    # (already done above per area)

    print("[5/6] Writing outputs to data dirs:")
    # Split features whose geometry came from the official projector shapefiles
    # (פורמט שכבות מידע - פרויקטים מומלצים.zip) into a separate "tzatal" layer.
    TZATAL_SOURCES = {
        "projector_official_layer1",
        "projector_official_layer2",
        "projector_official_layer3",
    }
    tzatal_features = [f for f in out_features
                       if (f.get("properties") or {}).get("geometry_source") in TZATAL_SOURCES]
    main_features = [f for f in out_features
                     if (f.get("properties") or {}).get("geometry_source") not in TZATAL_SOURCES]
    geojson_out = {"type": "FeatureCollection", "features": main_features}
    tzatal_out = {"type": "FeatureCollection", "features": tzatal_features}
    print(f"      main projector: {len(main_features)} features")
    print(f"      tzatal (פרוייקטים מרחביים): {len(tzatal_features)} features")
    files_to_write = {
        "projector_gonenim.geojson": geojson_out,
        "projector_gonenim_tzatal.geojson": tzatal_out,
        "projector_gonenim_overlap.json": overlap_idx,
        "projector_gonenim_summary.json": summary_out,
    }
    for d in DATA_DIRS:
        for fname, payload in files_to_write.items():
            out_path = d / fname
            out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"      wrote {out_path} ({out_path.stat().st_size:,} bytes)")

    # Persist any new Nominatim entries we picked up during the build.
    _save_geocode_cache(geocode_cache)

    print(f"[6/6] Done. {len(out_features)} features.")
    if PENDING_REVIEW:
        print(f"\n*** {len(PENDING_REVIEW)} features marked PENDING_REVIEW (need user-provided coords): ***")
        for k, note in PENDING_REVIEW.items():
            print(f"      {k}: {note}")


if __name__ == "__main__":
    main()
