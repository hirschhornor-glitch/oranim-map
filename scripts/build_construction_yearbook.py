# -*- coding: utf-8 -*-
"""
build_construction_yearbook.py
================================
Builds data/construction_yearbook.geojson — building STARTS & COMPLETIONS
(number of dwellings) per sub-quarter (תת-רובע), 2022-2025, dissolved onto
sub-quarter polygons derived from data/stat_areas.geojson.

Source: Jerusalem Statistical Yearbook 2026 (מכון ירושלים), table ט/7
        ("התחלות וגמר בנייה בירושלים, לפי תת-רובע ומספר דירות").
        Open API: POST admin-ajax.php?action=get_yearbook&yearbook_id=4373
        Per-table XLSX ("נתונים לעיבוד"): shnaton_I0726.xlsx

JOIN KEY: statistical-area 4-digit code -> sub-quarter = first 3 digits.
          e.g. stat_area_id 1011 -> sub-quarter 101 (בית הכרם).

Re-run after refreshing the XLSX or stat_areas.geojson.
Requires: openpyxl, shapely.  (py = python 3.14 on this machine)
"""
import json, re, sys, os
from openpyxl import load_workbook
from shapely.geometry import shape, mapping
from shapely.ops import unary_union

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, "data")

# XLSX sources — keep alongside script under _yearbook_src/
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    HERE, "_yearbook_src", "shnaton_I0726.xlsx")          # ט/7 construction
XLSX_T16 = os.path.join(HERE, "_yearbook_src", "shnaton_I1626.xlsx")  # ט/16 dwellings + rental %
STAT = os.path.join(DATA, "stat_areas.geojson")
OUT  = os.path.join(DATA, "construction_yearbook.geojson")

YEARS = [2022, 2023, 2024, 2025]
DW_YEARS = [2023, 2024, 2025]   # table ט/16 covers 2023-2025


def num(v):
    """Yearbook cell -> int dwellings. '-' / blank -> 0."""
    if v is None:
        return 0
    s = str(v).strip()
    if s in ("", "-", "–", "—"):
        return 0
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def parse_construction(path):
    """Parse table ט/7 ABSOLUTE-numbers block -> {subq_code: {name, years:{}}}.

    Layout (per row): colA may hold 'תת-רובע'/'רובע'/'', colB = 3-digit
    sub-quarter code, colC = Hebrew name, colsD..K = 8 values
    (start/complete x 4 years). The sheet repeats every code in a later
    'אחוזים' (percent) block; we stop at that header and keep first hits.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    out = {}
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        joined = " ".join(cells)
        # stop once we reach the percent section
        if "אחוזים" in joined or "Percent" in joined:
            break
        if len(cells) < 11:
            continue
        code = cells[1].strip()
        if not re.fullmatch(r"\d{3}", code):
            continue
        if code in out:           # keep first (absolute) occurrence only
            continue
        name = cells[2].strip()
        vals = [num(c) for c in cells[3:11]]   # 8 values
        years = {}
        for i, y in enumerate(YEARS):
            years[y] = {"started": vals[i * 2], "completed": vals[i * 2 + 1]}
        out[code] = {"name_he": name, "years": years}
    return out


def fnum(v):
    """Yearbook cell -> float (rounded). '-' / blank -> None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_dwellings(path):
    """Parse table ט/16 -> {subq_code: {dw:{year:int}, rent:{year:float%}}}.

    Layout per row: colB = 3-digit sub-quarter code, colC = name,
    colsD..F = dwelling units 2023/2024/2025, colsG..I = % rented 2023/2024/2025.
    First occurrence per code wins.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    out = {}
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if len(cells) < 9:
            continue
        code = cells[1].strip()
        if not re.fullmatch(r"\d{3}", code) or code in out:
            continue
        dw = {y: num(cells[3 + i]) for i, y in enumerate(DW_YEARS)}
        rent = {y: fnum(cells[6 + i]) for i, y in enumerate(DW_YEARS)}
        out[code] = {"dw": dw, "rent": rent}
    return out


def main():
    if not os.path.exists(XLSX):
        sys.exit("XLSX not found: %s\n(pass path as argv[1])" % XLSX)
    constr = parse_construction(XLSX)
    print("parsed %d sub-quarters from %s" % (len(constr), os.path.basename(XLSX)))
    dwell = parse_dwellings(XLSX_T16) if os.path.exists(XLSX_T16) else {}
    print("parsed %d sub-quarters from %s (dwellings/rental)"
          % (len(dwell), os.path.basename(XLSX_T16)))

    stat = json.load(open(STAT, encoding="utf-8"))
    # group stat-area features by derived sub-quarter code
    groups = {}
    for f in stat["features"]:
        sid = str(f["properties"]["stat_area_id"]).zfill(4)
        sq = sid[:3]
        groups.setdefault(sq, []).append(f)

    feats = []
    covered = 0
    for sq in sorted(groups):
        members = groups[sq]
        # dissolve geometry (unary_union handles MultiPolygon ring fill correctly)
        geoms = []
        pop = hh = 0.0
        for f in members:
            try:
                geoms.append(shape(f["geometry"]).buffer(0))
            except Exception:
                pass
            p = f["properties"]
            if isinstance(p.get("pop_approx"), (int, float)):
                pop += p["pop_approx"]
            if isinstance(p.get("hh_total"), (int, float)):
                hh += p["hh_total"]
        if not geoms:
            continue
        dissolved = unary_union(geoms)

        c = constr.get(sq)
        props = {
            "subq": sq,
            "name_he": (c["name_he"] if c else ""),
            "stat_area_ids": sorted(str(m["properties"]["stat_area_id"]) for m in members),
            "n_stat_areas": len(members),
            "pop_approx": round(pop) if pop else None,
            "hh_total": round(hh) if hh else None,
            "has_data": bool(c),
        }
        tot_s = tot_c = 0
        if c:
            covered += 1
            for y in YEARS:
                props["started_%d" % y] = c["years"][y]["started"]
                props["completed_%d" % y] = c["years"][y]["completed"]
                tot_s += c["years"][y]["started"]
                tot_c += c["years"][y]["completed"]
        props["started_total"] = tot_s
        props["completed_total"] = tot_c
        props["started_per_1000"] = round(tot_s / pop * 1000, 1) if pop else None

        # dwelling stock + rental share (table ט/16)
        dw = dwell.get(sq)
        props["has_dwellings"] = bool(dw)
        if dw:
            for y in DW_YEARS:
                props["dwellings_%d" % y] = dw["dw"].get(y)
                rv = dw["rent"].get(y)
                props["rented_pct_%d" % y] = round(rv, 1) if rv is not None else None
            latest = dw["dw"].get(DW_YEARS[-1])
            # construction starts 2022-25 as a share of the latest dwelling stock
            props["started_pct_of_stock"] = round(tot_s / latest * 100, 1) if latest else None

        feats.append({"type": "Feature", "properties": props,
                      "geometry": mapping(dissolved)})

    fc = {
        "type": "FeatureCollection",
        "metadata": {
            "title": "בנייה, מלאי דירות ושכירות לפי תת-רובע",
            "source": "השנתון הסטטיסטי לירושלים 2026, מכון ירושלים למחקרי מדיניות — לוחות ט/7 (בנייה) + ט/16 (מלאי דירות ושכירות)",
            "source_url": "https://jerusaleminstitute.org.il/yearbook/#/4373/31918",
            "metric": "מספר דירות (יח\"ד)",
            "years": YEARS,
            "dwelling_years": DW_YEARS,
            "join": "stat_area_id[:3] == sub-quarter code",
        },
        "features": feats,
    }
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(fc, fh, ensure_ascii=False)
    dwc = sum(1 for f in feats if f["properties"].get("has_dwellings"))
    print("wrote %d sub-quarter polygons (%d construction, %d dwellings/rental) -> %s"
          % (len(feats), covered, dwc, os.path.relpath(OUT, ROOT)))
    miss = [f["properties"]["subq"] for f in feats if not f["properties"]["has_data"]]
    if miss:
        print("  no construction data for:", miss)


if __name__ == "__main__":
    main()
